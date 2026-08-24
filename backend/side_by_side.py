# -*- coding: utf-8 -*-
from __future__ import annotations

from copy import copy
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any
import re
import unicodedata

from openpyxl import Workbook, load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

TAGS = {"ART", "CH3", "CH4", "CH5", "STOT", "TOTHT", "TVA", "TOTTTC"}
ORANGE = "FA4616"
GREY = "595959"
LIGHT_GREY = "E7E6E6"


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFD", str(value))
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    text = text.lower().replace("²", "2")
    text = re.sub(r"[^a-z0-9%]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


# NUMERIC_SEMANTIC_ALIGNMENT_V1
_NUMERIC_TEXT_RE = re.compile(r"^[+\-]?[0-9OIlS][0-9OIlS .,'’\u00a0]*(?:[%€])?$")
_UNIT_ALIASES={"u":"u","un":"u","unite":"u","m2":"m2","m²":"m2","m3":"m3","m³":"m3","ml":"ml","m l":"ml","ens":"ens","ensemble":"ens","forf":"forfait","forfait":"forfait","kg":"kg"}

def canonical_number(value):
    """Convertit uniquement une chaîne strictement numérique, avec OCR prudent."""
    if value is None or isinstance(value,bool):return None
    if isinstance(value,(int,float)):
        if isinstance(value,float) and (value!=value or abs(value)==float('inf')):return None
        return float(value)
    raw=str(value).strip()
    if not raw or not _NUMERIC_TEXT_RE.fullmatch(raw):return None
    # Les substitutions OCR ne sont admises que si au moins un vrai chiffre existe.
    if not re.search(r"\d",raw):return None
    s=raw.replace('\u00a0','').replace(' ','').replace('€','').replace('%','').replace("'",'').replace('’','')
    s=s.translate(str.maketrans({'O':'0','I':'1','l':'1','S':'5'}))
    if ',' in s and '.' in s:
        if s.rfind(',')>s.rfind('.'):s=s.replace('.','').replace(',','.')
        else:s=s.replace(',','')
    else:s=s.replace(',','.')
    try:return float(s)
    except ValueError:return None

def canonical_unit(value):
    text=clean_text(value).replace('.',' ').strip()
    return _UNIT_ALIASES.get(text.replace(' ',''),_UNIT_ALIASES.get(text,text))

def canonical_designation(value):
    text=clean_text(value)
    text=re.sub(r"\b0+(\d+)\b",r"\1",text)
    text=re.sub(r"\b(dim|dimensions?)\b"," ",text)
    text=re.sub(r"\b(ht|haut(?:eur)?)\b"," ht ",text)
    text=re.sub(r"\s+"," ",text).strip()
    return text

def _row_marker(ws,row,marker_col):
    return str(ws.cell(row,marker_col).value or '').strip().upper() if marker_col else ''

def _numeric_orphan(ws,row,table_end,marker_col):
    """Une ligne n'est décorative que si toutes ses valeurs visibles sont numériques ET nulles.

    Une quantité, un PU ou un montant non nul reste une donnée métier à conserver
    (voir ORPHAN_VALUE_ATTACHMENT_V1) : seule la combinaison zéro partout, sans
    aucun texte, correspond à un zéro de mise en forme (sous-total, chapitre...).
    """
    if _row_marker(ws,row,marker_col) in TAGS:return False
    values=[ws.cell(row,c).value for c in range(1,table_end+1) if ws.cell(row,c).value not in (None,'')]
    if not values:return True
    numbers=[canonical_number(v) for v in values]
    if any(n is None for n in numbers):return False
    return all((n or 0.0)==0.0 for n in numbers)


# ORPHAN_VALUE_ATTACHMENT_V1
def _is_blank_cell(value):
    return value is None or (isinstance(value,str) and not value.strip())

def resolve_orphan_attachments(ws,table_end,marker_col):
    """Rattache une ligne numérique non marquée à l'article ART qui la précède.

    Le rattachement n'est retenu que si toutes les conditions métier sont réunies :
    la ligne suit un ART et précède le prochain marqueur, elle ne contient ni
    référence, ni désignation, ni unité, ses valeurs sont numériques non nulles,
    et les cellules cibles de l'article parent sont vides pour ces colonnes.
    Retourne {ligne_orpheline: {'parent': ligne_art, 'columns': [...]}}.
    """
    attachments={}
    parent_row=None
    for row in range(1,ws.max_row+1):
        marker=_row_marker(ws,row,marker_col)
        if marker=='ART':
            parent_row=row;continue
        if marker in TAGS:
            parent_row=None;continue
        if parent_row is None:continue
        cells=[(c,ws.cell(row,c).value) for c in range(1,table_end+1) if ws.cell(row,c).value not in (None,'')]
        if not cells:continue
        numbers=[(c,canonical_number(v)) for c,v in cells]
        if any(n is None for _,n in numbers):continue
        nonzero=[(c,n) for c,n in numbers if (n or 0.0)!=0.0]
        if not nonzero:continue
        if not all(_is_blank_cell(ws.cell(parent_row,c).value) for c,_ in nonzero):continue
        attachments[row]={'parent':parent_row,'columns':[c for c,_ in nonzero]}
    return attachments

def parent_attachment_map(attachments):
    """Inverse {orpheline: parent} en {parent: orpheline la plus proche}."""
    reverse={}
    for orphan_row,info in sorted(attachments.items()):
        reverse.setdefault(info['parent'],orphan_row)
    return reverse

def _column_role(ws,col):
    sample=' '.join(clean_text(ws.cell(r,col).value) for r in range(1,min(ws.max_row,20)+1) if ws.cell(r,col).value not in (None,''))
    if 'quantite' in sample or 'qte' in sample:return 'quantity'
    if 'p u' in sample or 'prix unitaire' in sample:return 'unit_price'
    if 'montant' in sample or 'total' in sample:return 'amount'
    if 'tva' in sample:return 'vat'
    return ''

def normalize_business_value(ws,row,col,value):
    n=canonical_number(value)
    if n is None:return value,None
    role=_column_role(ws,col)
    digits=3 if role=='quantity' else 2 if role in {'unit_price','amount','vat'} else 6
    n=round(n,digits)
    if abs(n-round(n))<10**(-digits):n=int(round(n))
    fmt='#,##0.###' if role=='quantity' else '#,##0.00' if role in {'unit_price','amount','vat'} else '#,##0.######'
    return n,fmt

def technical_columns(ws):
    marker_col = id_col = None
    start = max(1, ws.max_column - 15)
    for row in range(1, ws.max_row + 1):
        for col in range(start, ws.max_column + 1):
            value = str(ws.cell(row, col).value or "").strip().upper()
            if value in TAGS:
                marker_col = col
                id_col = col + 1 if col < ws.max_column else None
                return marker_col, id_col
    return None, None


# STABILITY_FIX_AFTER_RESTORE_V1
def visible_table_end(ws, marker_col):
    """Retourne la fin du bloc métier visible, sans recopier les îlots techniques.

    Certains DPGF conservent ART/CHx et les identifiants plusieurs centaines de
    colonnes à droite. Ces balises restent lisibles via marker_col, mais les
    colonnes vides intermédiaires ne font pas partie du tableau à reproduire.
    Les colonnes TVA/TTC et multi-bâtiments restent conservées dès lors qu'elles
    appartiennent au bloc principal continu.
    """
    limit = (marker_col - 1) if marker_col else ws.max_column
    if limit <= 1:
        return max(1, limit)

    header_tokens = (
        "designation", "libelle", "description", "unite", "quantite", "qte",
        "prix unitaire", "p u", "pu", "montant", "total", "tva", "ttc",
        "batiment", "parking", "tranche",
    )
    populated = []
    header_limit = min(ws.max_row, 25)
    for col in range(1, limit + 1):
        count = 0
        header_match = False
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row, col).value
            if value not in (None, ""):
                count += 1
                if row <= header_limit:
                    normalized = clean_text(value)
                    if any(token in normalized for token in header_tokens):
                        header_match = True
        if count >= 2 or header_match:
            populated.append(col)

    if not populated:
        return 1

    # Le tableau métier commence dans le premier îlot de colonnes. Une rupture
    # de plus de 8 colonnes entièrement non significatives ferme ce bloc.
    last = populated[0]
    for col in populated[1:]:
        if col - last > 8:
            break
        last = col
    return max(1, last)

def meaningful_rows(ws, table_end, marker_col):
    attached=resolve_orphan_attachments(ws,table_end,marker_col)
    rows=[]
    for row in range(1,ws.max_row+1):
        if row in attached:continue
        marker=_row_marker(ws,row,marker_col)
        visible=any(ws.cell(row,col).value is not None for col in range(1,table_end+1))
        if (visible or marker in TAGS) and not _numeric_orphan(ws,row,table_end,marker_col):rows.append(row)
    return rows

def row_descriptor(ws,row,table_end,marker_col,id_col):
    marker=_row_marker(ws,row,marker_col)
    article_id=clean_text(ws.cell(row,id_col).value) if id_col else ''
    values=[ws.cell(row,col).value for col in range(1,table_end+1)]
    texts=[clean_text(v) for v in values if v not in (None,'')]
    visible=' | '.join(texts);reference=clean_text(values[0]) if values else ''
    designation=canonical_designation(values[1]) if len(values)>1 else canonical_designation(visible)
    unit=canonical_unit(values[2]) if len(values)>2 else ''
    if marker=='ART' and article_id:key=f'ART:{article_id}'
    elif marker:key=f'{marker}:{reference}:{designation}'
    elif 'designation' in visible and ('quantite' in visible or 'montant' in visible or 'p u' in visible):key='HEADER'
    else:key=f'TEXT:{reference}:{designation}'
    return {'row':row,'marker':marker,'id':article_id,'reference':reference,'designation':designation,'unit':unit,'visible':visible,'key':key}

def similarity(left,right):
    dsim=SequenceMatcher(None,left['designation'],right['designation']).ratio()
    same_unit=not left.get('unit') or not right.get('unit') or left.get('unit')==right.get('unit')
    if left['marker']==right['marker']=='ART' and left['id'] and left['id']==right['id']:
        # Un ART partagé mais réutilisé sur des postes différents (doublon de saisie)
        # ne doit pas verrouiller un mauvais appariement : la désignation tranche.
        # Ce test précède la clé générique car key=f'ART:{id}' ignore la désignation.
        return 40 if dsim>=0.5 else 15
    if left['key']==right['key']:return 40
    if left['reference'] and left['reference']==right['reference']:
        return 32 if dsim>=0.72 and same_unit else 18
    if left['marker'] and left['marker']==right['marker']:
        return 26 if dsim>=0.90 and same_unit else 10+int(dsim*8)
    if dsim>=0.94 and same_unit:return 24
    if dsim>=0.86 and same_unit:return 14
    ratio=SequenceMatcher(None,left['visible'],right['visible']).ratio()
    if ratio>=0.90:return 8
    if ratio>=0.80:return 3
    return -9


# BEGIN_HEADER_ANCHOR_ALIGNMENT_V1
def _nw_align_sequences(left, right):
    """Algorithme historique, inchangé, appliqué à une sous-séquence."""
    n, m = len(left), len(right)
    gap = -3
    score = [[0] * (m + 1) for _ in range(n + 1)]
    trace = [[None] * (m + 1) for _ in range(n + 1)]
    for i in range(1, n + 1):
        score[i][0], trace[i][0] = i * gap, "U"
    for j in range(1, m + 1):
        score[0][j], trace[0][j] = j * gap, "L"
    for i in range(1, n + 1):
        for j in range(1, m + 1):
            diag = score[i-1][j-1] + similarity(left[i-1], right[j-1])
            up = score[i-1][j] + gap
            west = score[i][j-1] + gap
            best = max(diag, up, west)
            score[i][j] = best
            trace[i][j] = "D" if best == diag else ("U" if best == up else "L")
    aligned = []
    i, j = n, m
    while i or j:
        direction = trace[i][j]
        if direction == "D":
            aligned.append((left[i-1], right[j-1])); i -= 1; j -= 1
        elif direction == "U":
            aligned.append((left[i-1], None)); i -= 1
        else:
            aligned.append((None, right[j-1])); j -= 1
    return list(reversed(aligned))


def _unique_header_index(sequence):
    """Retourne l'index de la première ancre métier fiable, sinon None."""
    indexes = [i for i, item in enumerate(sequence) if item and item.get("key") == "HEADER"]
    if not indexes:
        return None
    # Plusieurs lignes HEADER consécutives représentent un même en-tête multiniveau.
    first = indexes[0]
    if any(i > first + 2 for i in indexes[1:]):
        return None
    return first


def _bottom_align_prefix(left, right):
    """Conserve les lignes d'introduction mais aligne leur fin sur l'en-tête."""
    size = max(len(left), len(right))
    lpad = [None] * (size - len(left)) + list(left)
    rpad = [None] * (size - len(right)) + list(right)
    return list(zip(lpad, rpad))


# ART_ANCHOR_SEGMENTATION_V1
def _anchor_key(item):
    """Clé d'ancrage forte : identifiant ART, sinon référence non vide."""
    if item.get("marker") == "ART" and item.get("id"):
        return ("ART", item["id"])
    if item.get("reference"):
        return ("REF", item.get("marker") or "", item["reference"])
    return None


def _anchor_indexes(left, right):
    """Ancres communes, uniques des deux côtés et dans le même ordre relatif.

    Un identifiant ART ou une référence qui se répète sur un côté ne peut pas
    servir d'ancre (voir le cas M03-A058 dupliqué dans l'ACT du lot 2 22204) :
    seule une clé strictement unique à gauche ET à droite verrouille l'appariement.
    """
    from collections import Counter
    lkeys = [_anchor_key(item) for item in left]
    rkeys = [_anchor_key(item) for item in right]
    lcount = Counter(k for k in lkeys if k)
    rcount = Counter(k for k in rkeys if k)
    lpos = {k: i for i, k in enumerate(lkeys) if k and lcount[k] == 1}
    anchors = []
    last_i = last_j = -1
    for j, k in enumerate(rkeys):
        if not k or rcount[k] != 1 or k not in lpos:
            continue
        i = lpos[k]
        if i > last_i and j > last_j:
            anchors.append((i, j))
            last_i, last_j = i, j
    return anchors


def _segmented_align(left, right):
    """Découpe en intervalles bornés par les ancres ART/référence puis aligne localement.

    Une ligne parasite avant une ancre ne peut plus décaler l'ancre suivante :
    l'algorithme approximatif ne travaille que sur les lignes situées strictement
    entre deux ancres verrouillées (ou avant la première / après la dernière).
    """
    anchors = _anchor_indexes(left, right)
    if not anchors:
        return _nw_align_sequences(left, right)
    aligned = []
    li = ri = 0
    for i, j in anchors:
        aligned.extend(_nw_align_sequences(left[li:i], right[ri:j]))
        aligned.append((left[i], right[j]))
        li, ri = i + 1, j + 1
    aligned.extend(_nw_align_sequences(left[li:], right[ri:]))
    return aligned


def align_sequences(left, right):
    """Aligne d'abord les en-têtes métier, puis segmente par ancres ART/référence."""
    li = _unique_header_index(left)
    ri = _unique_header_index(right)
    if li is None or ri is None:
        return _segmented_align(left, right)
    prefix = _bottom_align_prefix(left[:li], right[:ri])
    # Les deux lignes HEADER sont verrouillées ensemble ; le reste garde l'algo historique.
    anchor = [(left[li], right[ri])]
    tail = _segmented_align(left[li + 1:], right[ri + 1:])
    return prefix + anchor + tail
# END_HEADER_ANCHOR_ALIGNMENT_V1

def business_sheet(ws):
    marker_col, _ = technical_columns(ws)
    end = visible_table_end(ws, marker_col)
    text = " ".join(clean_text(ws.cell(r, c).value) for r in range(1, min(ws.max_row, 25)+1)
                    for c in range(1, min(end, 30)+1) if ws.cell(r, c).value is not None)
    return bool(marker_col or ("designation" in text and ("quantite" in text or "montant" in text)))


def scope_type(ws):
    sample = clean_text(ws.title) + " " + " ".join(
        clean_text(ws.cell(r, c).value) for r in range(1, min(ws.max_row, 20)+1)
        for c in range(1, min(ws.max_column, 20)+1) if ws.cell(r, c).value is not None)
    if re.search(r"\b(pse|option|variante)\b", sample):
        return "OPTION"
    return "BASE"


def pair_sheets(dce_wb, act_wb):
    dce = [ws for ws in dce_wb.worksheets if business_sheet(ws)]
    act = [ws for ws in act_wb.worksheets if business_sheet(ws)]
    unused = set(range(len(act)))
    pairs = []
    for dws in dce:
        scored = []
        dmark, did = technical_columns(dws)
        dend = visible_table_end(dws, dmark)
        drows = meaningful_rows(dws, dend, dmark)
        dids = {clean_text(dws.cell(r, did).value) for r in drows if did and str(dws.cell(r, dmark).value or "").upper() == "ART"}
        for index in unused:
            aws = act[index]
            if scope_type(aws) != scope_type(dws):
                continue
            amark, aid = technical_columns(aws)
            aend = visible_table_end(aws, amark)
            arows = meaningful_rows(aws, aend, amark)
            aids = {clean_text(aws.cell(r, aid).value) for r in arows if aid and str(aws.cell(r, amark).value or "").upper() == "ART"}
            overlap = len(dids & aids)
            name = SequenceMatcher(None, clean_text(dws.title), clean_text(aws.title)).ratio()
            scored.append((overlap * 100 + name, index))
        if scored:
            _, best = max(scored)
            unused.remove(best)
            pairs.append((dws, act[best]))
        else:
            pairs.append((dws, None))
    for index in sorted(unused):
        pairs.append((None, act[index]))
    return pairs


def copy_cell(source, target, col_offset=0):
    target.value = source.value
    if isinstance(source.value, str) and source.value.startswith("=") and col_offset:
        try:
            target.value = Translator(source.value, origin=source.coordinate).translate_formula(target.coordinate)
        except Exception:
            target.value = source.value
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.font:
        target.font = copy(source.font)
    if source.fill:
        target.fill = copy(source.fill)
    if source.border:
        target.border = copy(source.border)
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.protection:
        target.protection = copy(source.protection)


def copy_block_row(src_ws, src_row, dst_ws, dst_row, start_col, end_col, orphan_row=None):
    """Copie une ligne métier ; complète les cellules vides avec une ligne orpheline rattachée."""
    for src_col in range(1, end_col + 1):
        source=src_ws.cell(src_row,src_col)
        cell=source
        if orphan_row and _is_blank_cell(source.value):
            orphan_cell=src_ws.cell(orphan_row,src_col)
            if not _is_blank_cell(orphan_cell.value):
                cell=orphan_cell
        target=dst_ws.cell(dst_row,start_col+src_col-1)
        copy_cell(cell,target,start_col-1)
        normalized,fmt=normalize_business_value(src_ws,cell.row,src_col,cell.value)
        if fmt and not (isinstance(cell.value,str) and cell.value.startswith('=')):
            target.value=normalized;target.number_format=fmt


def copy_dimensions(src_ws, dst_ws, start_col, end_col):
    for col in range(1, end_col + 1):
        src_letter = get_column_letter(col)
        dst_letter = get_column_letter(start_col + col - 1)
        dim = src_ws.column_dimensions[src_letter]
        dst_ws.column_dimensions[dst_letter].width = dim.width
        dst_ws.column_dimensions[dst_letter].hidden = dim.hidden


def copy_merges(src_ws, dst_ws, row_map, start_col, end_col):
    for merged in src_ws.merged_cells.ranges:
        if merged.max_col > end_col:
            continue
        mapped_rows = [row_map.get(r) for r in range(merged.min_row, merged.max_row + 1)]
        if not mapped_rows or any(r is None for r in mapped_rows):
            continue
        if mapped_rows != list(range(mapped_rows[0], mapped_rows[0] + len(mapped_rows))):
            continue
        dst_ws.merge_cells(start_row=mapped_rows[0], end_row=mapped_rows[-1],
                           start_column=start_col + merged.min_col - 1,
                           end_column=start_col + merged.max_col - 1)



# ---------- Alertes cellules manquantes V5 ----------
def _normalized_header(value):
    return clean_text(value).replace("p u", "pu")


def _detect_business_columns(ws, table_end):
    """Détecte les colonnes utiles sans considérer les zéros comme vides."""
    header_limit = min(ws.max_row, 12)
    found = {"designation": None, "unit": None, "quantity": [], "unit_price": [], "amount": [], "vat": []}
    for col in range(1, table_end + 1):
        values = [_normalized_header(ws.cell(row, col).value) for row in range(1, header_limit + 1)]
        joined = " ".join(value for value in values if value)
        last = next((value for value in reversed(values) if value), "")
        if found["designation"] is None and "designation" in joined:
            found["designation"] = col
        if found["unit"] is None and (last in {"u", "unite"} or " unite " in f" {joined} "):
            found["unit"] = col
        if any(token in joined for token in ("quantite", "qte", "qty")):
            found["quantity"].append(col)
        if last in {"pu", "p.u", "prix unitaire"} or "prix unitaire" in joined:
            found["unit_price"].append(col)
        if "montant" in joined and "tva" not in joined:
            found["amount"].append(col)
        if "tva" in joined:
            found["vat"].append(col)
    return found


def _is_blank(value):
    return value is None or (isinstance(value, str) and not value.strip())


def _row_alerts(ws, row, descriptor, columns, role):
    """N'alerte que les lignes ARTICLE. Les cellules à 0 sont valides."""
    if descriptor is None or descriptor.get("marker") != "ART":
        return []
    alerts = []
    designation_col = columns.get("designation")
    unit_col = columns.get("unit")
    if designation_col and _is_blank(ws.cell(row, designation_col).value):
        alerts.append(f"{role}: désignation absente")
    if unit_col and _is_blank(ws.cell(row, unit_col).value):
        alerts.append(f"{role}: unité absente")

    # La quantité entreprise est facultative par règle métier. Elle est signalée
    # comme information, pas comme anomalie bloquante.
    quantity_cols = columns.get("quantity") or []
    if role == "ACT" and quantity_cols and all(_is_blank(ws.cell(row, col).value) for col in quantity_cols):
        alerts.append("INFO ACT: quantité entreprise absente")

    pu_cols = columns.get("unit_price") or []
    amount_cols = columns.get("amount") or []
    if pu_cols and all(_is_blank(ws.cell(row, col).value) for col in pu_cols):
        alerts.append(f"{role}: prix unitaire absent")
    if amount_cols and all(_is_blank(ws.cell(row, col).value) for col in amount_cols):
        alerts.append(f"{role}: montant absent")

    # Si une colonne TVA existe réellement dans le tableau, sa valeur devient attendue.
    vat_cols = columns.get("vat") or []
    if vat_cols and all(_is_blank(ws.cell(row, col).value) for col in vat_cols):
        alerts.append(f"{role}: TVA article absente")
    return alerts


def _write_alert_summary(workbook, alerts):
    if not alerts:
        return
    title = "Alertes"
    if title in workbook.sheetnames:
        del workbook[title]
    ws = workbook.create_sheet(title, 0)
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Légende des codes couleurs"
    ws["A1"].font = Font(size=13, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=GREY)
    ws.merge_cells("A1:G1")
    alert_categories = set()
    if any(str(item.get("message", "")).startswith("INFO") for item in alerts):
        alert_categories.add("INFO")
    if any(not str(item.get("message", "")).startswith("INFO") for item in alerts):
        alert_categories.add("ERROR")
    _add_color_legend(ws, alert_categories, start_row=2, start_col=1, horizontal=False)
    visible_legend_rows = 4 if alert_categories == {"ERROR", "INFO"} else 2
    for hidden_row in range(2 + visible_legend_rows, 6):
        ws.row_dimensions[hidden_row].hidden = True
    headers = ["Feuille résultat", "Ligne", "Périmètre", "Côté", "Référence", "Désignation", "Alerte"]
    ws.append([])
    ws.append(headers)
    header_row = 7
    for cell in ws[header_row]:
        cell.fill = PatternFill("solid", fgColor=ORANGE)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for alert in alerts:
        ws.append([
            alert["sheet"], alert["row"], alert["scope"], alert["side"],
            alert["reference"], alert["designation"], alert["message"],
        ])
    widths = [22, 10, 14, 10, 16, 55, 42]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A8"
    ws.auto_filter.ref = f"A7:G{ws.max_row}"
    for row in range(8, ws.max_row + 1):
        ws.cell(row, 6).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, 7).alignment = Alignment(wrap_text=True, vertical="top")
        if str(ws.cell(row, 7).value or "").startswith("INFO"):
            ws.cell(row, 7).fill = PatternFill("solid", fgColor="FFF2CC")
        else:
            ws.cell(row, 7).fill = PatternFill("solid", fgColor="F4CCCC")


# ---------- Surlignage rouge cellules manquantes V6 ----------
def _missing_source_columns(ws, row, descriptor, columns, role):
    """Retourne les colonnes source réellement anormales pour une ligne ARTICLE."""
    if descriptor is None or descriptor.get("marker") != "ART":
        return []
    missing = []
    designation_col = columns.get("designation")
    unit_col = columns.get("unit")
    if designation_col and _is_blank(ws.cell(row, designation_col).value):
        missing.append((designation_col, f"{role}: désignation absente", "ERROR"))
    if unit_col and _is_blank(ws.cell(row, unit_col).value):
        missing.append((unit_col, f"{role}: unité absente", "ERROR"))

    quantity_cols = columns.get("quantity") or []
    if role == "ACT" and quantity_cols and all(_is_blank(ws.cell(row, col).value) for col in quantity_cols):
        # Quantité entreprise facultative : orange, pas rouge.
        for col in quantity_cols:
            missing.append((col, "INFO ACT: quantité entreprise absente", "INFO"))

    pu_cols = columns.get("unit_price") or []
    if pu_cols and all(_is_blank(ws.cell(row, col).value) for col in pu_cols):
        for col in pu_cols:
            missing.append((col, f"{role}: prix unitaire absent", "ERROR"))

    amount_cols = columns.get("amount") or []
    if amount_cols and all(_is_blank(ws.cell(row, col).value) for col in amount_cols):
        for col in amount_cols:
            missing.append((col, f"{role}: montant absent", "ERROR"))

    vat_cols = columns.get("vat") or []
    if vat_cols and all(_is_blank(ws.cell(row, col).value) for col in vat_cols):
        for col in vat_cols:
            missing.append((col, f"{role}: TVA article absente", "ERROR"))
    return missing


def _mark_missing_cells(dst_ws, dst_row, source_missing, dst_start_col):
    """Applique rouge aux erreurs, orange aux informations, avec commentaire."""
    for source_col, message, severity in source_missing:
        dst_col = dst_start_col + source_col - 1
        cell = dst_ws.cell(dst_row, dst_col)
        cell.fill = PatternFill(
            "solid",
            fgColor="F4CCCC" if severity == "ERROR" else "FCE5CD",
        )
        cell.border = Border(
            left=Side(style="thin", color="C00000" if severity == "ERROR" else "BF9000"),
            right=Side(style="thin", color="C00000" if severity == "ERROR" else "BF9000"),
            top=Side(style="thin", color="C00000" if severity == "ERROR" else "BF9000"),
            bottom=Side(style="thin", color="C00000" if severity == "ERROR" else "BF9000"),
        )
        existing = cell.comment.text + "\n" if cell.comment else ""
        cell.comment = __import__("openpyxl").comments.Comment(existing + message, "AnalyseAO")


def _mark_missing_counterpart(dst_ws, dst_row, start_col, end_col, message):
    """Marque en rouge la zone vide si un article n'existe que d'un côté."""
    if end_col < start_col:
        return
    for col in range(start_col, end_col + 1):
        cell = dst_ws.cell(dst_row, col)
        cell.fill = PatternFill("solid", fgColor="F4CCCC")
    anchor = dst_ws.cell(dst_row, start_col)
    anchor.value = "ARTICLE MANQUANT"
    anchor.font = Font(color="9C0006", bold=True)
    anchor.comment = __import__("openpyxl").comments.Comment(message, "AnalyseAO")


# ---------- Légende codes couleurs V7 ----------
# ---------- Légende conditionnelle V8 ----------
def _add_color_legend(ws, categories, start_row=1, start_col=1, horizontal=True):
    """Ajoute uniquement les codes réellement présents dans la feuille."""
    entries = [
        ("ERROR", "F4CCCC", "Rouge", "Anomalie : donnée attendue absente ou article correspondant manquant"),
        ("INFO", "FCE5CD", "Orange", "Information : quantité entreprise facultative absente"),
        ("ERROR", "C00000", "!", "Au moins une anomalie détectée sur la ligne"),
        ("INFO", "BF9000", "i", "Information non bloquante détectée sur la ligne"),
    ]
    entries = [entry[1:] for entry in entries if entry[0] in categories]
    if horizontal:
        col = start_col
        for color, code, description in entries:
            code_cell = ws.cell(start_row, col, code)
            code_cell.fill = PatternFill("solid", fgColor=color)
            code_cell.font = Font(
                bold=True,
                color="FFFFFF" if color in {"C00000", "BF9000"} else "000000",
            )
            code_cell.alignment = Alignment(horizontal="center", vertical="center")
            text_cell = ws.cell(start_row, col + 1, description)
            text_cell.alignment = Alignment(wrap_text=True, vertical="center")
            ws.column_dimensions[get_column_letter(col)].width = 10
            ws.column_dimensions[get_column_letter(col + 1)].width = 42
            col += 2
    else:
        for offset, (color, code, description) in enumerate(entries):
            row = start_row + offset
            code_cell = ws.cell(row, start_col, code)
            code_cell.fill = PatternFill("solid", fgColor=color)
            code_cell.font = Font(
                bold=True,
                color="FFFFFF" if color in {"C00000", "BF9000"} else "000000",
            )
            code_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row, start_col + 1, description).alignment = Alignment(wrap_text=True, vertical="center")

def build_side_by_side(dce_path, act_path, enterprise, lot, output_path):
    # data_only=True récupère les valeurs mises en cache par Excel et non les formules.
    dce_wb = load_workbook(dce_path, data_only=True)
    act_wb = load_workbook(act_path, data_only=True)
    out = Workbook()
    out.remove(out.active)
    used_titles = set()
    all_alerts = []

    for pair_index, (dws, aws) in enumerate(pair_sheets(dce_wb, act_wb), 1):
        scope = scope_type(dws or aws)
        base_title = f"{scope} {pair_index}"
        title = base_title[:31]
        suffix = 2
        while title in used_titles:
            title = f"{base_title[:27]} {suffix}"; suffix += 1
        used_titles.add(title)
        ws = out.create_sheet(title)
        ws.sheet_view.showGridLines = False
        sheet_alert_start = len(all_alerts)

        dmark, did = technical_columns(dws) if dws else (None, None)
        amark, aid = technical_columns(aws) if aws else (None, None)
        dend = visible_table_end(dws, dmark) if dws else 0
        aend = visible_table_end(aws, amark) if aws else 0
        drows = meaningful_rows(dws, dend, dmark) if dws else []
        arows = meaningful_rows(aws, aend, amark) if aws else []
        ddesc = [row_descriptor(dws, r, dend, dmark, did) for r in drows]
        adesc = [row_descriptor(aws, r, aend, amark, aid) for r in arows]
        dcols = _detect_business_columns(dws, dend) if dws else {}
        acols = _detect_business_columns(aws, aend) if aws else {}
        dattach = parent_attachment_map(resolve_orphan_attachments(dws, dend, dmark)) if dws else {}
        aattach = parent_attachment_map(resolve_orphan_attachments(aws, aend, amark)) if aws else {}
        alignment = align_sequences(ddesc, adesc)

        sep_width = 2
        act_start = dend + sep_width + 1
        total_end = act_start + aend - 1
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(1, dend))
        ws.cell(2, 1, f"DCE - {dws.title if dws else 'Aucune feuille correspondante'}")
        if aend:
            ws.merge_cells(start_row=2, start_column=act_start, end_row=2, end_column=total_end)
            ws.cell(2, act_start, f"ACT - {enterprise} - {aws.title if aws else 'Aucune feuille correspondante'}")
        for cell in (ws.cell(2, 1), ws.cell(2, act_start) if aend else None):
            if cell:
                cell.fill = PatternFill("solid", fgColor=ORANGE)
                cell.font = Font(color="FFFFFF", bold=True, size=12)
                cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 34
        ws.row_dimensions[2].height = 24

        dmap, amap = {}, {}
        output_row = 3
        for left, right in alignment:
            drow = left["row"] if left else None
            arow = right["row"] if right else None
            if drow:
                copy_block_row(dws, drow, ws, output_row, 1, dend, dattach.get(drow))
                dmap[drow] = output_row
            if arow:
                copy_block_row(aws, arow, ws, output_row, act_start, aend, aattach.get(arow))
                amap[arow] = output_row
            heights = []
            if drow and dws.row_dimensions[drow].height: heights.append(dws.row_dimensions[drow].height)
            if arow and aws.row_dimensions[arow].height: heights.append(aws.row_dimensions[arow].height)
            if heights: ws.row_dimensions[output_row].height = max(heights)

            row_alerts = []
            if left and dws:
                row_alerts.extend(_row_alerts(dws, drow, left, dcols, "DCE"))
            if right and aws:
                row_alerts.extend(_row_alerts(aws, arow, right, acols, "ACT"))
            if left is not None and right is None and left.get("marker") == "ART":
                row_alerts.append("ACT: article correspondant manquant")
            if right is not None and left is None and right.get("marker") == "ART":
                row_alerts.append("DCE: article correspondant manquant")

            # Marquage direct des cellules détectées dans les tableaux copiés.
            if left and dws:
                _mark_missing_cells(
                    ws, output_row,
                    _missing_source_columns(dws, drow, left, dcols, "DCE"),
                    1,
                )
            if right and aws:
                _mark_missing_cells(
                    ws, output_row,
                    _missing_source_columns(aws, arow, right, acols, "ACT"),
                    act_start,
                )
            if left is not None and right is None and left.get("marker") == "ART":
                _mark_missing_counterpart(
                    ws, output_row, act_start, total_end,
                    "Article présent dans le DCE mais absent de l'ACT.",
                )
            if right is not None and left is None and right.get("marker") == "ART":
                _mark_missing_counterpart(
                    ws, output_row, 1, dend,
                    "Article ajouté dans l'ACT sans correspondance DCE.",
                )

            if row_alerts:
                status_col = dend + 1
                status_cell = ws.cell(output_row, status_col)
                status_cell.value = "!" if any(not item.startswith("INFO") for item in row_alerts) else "i"
                status_cell.font = Font(bold=True, color="FFFFFF")
                status_cell.fill = PatternFill(
                    "solid",
                    fgColor="C00000" if status_cell.value == "!" else "BF9000",
                )
                status_cell.alignment = Alignment(horizontal="center", vertical="center")
                status_cell.comment = __import__("openpyxl").comments.Comment("\n".join(row_alerts), "AnalyseAO")
                source = left or right or {}
                for message in row_alerts:
                    all_alerts.append({
                        "sheet": title,
                        "row": output_row,
                        "scope": scope,
                        "side": message.split(":", 1)[0],
                        "reference": source.get("reference") or "",
                        "designation": source.get("designation") or "",
                        "message": message,
                    })
            output_row += 1

        if dws:
            copy_dimensions(dws, ws, 1, dend)
            copy_merges(dws, ws, dmap, 1, dend)
        if aws:
            copy_dimensions(aws, ws, act_start, aend)
            copy_merges(aws, ws, amap, act_start, aend)

        # ---------- Légende conditionnelle V8 ----------
        sheet_messages = [item["message"] for item in all_alerts[sheet_alert_start:]]
        sheet_categories = set()
        if any(message.startswith("INFO") for message in sheet_messages):
            sheet_categories.add("INFO")
        if any(not message.startswith("INFO") for message in sheet_messages):
            sheet_categories.add("ERROR")

        if sheet_categories:
            _add_color_legend(ws, sheet_categories, start_row=1, start_col=1, horizontal=True)
        else:
            # ---------- Correction titres sans légende V8.1 ----------
            # openpyxl ne translate pas automatiquement les plages fusionnées
            # lors d'un delete_rows. On retire donc les fusions de titres de la
            # ligne 2, on supprime la ligne vide, puis on recrée les fusions en
            # ligne 1. Les valeurs de titre restent dans leurs cellules d'origine.
            for merged_range in list(ws.merged_cells.ranges):
                if merged_range.min_row == 2 and merged_range.max_row == 2:
                    ws.unmerge_cells(str(merged_range))
            ws.delete_rows(1, 1)
            if dend:
                ws.merge_cells(
                    start_row=1,
                    start_column=1,
                    end_row=1,
                    end_column=max(1, dend),
                )
            if aend:
                ws.merge_cells(
                    start_row=1,
                    start_column=act_start,
                    end_row=1,
                    end_column=total_end,
                )
            ws.row_dimensions[1].height = 24

        # Séparateur visuel fixe entre les deux tableaux.
        for col in range(dend + 1, act_start):
            ws.column_dimensions[get_column_letter(col)].width = 2.5
            for row in range(1, ws.max_row + 1):
                ws.cell(row, col).fill = PatternFill("solid", fgColor=LIGHT_GREY)
        # Ne figer que la ligne de titre : le DCE et l’ACT défilent ensemble horizontalement.
        ws.freeze_panes = "A3" if sheet_categories else "A2"
        ws.auto_filter.ref = None
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.oddHeader.center.text = f"Lot {str(lot).zfill(2)} - {scope} - DCE / ACT côte à côte"

    target = Path(output_path)
    if target.suffix.lower() != ".xlsx": target = target.with_suffix(".xlsx")
    target.parent.mkdir(parents=True, exist_ok=True)
    _write_alert_summary(out, all_alerts)
    out.calculation.fullCalcOnLoad = True
    out.calculation.forceFullCalc = True
    out.calculation.calcMode = "auto"
    out.save(target)
    return target
