# -*- coding: utf-8 -*-
from __future__ import annotations
from dataclasses import dataclass, asdict
from pathlib import Path
import re, unicodedata
from openpyxl import load_workbook

TAGS={"ART","CH3","CH4","CH5","CH6","STOT","TOTHT","TVA","TOTTTC"}
CATEGORY_RULES=[
 ("REMISE", ("remise","rabais","ristourne")),
 ("MOINS-VALUE", ("moins value","moins-value")),
 ("PLUS-VALUE", ("plus value","plus-value")),
 ("COMPLEMENT", ("complement","avenant","travaux supplementaires")),
 ("VARIANTE", ("variante","alternative")),
 ("OPTION", ("option","tranche optionnelle","pse")),
 ("RECAP", ("recap","recapitulatif")),
 ("ANNEXE", ("annexe","precision","precisions")),
]
@dataclass
class SheetEntry:
    sheet:str; lot:str; category:str; score:int; valued:int; article_count:int; total_markers:int; warnings:list[str]
    def dict(self): return asdict(self)

def clean(v):
    s=unicodedata.normalize('NFKD',str(v or '')); s=''.join(c for c in s if not unicodedata.combining(c))
    return re.sub(r'\s+',' ',re.sub(r'[^a-z0-9]+',' ',s.casefold())).strip()
def norm_lot(v):
    m=re.search(r'\d{1,3}',str(v or '')); return m.group(0).zfill(2) if m else ''
def content_lot(ws):
    # Le contenu métier prévaut sur un titre ou un nom de fichier erroné.
    for row in ws.iter_rows(min_row=1,max_row=min(ws.max_row,80),values_only=True):
        for v in row[:40]:
            m=re.fullmatch(r'\s*lot\s*(?:n|no|numero)?\s*[°º]?\s*(\d{1,3})\s*',clean(v))
            if m: return m.group(1).zfill(2)
    return ''
def title_lot(title):
    m=re.search(r'\blot\s*(?:n|no|numero)?\s*[°º]?\s*(\d{1,3})\b',clean(title)); return m.group(1).zfill(2) if m else ''
def filename_lot(path): return title_lot(Path(path).stem)

# STABILITY_FIX_AFTER_RESTORE_V1
def classify(title, sample, is_business):
    """Classe la feuille sans promouvoir une DPGF BASE à cause d'une ligne interne.

    Les catégories métier spéciales doivent être explicites dans le titre de
    feuille. Le contenu n'est utilisé que pour qualifier une feuille non métier.
    """
    title_text = clean(title)
    sample_text = clean(sample)
    if any(value in title_text for value in ("page de garde", "pa garde", "pdg")):
        return "DOCUMENT"
    for category, words in CATEGORY_RULES:
        if any(clean(word) in title_text for word in words):
            return category
    if is_business:
        return "BASE"
    for category, words in CATEGORY_RULES:
        if any(clean(word) in sample_text for word in words):
            return category
    return "DOCUMENT"

def inspect_workbook(path, role='ACT'):
    wb=load_workbook(path,data_only=True,read_only=False)
    result=[]; file_lot=filename_lot(path)
    try:
        for ws in wb.worksheets:
            tags={}; valued=0; headers=set(); samples=[]
            max_scan_col=min(ws.max_column,100)
            for row in ws.iter_rows(min_row=1,max_row=min(ws.max_row,400),max_col=max_scan_col,values_only=True):
                for v in row:
                    if v is None: continue
                    s=str(v).strip(); u=s.upper(); c=clean(s)
                    if len(samples)<180: samples.append(s)
                    if u in TAGS: tags[u]=tags.get(u,0)+1
                    if any(h in c for h in ('designation','quantite','qte ent','prix unitaire','montant ht')): headers.add(c)
                    if isinstance(v,(int,float)) and v not in (0,): valued+=1
            sample=' '.join(samples)
            c_lot=content_lot(ws); t_lot=title_lot(ws.title); lot=c_lot or t_lot or file_lot
            warnings=[]
            declared={x for x in (c_lot,t_lot,file_lot) if x}
            if len(declared)>1: warnings.append('Conflit de numéro de lot corrigé par priorité au contenu métier')
            articles=tags.get('ART',0); totals=tags.get('STOT',0)+tags.get('TOTHT',0)
            is_business=bool(articles or tags.get('CH3') or ('designation' in clean(sample) and 'montant ht' in clean(sample)))
            category=classify(ws.title,sample,is_business)
            score=articles*20+sum(tags.get(t,0) for t in ('CH3','CH4','CH5','CH6'))*3+totals*15+min(valued,500)
            if category=='DOCUMENT': score-=100
            result.append(SheetEntry(ws.title,lot,category,score,valued,articles,totals,warnings))
    finally: wb.close()
    return result

def index_dce_files(dces):
    index={}; warnings=[]
    for original,path in dces:
        for entry in inspect_workbook(path,'DCE'):
            if not entry.lot or entry.category in {'DOCUMENT','RECAP'}: continue
            index.setdefault(entry.lot,[]).append({'original':original,'path':path,'entry':entry})
    for lot,items in index.items(): items.sort(key=lambda x:x['entry'].score,reverse=True)
    return index,warnings

def best_act_lot(path,fallback=''):
    entries=inspect_workbook(path,'ACT')
    candidates=[e for e in entries if e.lot and e.category!='DOCUMENT']
    if not candidates: return norm_lot(fallback),entries
    candidates.sort(key=lambda e:e.score,reverse=True)
    return candidates[0].lot,entries

# ROBUST_WORKBOOK_CATALOG_V21
# Redéfinitions finales volontaires. Les importateurs utilisent ces fonctions
# sans modification de multi_company.py, side_by_side.py ou offer_analysis.py.

_V21_ALIASES = {
    "designation": ("designation", "libelle", "description", "ouvrage", "prestation", "article"),
    "quantity": ("quantite", "qte", "quantite entreprise", "qte ent", "quantite moe"),
    "unit": ("u", "un", "unite"),
    "unit_price": ("p u", "pu", "prix unitaire", "prix unit", "prix u"),
    "amount": ("montant", "montant ht", "montant h t", "montant hors taxes", "prix total", "total ht"),
}
_V21_DOCUMENT_TITLES = (
    "page de garde", "pagedegarde", "sommaire", "notice", "presentation",
    "instruction", "coordonnees", "attestation",
)


def _v21_one_of(text, aliases):
    value = clean(text)
    return any(value == alias or alias in value for alias in aliases)


def _v21_header_roles(ws, max_rows=35, max_cols=120):
    """Détecte les rôles sur des en-têtes répartis sur plusieurs lignes."""
    roles = {key: set() for key in _V21_ALIASES}
    rows = min(ws.max_row, max_rows)
    cols = min(ws.max_column, max_cols)
    for col in range(1, cols + 1):
        fragments = []
        for row in range(1, rows + 1):
            value = ws.cell(row, col).value
            if value not in (None, ""):
                fragments.append(clean(value))
        joined = " ".join(fragments)
        for role, aliases in _V21_ALIASES.items():
            if any(fragment == alias or alias in joined for alias in aliases for fragment in fragments):
                roles[role].add(col)
    return roles


def _v21_scan(ws):
    """Accumule des preuves lexicales, techniques et arithmétiques."""
    max_row = min(ws.max_row, 800)
    max_col = min(ws.max_column, 160)
    tags = {}
    valued = 0
    nonempty = 0
    samples = []
    top_samples = []
    article_like = 0
    arithmetic_matches = 0

    for row in range(1, max_row + 1):
        row_values = []
        numeric = []
        text_count = 0
        for col in range(1, max_col + 1):
            value = ws.cell(row, col).value
            if value in (None, ""):
                continue
            nonempty += 1
            row_values.append(value)
            if row <= 80 and len(top_samples) < 500:
                top_samples.append(str(value))
            if len(samples) < 1200:
                samples.append(str(value))
            upper = str(value).strip().upper()
            if upper in TAGS:
                tags[upper] = tags.get(upper, 0) + 1
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                numeric.append(float(value))
                if value != 0:
                    valued += 1
            elif clean(value):
                text_count += 1

        # Faisceau de preuve structurel : texte + plusieurs nombres.
        if text_count and len(numeric) >= 2:
            article_like += 1
        # Vérifie une relation quantité x PU ~= montant parmi les nombres de la ligne.
        if len(numeric) >= 3:
            found = False
            for i in range(len(numeric) - 2):
                for j in range(i + 1, len(numeric) - 1):
                    expected = numeric[i] * numeric[j]
                    tolerance = max(0.05, abs(expected) * 0.001)
                    if any(abs(expected - candidate) <= tolerance for candidate in numeric[j + 1:]):
                        arithmetic_matches += 1
                        found = True
                        break
                if found:
                    break

    return {
        "tags": tags,
        "valued": valued,
        "nonempty": nonempty,
        "sample": " ".join(samples),
        "top_sample": " ".join(top_samples),
        "article_like": article_like,
        "arithmetic_matches": arithmetic_matches,
    }


def _v21_business_score(ws, scan, roles):
    title = clean(ws.title)
    reasons = []
    score = 0
    tags = scan["tags"]

    if any(token in title for token in _V21_DOCUMENT_TITLES):
        return -100, ["titre documentaire explicite"]

    articles = tags.get("ART", 0)
    subtotals = tags.get("STOT", 0)
    total_ht = tags.get("TOTHT", 0)
    chapters = sum(tags.get(tag, 0) for tag in ("CH3", "CH4", "CH5", "CH6"))

    if articles >= 2:
        score += 18; reasons.append(f"ART répété ({articles})")
    elif articles == 1:
        score += 8; reasons.append("balise ART")
    if total_ht:
        score += 14; reasons.append("balise TOTHT")
    if subtotals:
        score += min(12, 4 + subtotals); reasons.append(f"balise STOT ({subtotals})")
    if chapters:
        score += min(8, chapters); reasons.append(f"balises de chapitres ({chapters})")

    has_designation = bool(roles["designation"])
    has_pu = bool(roles["unit_price"])
    has_amount = bool(roles["amount"])
    has_quantity = bool(roles["quantity"])
    if has_designation and has_pu and has_amount:
        score += 18; reasons.append("en-têtes Désignation + P.U. + Montant")
    else:
        if has_designation: score += 4; reasons.append("en-tête Désignation/Libellé")
        if has_pu: score += 4; reasons.append("en-tête P.U./Prix unitaire")
        if has_amount: score += 4; reasons.append("en-tête Montant/Total")
    if has_quantity:
        score += 3; reasons.append("en-tête Quantité")

    if scan["article_like"] >= 3:
        score += 8; reasons.append(f"lignes d'articles structurées ({scan['article_like']})")
    if scan["arithmetic_matches"] >= 2:
        score += 10; reasons.append(f"cohérences quantité x prix = montant ({scan['arithmetic_matches']})")
    elif scan["arithmetic_matches"] == 1:
        score += 4; reasons.append("cohérence arithmétique détectée")
    if scan["valued"] >= 10:
        score += 4; reasons.append("données chiffrées nombreuses")
    if scan["nonempty"] < 10:
        score -= 20; reasons.append("feuille presque vide")

    return score, reasons


def _v21_category(title, top_sample, is_business):
    title_text = clean(title)
    top_text = clean(top_sample)
    if any(token in title_text for token in _V21_DOCUMENT_TITLES):
        return "DOCUMENT"

    # Catégories spéciales uniquement lorsqu'elles décrivent la feuille entière.
    for cat, words in CATEGORY_RULES:
        if any(clean(word) in title_text for word in words):
            return cat
    if "remplacement" in title_text or "prestation supplementaire eventuelle" in title_text:
        return "OPTION"

    if is_business:
        # PSE explicite dans la zone haute d'une feuille dédiée.
        if re.search(r"(?:^| )pse(?: |$)", top_text) and (
            "remplacement" in top_text or "prestation supplementaire" in top_text
        ):
            return "OPTION"
        return "BASE"

    for cat, words in CATEGORY_RULES:
        if any(clean(word) in top_text for word in words):
            return cat
    return "DOCUMENT"


def inspect_workbook(path, role="ACT"):
    """Catalogue robuste : score explicable, synonymes, balises et structure."""
    wb = load_workbook(path, data_only=True, read_only=False)
    result = []
    file_lot = filename_lot(path)
    try:
        for ws in wb.worksheets:
            scan = _v21_scan(ws)
            roles = _v21_header_roles(ws)
            score, reasons = _v21_business_score(ws, scan, roles)
            is_business = score >= 20

            c_lot = content_lot(ws)
            t_lot = title_lot(ws.title)
            lot = c_lot or t_lot or file_lot
            declared = {value for value in (c_lot, t_lot, file_lot) if value}
            warnings = []
            if len(declared) > 1:
                warnings.append("Conflit de numéro de lot corrigé par priorité au contenu métier")

            category = _v21_category(ws.title, scan["top_sample"], is_business)
            if category == "DOCUMENT":
                score -= 100
            confidence = "confirmée" if score >= 35 else "probable" if score >= 20 else "insuffisante"
            warnings.append(
                f"Détection métier {confidence} (score {score}) : " +
                (", ".join(reasons) if reasons else "aucun signal métier")
            )
            if not lot and is_business:
                warnings.append("Feuille métier reconnue mais numéro de lot non déterminé")

            tags = scan["tags"]
            result.append(SheetEntry(
                ws.title,
                lot,
                category,
                score,
                scan["valued"],
                tags.get("ART", 0),
                tags.get("STOT", 0) + tags.get("TOTHT", 0),
                warnings,
            ))
    finally:
        wb.close()
    return result

# TCE_VIRTUAL_SHEETS_V22
_v22_previous_inspect_workbook = inspect_workbook

def inspect_workbook(path, role="ACT"):
    """Ajoute BPU par titre explicite, sans changer le classement des autres feuilles."""
    entries = _v22_previous_inspect_workbook(path, role)
    for entry in entries:
        title = clean(entry.sheet)
        if re.search(r"(?:^| )bpu(?: |$)", title) and entry.category != "DOCUMENT":
            entry.category = "BPU"
            entry.warnings.append("Catégorie BPU détectée explicitement dans le titre")
    return entries

