# -*- coding: utf-8 -*-
from __future__ import annotations

from dataclasses import asdict, dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
import json
import re
import unicodedata

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

ARTICLE_TAG = "ART"
STRUCTURE_TAGS = {"CH3", "CH4", "CH5", "STOT", "TOTHT", "TVA", "TOTTTC"}
QUANTITY_WORDS = ("quantite", "qte", "qty")
UNIT_PRICE_WORDS = ("p.u", "pu", "prix unitaire")
AMOUNT_WORDS = ("montant", "total")
VAT_WORDS = ("tva", "taux tva", "% tva")


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFD", str(value))
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = text.lower().replace("²", "2")
    text = re.sub(r"[^a-z0-9%]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_reference(value: Any) -> str:
    text = clean_text(value)
    return re.sub(r"\s+", ".", text).strip(".")


def normalize_unit(value: Any) -> str:
    text = clean_text(value).replace(" ", "")
    aliases = {"m²": "m2", "m2": "m2", "ml": "ml", "m.l": "ml", "u": "u", "un": "u", "ens": "ens", "ens.": "ens"}
    return aliases.get(text, text)


# NUMERIC_SEMANTIC_ALIGNMENT_V1
_NUMERIC_TEXT_RE=re.compile(r"^[+\-]?[0-9OIlS][0-9OIlS .,'’\u00a0]*(?:[%€])?$")
def decimal_or_none(value: Any) -> Decimal | None:
    if value in (None,'') or isinstance(value,bool):return None
    try:
        if isinstance(value,str):
            raw=value.strip()
            if not _NUMERIC_TEXT_RE.fullmatch(raw) or not re.search(r"\d",raw):return None
            value=raw.replace(' ','').replace('\u00a0','').replace('€','').replace('%','').replace("'",'').replace('’','')
            value=value.translate(str.maketrans({'O':'0','I':'1','l':'1','S':'5'}))
            if ',' in value and '.' in value:
                value=value.replace('.','').replace(',','.') if value.rfind(',')>value.rfind('.') else value.replace(',','')
            else:value=value.replace(',','.')
        return Decimal(str(value))
    except (InvalidOperation,ValueError,TypeError):return None

def json_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, dict):
        return {key: json_value(item) for key, item in value.items()}
    if isinstance(value, list):
        return [json_value(item) for item in value]
    return value


@dataclass
class ColumnBlock:
    label: str
    quantity_col: int | None = None
    unit_price_col: int | None = None
    amount_col: int | None = None
    vat_col: int | None = None


@dataclass
class SheetProfile:
    name: str
    scope_type: str
    header_rows: list[int]
    designation_col: int | None
    unit_col: int | None
    reference_col: int | None
    marker_col: int | None
    article_id_col: int | None
    blocks: list[ColumnBlock]
    article_rows: list[int]
    warnings: list[str] = field(default_factory=list)


@dataclass
class Article:
    workbook_role: str
    sheet_name: str
    scope_type: str
    source_row: int
    marker: str
    article_id: str | None
    reference: str | None
    designation: str
    unit: str | None
    quantity: Decimal | None
    unit_price: Decimal | None
    amount: Decimal | None
    vat_rate: Decimal | None
    details: dict[str, dict[str, Decimal | None]] = field(default_factory=dict)


@dataclass
class Alignment:
    dce: Article | None
    act: Article | None
    method: str
    confidence: str
    issues: list[str] = field(default_factory=list)


def detect_scope(sheet_name: str, ws: Worksheet) -> str:
    sample = clean_text(sheet_name)
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True):
        sample += " " + clean_text(" ".join(str(v) for v in row if v is not None))
    if re.search(r"\b(pse|option|variante)\b", sample):
        return "OPTION"
    if "tranche" in sample:
        return "TRANCHE"
    return "BASE"


def find_markers(ws: Worksheet) -> tuple[int | None, int | None, list[int]]:
    marker_col = article_id_col = None
    article_rows: list[int] = []
    start_col = max(1, ws.max_column - 12)
    for row in range(1, ws.max_row + 1):
        for col in range(start_col, ws.max_column + 1):
            value = str(ws.cell(row, col).value or "").strip().upper()
            if value == ARTICLE_TAG:
                marker_col = col
                article_id_col = col + 1 if col < ws.max_column else None
                article_rows.append(row)
                break
    return marker_col, article_id_col, article_rows


def find_header_rows(ws: Worksheet) -> list[int]:
    rows = []
    for row in range(1, min(ws.max_row, 30) + 1):
        texts = [clean_text(ws.cell(row, col).value) for col in range(1, min(ws.max_column, 80) + 1)]
        joined = " | ".join(texts)
        if "designation" in joined and ("quantite" in joined or "montant" in joined or "p u" in joined):
            rows.append(row)
        elif any("quantite" in text or "prix unitaire" in text or text == "p u" for text in texts):
            rows.append(row)
    return sorted(set(rows))


def header_text(ws: Worksheet, col: int, header_rows: list[int]) -> str:
    values = []
    for row in range(1, max(header_rows or [1]) + 1):
        value = ws.cell(row, col).value
        if value is not None:
            values.append(str(value))
    return clean_text(" ".join(values))


def profile_sheet(ws: Worksheet) -> SheetProfile | None:
    marker_col, article_id_col, article_rows = find_markers(ws)
    header_rows = find_header_rows(ws)
    if not article_rows and not header_rows:
        return None

    scan_limit = min(ws.max_column, (marker_col - 1) if marker_col else 100)
    designation_col = unit_col = reference_col = None
    quantity_cols: list[int] = []
    pu_cols: list[int] = []
    amount_cols: list[int] = []
    vat_cols: list[int] = []

    for col in range(1, scan_limit + 1):
        text = header_text(ws, col, header_rows)
        last_header = clean_text(ws.cell(max(header_rows or [1]), col).value)
        if designation_col is None and "designation" in text:
            designation_col = col
        if unit_col is None and (last_header in {"u", "unite"} or " unite " in f" {text} "):
            unit_col = col
        if any(word in text for word in QUANTITY_WORDS):
            quantity_cols.append(col)
        if any(word in text for word in UNIT_PRICE_WORDS) or last_header in {"p u", "pu"}:
            pu_cols.append(col)
        if any(word in text for word in AMOUNT_WORDS) and "tva" not in text:
            amount_cols.append(col)
        if any(word in text for word in VAT_WORDS):
            vat_cols.append(col)

    if designation_col and designation_col > 1:
        reference_col = designation_col - 1
    elif article_rows:
        designation_col, unit_col, reference_col = 2, 3, 1

    # Les blocs sont construits à partir des colonnes quantité répétées.
    blocks: list[ColumnBlock] = []
    for index, qcol in enumerate(quantity_cols):
        next_q = quantity_cols[index + 1] if index + 1 < len(quantity_cols) else scan_limit + 1
        label = header_text(ws, qcol, [row for row in header_rows if row < max(header_rows or [1])]) or f"bloc_{index + 1}"
        pu = next((col for col in pu_cols if qcol < col < next_q), None)
        amount = next((col for col in amount_cols if qcol < col < next_q), None)
        vat = next((col for col in vat_cols if qcol <= col < next_q), None)
        blocks.append(ColumnBlock(label=label, quantity_col=qcol, unit_price_col=pu, amount_col=amount, vat_col=vat))

    # Cas sans quantité entreprise : on conserve tout de même PU, montant et TVA.
    if not blocks and (pu_cols or amount_cols or vat_cols):
        blocks.append(ColumnBlock(
            label="offre",
            quantity_col=None,
            unit_price_col=pu_cols[0] if pu_cols else None,
            amount_col=amount_cols[0] if amount_cols else None,
            vat_col=vat_cols[0] if vat_cols else None,
        ))

    warnings = []
    if not article_rows:
        warnings.append("Aucun marqueur ART détecté : un mode de secours sera nécessaire.")
    if not designation_col:
        warnings.append("Colonne Désignation non détectée.")
    if not blocks:
        warnings.append("Aucun bloc quantité / PU / montant détecté.")

    return SheetProfile(
        name=ws.title,
        scope_type=detect_scope(ws.title, ws),
        header_rows=header_rows,
        designation_col=designation_col,
        unit_col=unit_col,
        reference_col=reference_col,
        marker_col=marker_col,
        article_id_col=article_id_col,
        blocks=blocks,
        article_rows=article_rows,
        warnings=warnings,
    )


def read_vat(value: Any) -> Decimal | None:
    rate = decimal_or_none(value)
    if rate is None:
        return None
    if rate > 1:
        rate /= Decimal("100")
    if rate < 0 or rate > 1:
        return None
    return rate


def choose_summary_block(profile: SheetProfile, role: str) -> ColumnBlock | None:
    if not profile.blocks:
        return None
    # Dans un DCE détaillé, le dernier bloc est généralement le cumul.
    return profile.blocks[-1] if role == "DCE" else profile.blocks[0]


def parse_articles(ws: Worksheet, profile: SheetProfile, role: str) -> list[Article]:
    articles = []
    summary = choose_summary_block(profile, role)
    for row in profile.article_rows:
        designation = str(ws.cell(row, profile.designation_col).value or "").strip() if profile.designation_col else ""
        if not designation:
            continue
        details = {}
        for block in profile.blocks:
            details[block.label] = {
                "quantity": decimal_or_none(ws.cell(row, block.quantity_col).value) if block.quantity_col else None,
                "unit_price": decimal_or_none(ws.cell(row, block.unit_price_col).value) if block.unit_price_col else None,
                "amount": decimal_or_none(ws.cell(row, block.amount_col).value) if block.amount_col else None,
                "vat_rate": read_vat(ws.cell(row, block.vat_col).value) if block.vat_col else None,
            }
        articles.append(Article(
            workbook_role=role,
            sheet_name=ws.title,
            scope_type=profile.scope_type,
            source_row=row,
            marker=str(ws.cell(row, profile.marker_col).value or "") if profile.marker_col else "",
            article_id=str(ws.cell(row, profile.article_id_col).value or "").strip() or None if profile.article_id_col else None,
            reference=str(ws.cell(row, profile.reference_col).value or "").strip() or None if profile.reference_col else None,
            designation=designation,
            unit=str(ws.cell(row, profile.unit_col).value or "").strip() or None if profile.unit_col else None,
            quantity=decimal_or_none(ws.cell(row, summary.quantity_col).value) if summary and summary.quantity_col else None,
            unit_price=decimal_or_none(ws.cell(row, summary.unit_price_col).value) if summary and summary.unit_price_col else None,
            amount=decimal_or_none(ws.cell(row, summary.amount_col).value) if summary and summary.amount_col else None,
            vat_rate=read_vat(ws.cell(row, summary.vat_col).value) if summary and summary.vat_col else None,
            details=details,
        ))
    return articles


def parse_workbook(path: str | Path, role: str) -> dict[str, Any]:
    path = Path(path)
    workbook = load_workbook(path, data_only=True, read_only=False)
    sheets = []
    articles = []
    for ws in workbook.worksheets:
        profile = profile_sheet(ws)
        if profile is None:
            continue
        parsed = parse_articles(ws, profile, role)
        sheets.append(profile)
        articles.extend(parsed)
    return {"path": str(path), "role": role, "sheets": sheets, "articles": articles}


def article_keys(article: Article) -> list[tuple[str, str]]:
    keys = []
    if article.article_id:
        keys.append(("EXACT_ID", clean_text(article.article_id)))
    if article.reference:
        keys.append(("REFERENCE", normalize_reference(article.reference)))
    keys.append(("DESIGNATION_UNIT", clean_text(article.designation) + "|" + normalize_unit(article.unit)))
    return keys


def compare_values(dce: Article, act: Article) -> list[str]:
    issues = []
    if normalize_unit(dce.unit) != normalize_unit(act.unit):
        issues.append("UNITE_DIFFERENTE")
    if act.quantity is None:
        issues.append("QUANTITE_ENTREPRISE_ABSENTE")
    if act.unit_price is None:
        issues.append("PRIX_UNITAIRE_ABSENT")
    if act.amount is None:
        issues.append("MONTANT_ABSENT")
    if act.quantity is not None and act.unit_price is not None and act.amount is not None:
        expected = act.quantity * act.unit_price
        tolerance = max(Decimal("0.02"), abs(act.amount) * Decimal("0.0001"))
        if abs(expected - act.amount) > tolerance:
            issues.append("MONTANT_ACT_INCOHERENT")
    if dce.quantity is not None and act.quantity is not None and dce.quantity != 0:
        ratio = abs(act.quantity - dce.quantity) / abs(dce.quantity)
        if ratio >= Decimal("0.10"):
            issues.append("ECART_QUANTITE_GE_10_PCT")
    return issues


def align_articles(dce_articles: list[Article], act_articles: list[Article]) -> list[Alignment]:
    unused = set(range(len(act_articles)))
    indexes: dict[tuple[str, str, str], list[int]] = {}
    for index, article in enumerate(act_articles):
        for method, key in article_keys(article):
            indexes.setdefault((article.scope_type, method, key), []).append(index)

    results = []
    for dce in dce_articles:
        selected = None
        selected_method = ""
        for method, key in article_keys(dce):
            candidates = [i for i in indexes.get((dce.scope_type, method, key), []) if i in unused]
            if len(candidates) == 1:
                selected = candidates[0]
                selected_method = method
                break
        if selected is None:
            results.append(Alignment(dce=dce, act=None, method="UNMATCHED_DCE", confidence="NONE", issues=["ARTICLE_DCE_ABSENT_ACT"]))
            continue
        unused.remove(selected)
        act = act_articles[selected]
        confidence = "HIGH" if selected_method in {"EXACT_ID", "REFERENCE"} else "MEDIUM"
        results.append(Alignment(dce=dce, act=act, method=selected_method, confidence=confidence, issues=compare_values(dce, act)))

    for index in sorted(unused):
        results.append(Alignment(dce=None, act=act_articles[index], method="UNMATCHED_ACT", confidence="NONE", issues=["ARTICLE_SUPPLEMENTAIRE_ACT"]))
    return results


def analyze_offer(dce_path: str | Path, act_path: str | Path, enterprise: str, lot: str) -> dict[str, Any]:
    dce = parse_workbook(dce_path, "DCE")
    act = parse_workbook(act_path, "ACT")
    alignments = align_articles(dce["articles"], act["articles"])
    issue_counts: dict[str, int] = {}
    for alignment in alignments:
        for issue in alignment.issues:
            issue_counts[issue] = issue_counts.get(issue, 0) + 1
    return {
        "lot": str(lot).zfill(2),
        "enterprise": enterprise,
        "dce": {
            "path": dce["path"],
            "sheets": [asdict(sheet) for sheet in dce["sheets"]],
            "article_count": len(dce["articles"]),
        },
        "act": {
            "path": act["path"],
            "sheets": [asdict(sheet) for sheet in act["sheets"]],
            "article_count": len(act["articles"]),
        },
        "alignment_count": len(alignments),
        "issue_counts": issue_counts,
        "alignments": [asdict(item) for item in alignments],
    }


def write_analysis_json(result: dict[str, Any], output_path: str | Path) -> Path:
    target = Path(output_path)
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(json.dumps(json_value(result), ensure_ascii=False, indent=2), encoding="utf-8")
    return target


# ---------- Export Excel V2 ----------
def _safe_sheet_title(value: str) -> str:
    value = re.sub(r"[\\/*?:\[\]]", "_", value).strip()
    return value[:31] or "Feuille"


def _alignment_scope(item: dict[str, Any]) -> str:
    article = item.get("dce") or item.get("act") or {}
    return article.get("scope_type") or "BASE"


def _alignment_sheet(item: dict[str, Any], side: str) -> str:
    article = item.get(side) or {}
    return article.get("sheet_name") or ""


def _status_for_alignment(item: dict[str, Any]) -> str:
    if item.get("dce") is None and item.get("act") is not None:
        return "OPTION AJOUTEE ACT" if _alignment_scope(item) == "OPTION" else "ARTICLE AJOUTE ACT"
    if item.get("dce") is not None and item.get("act") is None:
        return "ABSENT ACT"
    return "ALIGNE"


def _create_comparison_sheet(workbook, title: str, items: list[dict[str, Any]], orange: str = "FA4616"):
    from openpyxl.styles import Alignment as CellAlignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws = workbook.create_sheet(_safe_sheet_title(title))
    ws.sheet_view.showGridLines = False
    headers = [
        "Statut", "Méthode", "Confiance", "Feuille DCE", "Feuille ACT",
        "Identifiant", "Référence", "Désignation", "Unité",
        "Quantité MOE", "Quantité entreprise", "Écart quantité", "Écart quantité %",
        "PU MOE", "PU entreprise", "Écart PU", "Écart PU %",
        "Montant MOE", "Montant entreprise", "Écart montant", "Écart montant %",
        "TVA article MOE", "TVA article entreprise", "Anomalies",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=orange)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = CellAlignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for item in items:
        dce = item.get("dce") or {}
        act = item.get("act") or {}
        source = dce or act
        row = ws.max_row + 1
        ws.append([
            _status_for_alignment(item), item.get("method"), item.get("confidence"),
            dce.get("sheet_name"), act.get("sheet_name"),
            source.get("article_id"), source.get("reference"), source.get("designation"), source.get("unit"),
            dce.get("quantity"), act.get("quantity"), None, None,
            dce.get("unit_price"), act.get("unit_price"), None, None,
            dce.get("amount"), act.get("amount"), None, None,
            dce.get("vat_rate"), act.get("vat_rate"), ", ".join(item.get("issues") or []),
        ])
        # Les écarts restent des formules Excel et restent vides si une valeur manque.
        ws.cell(row, 12, f'=IF(OR(J{row}="",K{row}=""),"",K{row}-J{row})')
        ws.cell(row, 13, f'=IF(OR(J{row}="",J{row}=0,L{row}=""),"",L{row}/J{row})')
        ws.cell(row, 16, f'=IF(OR(N{row}="",O{row}=""),"",O{row}-N{row})')
        ws.cell(row, 17, f'=IF(OR(N{row}="",N{row}=0,P{row}=""),"",P{row}/N{row})')
        ws.cell(row, 20, f'=IF(OR(R{row}="",S{row}=""),"",S{row}-R{row})')
        ws.cell(row, 21, f'=IF(OR(R{row}="",R{row}=0,T{row}=""),"",T{row}/R{row})')

        status = ws.cell(row, 1).value
        if status in {"OPTION AJOUTEE ACT", "ARTICLE AJOUTE ACT"}:
            ws.cell(row, 1).fill = PatternFill("solid", fgColor="FFF2CC")
        elif status == "ABSENT ACT":
            ws.cell(row, 1).fill = PatternFill("solid", fgColor="F4CCCC")
        elif item.get("issues"):
            ws.cell(row, 24).fill = PatternFill("solid", fgColor="FCE5CD")

    widths = {
        1: 22, 2: 18, 3: 12, 4: 28, 5: 28, 6: 16, 7: 14, 8: 58, 9: 10,
        10: 15, 11: 18, 12: 15, 13: 16, 14: 14, 15: 16, 16: 14, 17: 14,
        18: 16, 19: 19, 20: 16, 21: 16, 22: 16, 23: 20, 24: 40,
    }
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in ws.iter_rows(min_row=2):
        row[7].alignment = CellAlignment(wrap_text=True, vertical="top")
        row[23].alignment = CellAlignment(wrap_text=True, vertical="top")
    for col in range(10, 22):
        for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2, max_row=ws.max_row):
            for item_cell in cell:
                item_cell.number_format = '#,##0.00;[Red](#,##0.00);-'
    for col in (13, 17, 21, 22, 23):
        for row in range(2, ws.max_row + 1):
            ws.cell(row, col).number_format = '0.0%'
    return ws


def export_analysis_workbook(result: dict[str, Any], output_path: str | Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment as CellAlignment, Font, PatternFill

    target = Path(output_path)
    if target.suffix.lower() != ".xlsx":
        target = target.with_suffix(".xlsx")
    target.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    summary = wb.active
    summary.title = "Synthèse"
    summary.sheet_view.showGridLines = False
    summary["A1"] = f"Analyse des offres - Lot {result.get('lot', '')}"
    summary["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor="FA4616")
    summary.merge_cells("A1:D1")
    summary["A3"] = "Entreprise"
    summary["B3"] = result.get("enterprise")
    summary["A4"] = "Articles DCE"
    summary["B4"] = result.get("dce", {}).get("article_count")
    summary["A5"] = "Articles ACT"
    summary["B5"] = result.get("act", {}).get("article_count")
    summary["A6"] = "Alignements"
    summary["B6"] = result.get("alignment_count")

    base_items = [item for item in result.get("alignments", []) if _alignment_scope(item) != "OPTION"]
    option_items = [item for item in result.get("alignments", []) if _alignment_scope(item) == "OPTION"]
    summary["A8"] = "Lignes BASE"
    summary["B8"] = len(base_items)
    summary["A9"] = "Lignes OPTIONS"
    summary["B9"] = len(option_items)
    summary["A10"] = "Options ajoutées uniquement dans l'ACT"
    summary["B10"] = sum(1 for item in option_items if item.get("dce") is None and item.get("act") is not None)
    summary["A11"] = "Options DCE absentes de l'ACT"
    summary["B11"] = sum(1 for item in option_items if item.get("dce") is not None and item.get("act") is None)

    summary["A13"] = "Anomalie"
    summary["B13"] = "Nombre"
    for cell in summary[13]:
        if cell.column <= 2:
            cell.fill = PatternFill("solid", fgColor="595959")
            cell.font = Font(color="FFFFFF", bold=True)
    current = 14
    for issue, count in sorted((result.get("issue_counts") or {}).items()):
        summary.cell(current, 1, issue)
        summary.cell(current, 2, count)
        current += 1
    summary.column_dimensions["A"].width = 46
    summary.column_dimensions["B"].width = 18
    summary.column_dimensions["C"].width = 16
    summary.column_dimensions["D"].width = 16

    _create_comparison_sheet(wb, "BASE", base_items)
    # La feuille OPTIONS existe toujours, même si le DCE n'en contient aucune.
    _create_comparison_sheet(wb, "OPTIONS", option_items)

    # Une vue par feuille d'option ACT évite de mélanger plusieurs options ajoutées par l'entreprise.
    act_option_names = sorted({
        _alignment_sheet(item, "act") for item in option_items if _alignment_sheet(item, "act")
    })
    if len(act_option_names) > 1:
        for index, name in enumerate(act_option_names, 1):
            scoped = [item for item in option_items if _alignment_sheet(item, "act") == name]
            _create_comparison_sheet(wb, f"OPT {index} {name}", scoped)

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(target)
    return target


# ---------- Alignement V3 ----------
from difflib import SequenceMatcher


def _similarity(left: Any, right: Any) -> float:
    a, b = clean_text(left), clean_text(right)
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a, b).ratio()


def _sheet_key(name: str) -> str:
    text = clean_text(name)
    text = re.sub(r"\blot\s*n?\s*\d+\b", " ", text)
    text = re.sub(r"\b(dpgf|dqe|bpu|devis|offre|act|dce)\b", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _group_by_sheet(articles: list[Article]) -> dict[tuple[str, str], list[Article]]:
    groups: dict[tuple[str, str], list[Article]] = {}
    for article in articles:
        groups.setdefault((article.scope_type, article.sheet_name), []).append(article)
    return groups


def _article_id_set(articles: list[Article]) -> set[str]:
    return {clean_text(a.article_id) for a in articles if a.article_id}


def _reference_set(articles: list[Article]) -> set[str]:
    return {normalize_reference(a.reference) for a in articles if a.reference}


def pair_business_sheets(dce_articles: list[Article], act_articles: list[Article]) -> list[dict[str, Any]]:
    """Apparie les feuilles avant les lignes, sans forcer les options absentes d'un côté."""
    dce_groups = _group_by_sheet(dce_articles)
    act_groups = _group_by_sheet(act_articles)
    unused_act = set(act_groups)
    pairs: list[dict[str, Any]] = []

    for dce_key, dce_rows in dce_groups.items():
        scope, dce_name = dce_key
        candidates = []
        dce_ids = _article_id_set(dce_rows)
        dce_refs = _reference_set(dce_rows)
        for act_key in unused_act:
            act_scope, act_name = act_key
            if act_scope != scope:
                continue
            act_rows = act_groups[act_key]
            act_ids = _article_id_set(act_rows)
            act_refs = _reference_set(act_rows)
            id_overlap = len(dce_ids & act_ids)
            ref_overlap = len(dce_refs & act_refs)
            name_score = _similarity(_sheet_key(dce_name), _sheet_key(act_name))
            # Les identifiants dominent, puis les références, puis le nom de feuille.
            score = id_overlap * 1000 + ref_overlap * 100 + name_score * 10
            candidates.append((score, id_overlap, ref_overlap, name_score, act_key))
        candidates.sort(reverse=True, key=lambda item: item[:4])
        if candidates and candidates[0][0] > 0:
            score, id_overlap, ref_overlap, name_score, act_key = candidates[0]
            unused_act.remove(act_key)
            pairs.append({
                "scope": scope,
                "dce_sheet": dce_name,
                "act_sheet": act_key[1],
                "dce_articles": dce_rows,
                "act_articles": act_groups[act_key],
                "method": "SHEET_ID_OVERLAP" if id_overlap else ("SHEET_REFERENCE_OVERLAP" if ref_overlap else "SHEET_NAME"),
                "score": score,
            })
        else:
            pairs.append({
                "scope": scope, "dce_sheet": dce_name, "act_sheet": None,
                "dce_articles": dce_rows, "act_articles": [],
                "method": "DCE_SHEET_ONLY", "score": 0,
            })

    for act_key in sorted(unused_act):
        scope, act_name = act_key
        pairs.append({
            "scope": scope, "dce_sheet": None, "act_sheet": act_name,
            "dce_articles": [], "act_articles": act_groups[act_key],
            "method": "ACT_SHEET_ONLY", "score": 0,
        })
    return pairs


def _candidate_score(dce: Article, act: Article) -> tuple[int, list[str]]:
    score = 0
    reasons = []
    if dce.article_id and act.article_id and clean_text(dce.article_id) == clean_text(act.article_id):
        score += 10000
        reasons.append("ID")
    if dce.reference and act.reference and normalize_reference(dce.reference) == normalize_reference(act.reference):
        score += 2500
        reasons.append("REFERENCE")
    designation_similarity = _similarity(dce.designation, act.designation)
    if designation_similarity >= 0.98:
        score += 1500
        reasons.append("DESIGNATION_EXACTE")
    elif designation_similarity >= 0.90:
        score += 800
        reasons.append("DESIGNATION_PROCHE")
    elif designation_similarity >= 0.78:
        score += 300
        reasons.append("DESIGNATION_APPROX")
    if normalize_unit(dce.unit) and normalize_unit(dce.unit) == normalize_unit(act.unit):
        score += 200
        reasons.append("UNITE")
    return score, reasons


def _align_sheet_pair(dce_rows: list[Article], act_rows: list[Article]) -> list[Alignment]:
    unused = set(range(len(act_rows)))
    result: list[Alignment] = []
    for dce in dce_rows:
        ranked = []
        for index in unused:
            score, reasons = _candidate_score(dce, act_rows[index])
            ranked.append((score, index, reasons))
        ranked.sort(reverse=True, key=lambda item: item[0])
        if not ranked or ranked[0][0] < 1000:
            result.append(Alignment(dce=dce, act=None, method="UNMATCHED_DCE", confidence="NONE", issues=["ARTICLE_DCE_ABSENT_ACT"]))
            continue
        best_score, best_index, reasons = ranked[0]
        second_score = ranked[1][0] if len(ranked) > 1 else -1
        ambiguous = second_score == best_score
        # Un match faible basé uniquement sur la désignation doit rester à valider.
        if "ID" in reasons:
            method, confidence = "EXACT_ID", "HIGH"
        elif "REFERENCE" in reasons and "DESIGNATION_EXACTE" in reasons:
            method, confidence = "REFERENCE_DESIGNATION", "HIGH"
        elif "REFERENCE" in reasons:
            method, confidence = "REFERENCE", "MEDIUM"
        elif "DESIGNATION_EXACTE" in reasons and "UNITE" in reasons:
            method, confidence = "DESIGNATION_UNIT", "MEDIUM"
        else:
            method, confidence = "DESIGNATION_APPROX", "LOW"
        act = act_rows[best_index]
        issues = compare_values(dce, act)
        if ambiguous:
            issues.append("ALIGNEMENT_AMBIGU")
            confidence = "LOW"
        if confidence == "LOW":
            issues.append("ALIGNEMENT_A_VALIDER")
        unused.remove(best_index)
        result.append(Alignment(dce=dce, act=act, method=method, confidence=confidence, issues=issues))
    for index in sorted(unused):
        result.append(Alignment(dce=None, act=act_rows[index], method="UNMATCHED_ACT", confidence="NONE", issues=["ARTICLE_SUPPLEMENTAIRE_ACT"]))
    return result


def align_articles(dce_articles: list[Article], act_articles: list[Article]) -> list[Alignment]:
    """Alignement propre : feuilles d'abord, articles ensuite, ordre DCE conservé."""
    results: list[Alignment] = []
    for pair in pair_business_sheets(dce_articles, act_articles):
        if pair["dce_articles"] and pair["act_articles"]:
            results.extend(_align_sheet_pair(pair["dce_articles"], pair["act_articles"]))
        elif pair["dce_articles"]:
            for dce in pair["dce_articles"]:
                results.append(Alignment(dce=dce, act=None, method="DCE_SHEET_ONLY", confidence="NONE", issues=["FEUILLE_DCE_ABSENTE_ACT", "ARTICLE_DCE_ABSENT_ACT"]))
        else:
            for act in pair["act_articles"]:
                issue = "OPTION_AJOUTEE_ACT" if act.scope_type == "OPTION" else "FEUILLE_AJOUTEE_ACT"
                results.append(Alignment(dce=None, act=act, method="ACT_SHEET_ONLY", confidence="NONE", issues=[issue, "ARTICLE_SUPPLEMENTAIRE_ACT"]))
    return results


def alignment_diagnostics(dce_articles: list[Article], act_articles: list[Article]) -> list[dict[str, Any]]:
    diagnostics = []
    for pair in pair_business_sheets(dce_articles, act_articles):
        diagnostics.append({
            "scope": pair["scope"],
            "dce_sheet": pair["dce_sheet"],
            "act_sheet": pair["act_sheet"],
            "method": pair["method"],
            "score": pair["score"],
            "dce_article_count": len(pair["dce_articles"]),
            "act_article_count": len(pair["act_articles"]),
        })
    return diagnostics


_original_analyze_offer_v3 = analyze_offer

def analyze_offer(dce_path: str | Path, act_path: str | Path, enterprise: str, lot: str) -> dict[str, Any]:
    dce = parse_workbook(dce_path, "DCE")
    act = parse_workbook(act_path, "ACT")
    alignments = align_articles(dce["articles"], act["articles"])
    issue_counts: dict[str, int] = {}
    method_counts: dict[str, int] = {}
    confidence_counts: dict[str, int] = {}
    for alignment in alignments:
        method_counts[alignment.method] = method_counts.get(alignment.method, 0) + 1
        confidence_counts[alignment.confidence] = confidence_counts.get(alignment.confidence, 0) + 1
        for issue in alignment.issues:
            issue_counts[issue] = issue_counts.get(issue, 0) + 1
    return {
        "lot": str(lot).zfill(2), "enterprise": enterprise,
        "dce": {"path": dce["path"], "sheets": [asdict(s) for s in dce["sheets"]], "article_count": len(dce["articles"])},
        "act": {"path": act["path"], "sheets": [asdict(s) for s in act["sheets"]], "article_count": len(act["articles"])},
        "sheet_alignment": alignment_diagnostics(dce["articles"], act["articles"]),
        "alignment_count": len(alignments),
        "alignment_method_counts": method_counts,
        "alignment_confidence_counts": confidence_counts,
        "issue_counts": issue_counts,
        "alignments": [asdict(item) for item in alignments],
    }
