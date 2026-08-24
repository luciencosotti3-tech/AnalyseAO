# -*- coding: utf-8 -*-
"""Moteur direct V2 : comparaison, calculs et moyenne PU sans correction des offres."""
from __future__ import annotations
from collections import defaultdict
import math, re, unicodedata
from typing import Any
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
VERSION="3.6"
TOLERANCE=0.02
TEXT_VALUES=("compris","inclus","pm","so","sans objet","non chiffre","non valorise")
# CALC_QUALIFICATION_THRESHOLDS_V1
CALC_ABSOLUTE_SIGNIFICANT=1.00   # ecart absolu (EUR) a partir duquel un calcul est une erreur significative
CALC_RELATIVE_SIGNIFICANT=0.001   # ecart relatif (0.1 %) a partir duquel un calcul est une erreur significative
CALC_TECHNICAL_MINIMUM=0.02       # en-deca, l'ecart est un arrondi technique non significatif
# TOTAL_CONTROL_THRESHOLDS_V1
TOTAL_ABSOLUTE_TOLERANCE=0.02      # ecart absolu (EUR) tolere sur un controle de total certain
TOTAL_RELATIVE_TOLERANCE=0.00001  # ecart relatif tolere sur un controle de total certain (identique au seuil deja utilise avant ce patch)

def norm(value:Any)->str:
    if value is None:return ""
    text=unicodedata.normalize("NFKD",str(value));text="".join(c for c in text if not unicodedata.combining(c))
    return re.sub(r"\s+"," ",re.sub(r"[^a-z0-9%]+"," ",text.casefold().replace("²","2").replace("³","3"))).strip()

# NUMERIC_SEMANTIC_ALIGNMENT_V1
_NUMERIC_TEXT_RE=re.compile(r"^[+\-]?[0-9OIlS][0-9OIlS .,'’\u00a0]*(?:[%€])?$")
def number(value):
    if value is None or isinstance(value,bool):return None
    if isinstance(value,(int,float)):
        if isinstance(value,float) and (math.isnan(value) or math.isinf(value)):return None
        return round(float(value),8)
    raw=str(value).strip()
    if not raw or not _NUMERIC_TEXT_RE.fullmatch(raw) or not re.search(r"\d",raw):return None
    text=raw.replace('€','').replace('\u00a0','').replace(' ','').replace('%','').replace("'",'').replace('’','')
    text=text.translate(str.maketrans({'O':'0','I':'1','l':'1','S':'5'}))
    if ',' in text and '.' in text:
        if text.rfind(',')>text.rfind('.'):text=text.replace('.','').replace(',','.')
        else:text=text.replace(',','')
    else:text=text.replace(',','.')
    try:return round(float(text),8)
    except ValueError:return None

def money(value):return f"{value:,.2f}".replace(",","X").replace(".",",").replace("X"," ")
def display(value):
    n=number(value)
    if n is not None:return f"{n:g}"
    return str(value or "").strip()

# TEXT_EQUIVALENCE_V1
# Filtre les ecarts de designation/unite purement typographiques (section 10/11 du cahier des charges).
_UNIT_ALIASES={"u":"u","un":"u","unite":"u","m2":"m2","m3":"m3","ml":"ml","ens":"ens","ensemble":"ens","forf":"forf","forfait":"forf","kg":"kg"}
def _tight_signature(normalized_text):
    """Ignore les espaces et les zeros de tete (ex: x204ht/x 204 ht, PF3/PF03)."""
    tight=normalized_text.replace(" ","")
    return re.sub(r"(?<!\d)0+(\d)",r"\1",tight)
def designation_equivalent(dv,av):
    dn,an=norm(dv),norm(av)
    if dn==an:return True
    return _tight_signature(dn)==_tight_signature(an)
def unit_equivalent(dv,av):
    dn,an=norm(dv),norm(av)
    if dn==an:return True
    return _UNIT_ALIASES.get(dn,dn)==_UNIT_ALIASES.get(an,an)

def pure_blocks(ws):
    ribbon=None
    for row in range(1,min(ws.max_row,20)+1):
        if any(norm(c.value)=="estimation" for c in ws[row]):ribbon=row;break
    if ribbon is None:return None,[]
    blocks=[]
    for rg in ws.merged_cells.ranges:
        if rg.min_row==ribbon and rg.max_row==ribbon:
            label=str(ws.cell(ribbon,rg.min_col).value or "").strip()
            if label:blocks.append((label,rg.min_col,rg.max_col))
    return ribbon,sorted(blocks,key=lambda x:x[1])

def header_limit(ws,start,end,ribbon):
    """En-tête adaptatif : fenêtre historique, puis recherche prudente jusqu'à +60."""
    candidates=[]
    upper=min(ws.max_row,ribbon+60)
    for row in range(ribbon+1,upper+1):
        values=[norm(ws.cell(row,col).value) for col in range(start,end+1)]
        if "designation" not in values:continue
        # Une ligne Désignation est acceptée seulement si son voisinage vertical contient
        # au moins deux familles numériques métier, afin d'éviter une page de garde.
        lo=max(ribbon+1,row-4);hi=min(upper,row+4)
        nearby=" ".join(norm(ws.cell(r,col).value) for r in range(lo,hi+1) for col in range(start,end+1))
        evidence=sum(token in nearby for token in ("quantite","qte","p u","prix unitaire","montant"))
        if evidence>=2:candidates.append(row)
    return max(candidates) if candidates else min(ws.max_row,ribbon+8)

def header_tokens(ws,col,ribbon,limit):
    result=[]
    for row in range(ribbon+1,limit+1):
        value=ws.cell(row,col).value
        if value not in (None,""):result.append((row,norm(value)))
    return result

def role_map(ws,start,end,ribbon,is_dce=False):
    roles=defaultdict(list);limit=header_limit(ws,start,end,ribbon)
    for col in range(start,end+1):
        tokens=header_tokens(ws,col,ribbon,limit);joined=" ".join(v for _,v in tokens);last=tokens[-1][1] if tokens else ""
        if "moyenne p u entreprises" in joined:roles["average_pu"].append(col);continue
        if any(v=="designation" for _,v in tokens):roles["designation"].append(col);continue
        if last in {"u","un","unite"}:roles["unit"].append(col);continue
        if last in {"p u","pu","prix unitaire"}:roles["unit_price"].append(col);continue
        if "tva" in last or ("montant" in joined and "tva" in joined):roles["vat"].append(col);continue
        if "montant" in last:roles["amount"].append(col);continue
        if "quantite entreprise" in last or "qte ent" in last:roles["quantity_company"].append(col);continue
        if "quantite moe" in last or last in {"quantite","qte"}:roles["quantity_reference"].append(col);continue
    return dict(roles)

def summary_row(ws):
    for row in range(1,ws.max_row+1):
        if any(norm(c.value)=="synthese des controles par entreprise" for c in ws[row]):return row
    return None

def article_label(ws,row,roles,start,end):
    designation=""
    for col in roles.get("designation",[]):
        if ws.cell(row,col).value not in (None,""):designation=str(ws.cell(row,col).value).strip();break
    reference=""
    dcol=min(roles.get("designation",[min(end,start+1)]))
    for col in range(start,dcol):
        if ws.cell(row,col).value not in (None,""):reference=str(ws.cell(row,col).value).strip();break
    return " ".join(x for x in (reference,designation) if x).strip() or f"Ligne {row}"

def row_marker(ws,row):
    """Retourne le marqueur technique de ligne lorsqu'il est présent."""
    start=max(1,ws.max_column-20)
    for col in range(start,ws.max_column+1):
        value=str(ws.cell(row,col).value or "").strip().upper()
        if value in {"ART","CH3","CH4","CH5","CH6","STOT","TOTHT","TVA","TOTTTC"}:
            return value
    return ""


def article_row(ws,row,roles,start,end):
    """ART est prioritaire; l'heuristique n'est qu'un repli sans marqueurs."""
    marker=row_marker(ws,row)
    if marker:
        return marker=="ART"
    dcols=roles.get("designation",[])
    if not dcols:return False
    dcol=dcols[0];designation=norm(ws.cell(row,dcol).value)
    if not designation:return False
    forbidden=("total","sous total","montant ht","montant tva","montant ttc","tranche ferme","tranche optionnelle")
    if any(x in designation for x in forbidden):return False
    reference=""
    for col in range(start,dcol):
        raw=ws.cell(row,col).value
        if raw not in (None,""):
            reference=str(raw).strip();break
    if not re.search(r"\d",reference):return False
    unit_ok=any(norm(ws.cell(row,c).value) for c in roles.get("unit",[]))
    numeric_ok=any(number(ws.cell(row,c).value) is not None for role in ("quantity_reference","unit_price","amount") for c in roles.get(role,[]))
    return unit_ok or numeric_ok

def data_row(ws,row,roles,start=1,end=None):
    return article_row(ws,row,roles,start,end or ws.max_column)
def column_cardinality_issues(store,company,dce,roles):
    """Diagnostic unique de couverture, jamais preuve d'une offre erronée."""
    labels=(("quantity_reference","quantités"),("unit_price","prix unitaires"),("amount","montants"))
    for role,label in labels:
        expected=len(dce.get(role,[]))
        found=len(roles.get("quantity_company",[]) or roles.get(role,[])) if role=="quantity_reference" else len(roles.get(role,[]))
        if expected==found:continue
        if found==1 and expected>=1:continue
        add_issue(store,company,"STRUCTURE",0,"Structure des colonnes",f"Structure ACT différente du DCE : {label}, {expected} colonne(s) de référence et {found} colonne(s) ACT ; comparaison partielle.",(role,expected,found))
def paired(reference,company,allow_single_common=False):
    """Apparie par ordre ; conserve explicitement le cas d'une valeur commune."""
    if not reference or not company:return []
    if allow_single_common and len(company)==1:return [(reference[i],company[0],i+1) for i in range(len(reference))]
    if len(company)==1 and len(reference)>1:return [(reference[-1],company[0],len(reference))]
    return [(reference[i],company[i],i+1) for i in range(min(len(reference),len(company)))]

# CONFIRMED_CUMULATIVE_COLUMNS_V3
CUMULATIVE_CONFIRMATION_RATIO=0.80
CUMULATIVE_CONFIRMATION_MIN_ROWS=3
_STRONG_CUMULATIVE_HEADERS=(
    "cumul","montant total","total montant","total general",
    "total offre","total entreprise",
)

def _merged_header_anchor_map(ws,last_row):
    """Valeur d'ancrage des fusions d'en-tête, propagée sans modifier la feuille."""
    mapping={}
    for merged in ws.merged_cells.ranges:
        if merged.min_row>last_row or merged.max_row<1:continue
        value=ws.cell(merged.min_row,merged.min_col).value
        if value in (None,""):continue
        for row in range(max(1,merged.min_row),min(last_row,merged.max_row)+1):
            for col in range(merged.min_col,merged.max_col+1):mapping[(row,col)]=value
    return mapping


def _effective_header_text(ws,col,ribbon,limit=None,anchor_map=None):
    limit=limit or header_limit(ws,col,col,ribbon)
    anchor_map=anchor_map or _merged_header_anchor_map(ws,limit)
    values=[]
    for row in range(ribbon+1,limit+1):
        value=ws.cell(row,col).value
        if value in (None,""):value=anchor_map.get((row,col))
        if value not in (None,""):values.append(str(value))
    return norm(" ".join(values))


def _strong_cumulative_header(text):
    return any(token in text for token in _STRONG_CUMULATIVE_HEADERS)


def _candidate_validation(ws,candidate,details):
    comparable=coherent=0
    for row in range(1,ws.max_row+1):
        declared=number(ws.cell(row,candidate).value)
        values=[number(ws.cell(row,col).value) for col in details]
        if declared is None or not values or any(value is None for value in values):continue
        comparable+=1;expected=sum(values)
        tolerance=max(TOTAL_ABSOLUTE_TOLERANCE,abs(expected)*TOTAL_RELATIVE_TOLERANCE)
        if abs(declared-expected)<=tolerance:coherent+=1
    ratio=(coherent/comparable) if comparable else 0.0
    minimum=min(CUMULATIVE_CONFIRMATION_MIN_ROWS,comparable) if comparable else CUMULATIVE_CONFIRMATION_MIN_ROWS
    confirmed=comparable>=minimum and coherent>=minimum and ratio>=CUMULATIVE_CONFIRMATION_RATIO
    return {"comparable":comparable,"coherent":coherent,"ratio":ratio,"confirmed":confirmed}


def _row_has_significant_amount(ws,row,roles):
    return any((number(ws.cell(row,col).value) is not None and abs(number(ws.cell(row,col).value))>TOTAL_ABSOLUTE_TOLERANCE) for col in roles.get("amount",[]))


def _chapter_scope_coverage(ws,start_row,total_row,article_rows,dce_roles,company_roles):
    """Refuse un chapitre si une ligne chiffrée non qualifiée est ignorée."""
    articles=set(article_rows);dcols=dce_roles.get("designation",[]);dcol=dcols[0] if dcols else None
    ignored=[]
    for row in range(start_row+1,total_row):
        if row in articles:continue
        label=str(ws.cell(row,dcol).value or "").strip() if dcol else ""
        if _explicit_total_label(label):continue
        if _row_has_significant_amount(ws,row,company_roles):ignored.append(row)
    return {"certain":not ignored,"ignored_rows":ignored}

def header_text(ws,col,ribbon):
    """En-tête effectif, incluant le texte des cellules fusionnées couvrant col."""
    limit=min(ws.max_row,ribbon+60)
    return _effective_header_text(ws,col,ribbon,limit,_merged_header_anchor_map(ws,limit))

def is_total_column(ws,col,ribbon):
    """Candidat total/cumul sémantique. La confirmation numérique est faite par le layout."""
    return _strong_cumulative_header(header_text(ws,col,ribbon))

def add_issue(store,company,category,row,label,message,key,extra=None):
    dedupe=(company,category,row,key,norm(message))
    if dedupe in store["seen"]:return
    store["seen"].add(dedupe)
    item={"company":company,"category":category,"row":row,"label":label,"message":message}
    if extra:item.update(extra)
    store["items"].append(item)

# TOTAL_ALERTS_V1 : detection certaine des erreurs de total (horizontale, verticale,
# lot HT), mise en evidence de la cellule fautive, valeur declaree jamais modifiee.
def total_row_label(ws,row,roles,start,end):
    """Libelle d'une ligne de total/sous-total : n'exclut pas les mots 'total'."""
    for col in roles.get("designation",[]):
        v=ws.cell(row,col).value
        if v not in (None,""):return str(v).strip()
    for col in range(start,end+1):
        v=ws.cell(row,col).value
        if v not in (None,"") and not isinstance(v,(int,float)):return str(v).strip()
    return f"Ligne {row}"

def _total_error_border():
    thick=Side(style="thick",color="990000")
    return Border(left=thick,right=thick,top=thick,bottom=thick)

def _style_total_error_cell(ws,row,col,comment_text):
    from openpyxl.comments import Comment
    cell=ws.cell(row,col)
    cell.fill=PatternFill("solid",fgColor="F4CCCC")
    cell.border=_total_error_border()
    cell.font=Font(bold=True)
    current=cell.comment.text.splitlines() if cell.comment else []
    if comment_text not in current:
        cell.comment=Comment(("\n".join(current)+"\n" if current else "")+comment_text,"AnalyseAO")

_TOTAL_ERROR_PREFIX={
    "CHAPTER_TOTAL_ERROR":"ERREUR DE TOTAL (sous-total de chapitre)",
    "HT_TOTAL_ERROR":"ERREUR DE TOTAL (montant HT du lot)",
    "VAT_TOTAL_ERROR":"ERREUR DE TOTAL (TVA)",
    "TTC_TOTAL_ERROR":"ERREUR DE TOTAL (montant TTC)",
    "ROW_AGGREGATE_ERROR":"ERREUR DE CUMUL DE LIGNE (hors total général)",
}
_TOTAL_ERROR_TYPE_LABEL={
    "CHAPTER_TOTAL_ERROR":"sous-total de chapitre",
    "HT_TOTAL_ERROR":"montant HT du lot",
    "VAT_TOTAL_ERROR":"TVA",
    "TTC_TOTAL_ERROR":"montant TTC",
    "ROW_AGGREGATE_ERROR":"cumul de ligne (hors total général)",
}

def _highlight_total_label(ws,row,roles,start,end):
    """Met en évidence le libellé du total en erreur, sans jamais modifier son texte."""
    for col in roles.get("designation",[]):
        v=ws.cell(row,col).value
        if v not in (None,""):
            cell=ws.cell(row,col);cell.fill=PatternFill("solid",fgColor="F4CCCC");cell.font=Font(bold=True)
            return
    for col in range(start,end+1):
        v=ws.cell(row,col).value
        if v not in (None,"") and not isinstance(v,(int,float)):
            cell=ws.cell(row,col);cell.fill=PatternFill("solid",fgColor="F4CCCC");cell.font=Font(bold=True)
            return

def add_total_error(store,company,row,label,category,total_kind,expected,declared,ws,target_col,scope,roles=None,start=None,end=None):
    """Enregistre une erreur de total structurée et met en évidence la cellule déclarée (jamais modifiée)."""
    delta=declared-expected
    if abs(expected)>1e-9:
        relative_delta=abs(delta)/abs(expected);relative_text=f"{money(relative_delta*100)} %"
    else:
        relative_delta=None;relative_text="non applicable (total attendu nul)"
    type_label=_TOTAL_ERROR_TYPE_LABEL.get(category,category)
    message=(f"Type de total : {type_label}. "
             f"Total attendu pour le contrôle = {money(expected)} € ; "
             f"total déclaré par l'entreprise = {money(declared)} € ; "
             f"écart = {money(delta)} € (écart relatif au résultat attendu = {relative_text}). "
             f"Périmètre contrôlé : {scope}")
    target_cell=ws.cell(row,target_col).coordinate
    extra={
        "total_kind":total_kind,"expected":round(expected,2),"declared":round(declared,2),
        "delta":round(delta,2),"relative_delta":round(relative_delta,6) if relative_delta is not None else None,
        "target_cell":target_cell,"scope":scope,
    }
    before=len(store["items"])
    add_issue(store,company,category,row,label,message,("total_error",category,total_kind,target_col),extra)
    if len(store["items"])>before:
        prefix=_TOTAL_ERROR_PREFIX.get(category,"ERREUR DE TOTAL")
        comment=(f"{prefix} — {message} "
                 f"Aucune correction n'a été appliquée ; la valeur déclarée par l'entreprise est conservée.")
        _style_total_error_cell(ws,row,target_col,comment)
        if roles is not None and start is not None and end is not None:
            _highlight_total_label(ws,row,roles,start,end)

def horizontal_total_checks(ws,row,company,article,dce_roles,roles,ribbon,store,stats):
    """Cumul horizontal d'une ligne (bâtiments/zones) : ROW_AGGREGATE_ERROR, jamais un total général."""
    dce_amounts=dce_roles.get("amount",[])
    act_amounts=roles.get("amount",[])
    dce_totals=[c for c in dce_amounts if is_total_column(ws,c,ribbon)]
    act_totals=[c for c in act_amounts if is_total_column(ws,c,ribbon)]
    if len(dce_totals)!=1 or len(act_totals)!=1:return
    dtotal,atotal=dce_totals[0],act_totals[0]
    ddetails=[c for c in dce_amounts if c<dtotal and c not in dce_totals]
    adetails=[c for c in act_amounts if c<atotal and c not in act_totals]
    if len(ddetails)<2 or len(ddetails)!=len(adetails):return
    vals=[number(ws.cell(row,c).value) for c in adetails]
    declared=number(ws.cell(row,atotal).value)
    if declared is None or any(v is None for v in vals):return
    expected=sum(vals);stats["totals_checked"]+=1
    tolerance=max(TOTAL_ABSOLUTE_TOLERANCE,abs(expected)*TOTAL_RELATIVE_TOLERANCE)
    row_delta=abs(expected-declared)
    _reliability_register(store,company,"ROW_AGGREGATE",min(1.0,row_delta/tolerance),row_delta>tolerance)
    if abs(expected-declared)>tolerance:
        stats["total_errors"]=stats.get("total_errors",0)+1
        scope=f"Somme des {len(adetails)} montant(s) détaillé(s) de cette ligne (colonnes ACT alignées sur le périmètre DCE)."
        add_total_error(store,company,row,article,"ROW_AGGREGATE_ERROR","horizontal",expected,declared,ws,atotal,scope)

# CHAPTER_LOT_VAT_TTC_TOTALS_V1 : sous-totaux de chapitre, total HT, TVA et TTC du
# lot. Marqueur technique (STOT/TOTHT/TVA/TOTTTC) lorsqu'il est présent ; à défaut,
# libellé de la ligne de référence DCE (cas réel restructuré, cf. audit
# Fichiers_tests/_resultats/22204_multi_company_GCC_tous_lots.xlsx : les marqueurs
# ne survivent pas à la restructuration multi-entreprises).
_VAT_RATE_HEADER_RE=re.compile(r"tva\D{0,6}([0-9]+(?:[.,][0-9]+)?)\s*%",re.IGNORECASE)
_VAT_RATE_LABEL_RE=re.compile(r"([0-9]+(?:[.,][0-9]+)?)\s*%")
_GENERAL_HT_RE=re.compile(r"(montant|total).*\bht\b|\btotal\b.*\blot\b")

def _vat_rate_amount_columns(ws,start,end,ribbon):
    """Colonnes 'Montant TVA X%' : montant de TVA déjà calculé pour ce taux (une colonne par taux)."""
    limit=header_limit(ws,start,end,ribbon);found=[]
    for col in range(start,end+1):
        raw=" ".join(str(ws.cell(r,col).value or "") for r in range(ribbon+1,limit+1))
        low=raw.casefold()
        if "montant" not in low or "tva" not in low:continue
        if _VAT_RATE_HEADER_RE.search(raw):found.append(col)
    return found

def _rate_from_text(text):
    m=_VAT_RATE_LABEL_RE.search(text)
    if not m:return None
    try:return round(float(m.group(1).replace(",","."))/100.0,4)
    except ValueError:return None

def _row_text_label(ws,row,roles,start,end):
    """Libellé textuel d'une ligne (désignation, ou première cellule non numérique)."""
    for col in roles.get("designation",[]):
        v=ws.cell(row,col).value
        if v not in (None,""):return str(v).strip()
    for col in range(start,end+1):
        v=ws.cell(row,col).value
        if v not in (None,"") and not isinstance(v,(int,float)):return str(v).strip()
    return None

# ARTICLE_SUM_TOTALS_V1
_ARTICLE_TOTAL_TTC_RE = re.compile(
    r"\b(?:montant|total)\s+(?:t\s*t\s*c|ttc)\b|"
    r"\b(?:montant|total)\s+toutes\s+taxes\s+comprises\b"
)
_ARTICLE_TOTAL_VAT_RE = re.compile(
    r"\b(?:montant\s+(?:de\s+la\s+)?)?(?:t\s*v\s*a|tva)\b|"
    r"\btaxe\s+sur\s+la\s+valeur\s+ajoutee\b"
)
_ARTICLE_TOTAL_HT_RE = re.compile(
    r"\b(?:montant|total)\s+(?:h\s*t|ht)\b|"
    r"\b(?:montant|total)\s+hors\s+taxes?\b"
)
_ARTICLE_VAT_RATE_RE = re.compile(
    r"\b(?:t\s*v\s*a|tva|taxe\s+sur\s+la\s+valeur\s+ajoutee)\b"
    r"[^0-9]{0,20}(\d{1,2}(?:[.,]\d+)?)\s*%",
    re.IGNORECASE,
)


def _article_total_text(value):
    text = norm(value)
    text = re.sub(r"\bt\s+t\s+c\b", "ttc", text)
    text = re.sub(r"\bt\s+v\s+a\b", "tva", text)
    text = re.sub(r"\bh\s+t\b", "ht", text)
    return re.sub(r"\s+", " ", text).strip()


def _article_classify_total(label, marker=""):
    """Classification exclusive : TTC, TVA, HT, chapitre marque, aucun."""
    text = _article_total_text(label)
    marker = str(marker or "").strip().upper()
    if marker == "TOTTTC" or _ARTICLE_TOTAL_TTC_RE.search(text):
        return "TTC"
    if marker == "TVA" or _ARTICLE_TOTAL_VAT_RE.search(text):
        return "VAT"
    if "ttc" in text or "tva" in text:
        return "NONE"
    if marker == "TOTHT" or _ARTICLE_TOTAL_HT_RE.search(text):
        return "HT"
    if marker == "STOT":
        return "CHAPTER"
    return "NONE"


def _article_explicit_vat_rate(label):
    """Taux explicite uniquement : aucun repli implicite a 20 %."""
    match = _ARTICLE_VAT_RATE_RE.search(str(label or ""))
    if not match:
        return None
    try:
        rate = float(match.group(1).replace(",", ".")) / 100.0
    except ValueError:
        return None
    return rate if 0.0 < rate <= 1.0 else None


def _article_amount_layout(ws,roles,ribbon):
    """Choisit exactement une source par article : cumul confirmé OU détails.

    Un candidat cumul est confirmé par la cohérence numérique avec les autres
    colonnes de montant. En cas d'ambiguïté, aucun total général n'est contrôlé.
    """
    cols=list(roles.get("amount",[]))
    if not cols:return {"status":"AMBIGUOUS","details":[],"total":None,"reason":"aucune colonne montant","validation":[]}
    limit=header_limit(ws,min(cols),max(cols),ribbon);anchors=_merged_header_anchor_map(ws,limit)
    headers={col:_effective_header_text(ws,col,ribbon,limit,anchors) for col in cols}
    candidates=[]
    for candidate in cols:
        if not _strong_cumulative_header(headers[candidate]):continue
        details=[col for col in cols if col!=candidate]
        validation=_candidate_validation(ws,candidate,details)
        candidates.append({"column":candidate,"details":details,"header":headers[candidate],**validation})
    confirmed=[item for item in candidates if item["confirmed"]]
    if len(confirmed)==1:
        item=confirmed[0]
        return {"status":"CERTAIN","details":item["details"],"total":item["column"],"reason":f"colonne cumul confirmée ({item['coherent']}/{item['comparable']} lignes cohérentes)","validation":candidates}
    if len(confirmed)>1:
        return {"status":"AMBIGUOUS","details":[],"total":None,"reason":"plusieurs colonnes cumul confirmées","validation":candidates}
    if candidates:
        return {"status":"AMBIGUOUS","details":[],"total":None,"reason":"candidat cumul non confirmé numériquement","validation":candidates}
    if len(cols)==1:
        return {"status":"CERTAIN","details":[],"total":cols[0],"reason":"colonne montant unique","validation":[]}
    return {"status":"CERTAIN","details":cols,"total":None,"reason":"somme des montants détaillés, aucun cumul annoncé","validation":[]}


# ADAPTIVE_ARTICLE_TOTALS_V2

def _markers_available(ws, first_row, last_row, wanted=None):
    wanted=set(wanted or {"ART"})
    return any(row_marker(ws,row) in wanted for row in range(max(1,first_row),min(ws.max_row,last_row)+1))


def elementary_article_rows(ws, first_row, last_row, dce_roles, ds, de):
    """Articles adaptatifs : marqueurs stricts s'ils existent, sinon article_row()."""
    first_row=max(1,first_row);last_row=min(ws.max_row,last_row)
    marker_mode=_markers_available(ws,first_row,last_row,{"ART"})
    result=[]
    for row in range(first_row,last_row+1):
        valid=(row_marker(ws,row)=="ART") if marker_mode else article_row(ws,row,dce_roles,ds,de)
        if valid:result.append(row)
    return result


def _chapter_name(label):
    text=norm(label)
    return re.sub(r"^(?:sous\s+total|total)\s+","",text).strip()


def _explicit_total_label(label):
    text=norm(label)
    if not text.startswith(("total ","sous total ")):return False
    if any(x in text for x in ("montant ht","montant tva","montant ttc","toutes taxes")):return False
    return bool(_chapter_name(text))


def _visible_chapter_scopes(ws,ribbon,stop,dce_roles,ds,de):
    """Paires exactes TITRE / Total TITRE avec contrôle de couverture."""
    dcols=dce_roles.get("designation",[])
    if not dcols:return []
    dcol=dcols[0];scopes=[];used=set()
    for total_row in range(ribbon+1,stop):
        total_label=str(ws.cell(total_row,dcol).value or "").strip()
        if not _explicit_total_label(total_label):continue
        name=_chapter_name(total_label);headings=[]
        for row in range(total_row-1,ribbon,-1):
            label=str(ws.cell(row,dcol).value or "").strip()
            if _explicit_total_label(label):break
            if norm(label)==name:headings.append(row)
        if len(headings)!=1:continue
        heading=headings[0]
        article_rows=elementary_article_rows(ws,heading+1,total_row-1,dce_roles,ds,de)
        key=(heading,total_row,tuple(article_rows))
        if article_rows and key not in used:
            used.add(key);scopes.append({"row":total_row,"start":heading,"level":None,"article_rows":article_rows,"mode":"VISIBLE_TITLE_TOTAL","coverage_pending":True})
    return scopes

def representative_article_amount(ws, row, roles, ribbon):
    """Un montant ACT unique pour une ligne déjà qualifiée comme article.

    La qualification appartient à elementary_article_rows(); aucun second test ART
    n'est effectué ici, car les marqueurs ne sont pas exportés après restructuration.
    """
    layout=_article_amount_layout(ws,roles,ribbon)
    if layout["status"]!="CERTAIN":
        return {"status":"AMBIGUOUS","amount":None,"source":None,"columns":[],"reason":layout["reason"]}
    total_col=layout["total"]
    if total_col is not None:
        total=number(ws.cell(row,total_col).value)
        if total is not None:
            return {"status":"CERTAIN","amount":total,"source":"TOTAL_COLUMN","columns":[total_col],"reason":layout["reason"]}
        if not layout["details"]:
            return {"status":"INCOMPLETE","amount":None,"source":None,"columns":[total_col],"reason":"montant article non numérique"}
    detail_cols=layout["details"]
    if not detail_cols:
        return {"status":"INCOMPLETE","amount":None,"source":None,"columns":[],"reason":"aucun montant représentatif"}
    values=[number(ws.cell(row,col).value) for col in detail_cols]
    if any(value is None for value in values):
        return {"status":"INCOMPLETE","amount":None,"source":None,"columns":detail_cols,"reason":"montants détaillés incomplets ou textuels"}
    return {"status":"CERTAIN","amount":sum(values),"source":"DETAIL_SUM","columns":detail_cols,"reason":"somme complète des montants détaillés"}


def _article_scope_sum(ws, rows, roles, ribbon):
    items = []
    for row in rows:
        item = representative_article_amount(ws, row, roles, ribbon)
        item["row"] = row
        items.append(item)
    invalid = [item for item in items if item["status"] != "CERTAIN"]
    if invalid:
        return {
            "status": "NOT_CHECKED", "amount": None, "items": items,
            "reason": f"{len(invalid)} article(s) sans montant representatif certain",
        }
    if not items:
        return {"status": "NOT_CHECKED", "amount": None, "items": [], "reason": "aucun article elementaire"}
    return {"status": "CERTAIN", "amount": sum(item["amount"] for item in items), "items": items, "reason": f"somme de {len(items)} article(s) elementaire(s) ACT"}


def _article_financial_rows(ws, ribbon, stop, dce_roles, ds, de):
    result = []
    for row in range(ribbon + 1, stop):
        label = _row_text_label(ws, row, dce_roles, ds, de) or ""
        marker = row_marker(ws, row)
        kind = _article_classify_total(label, marker)
        if kind != "NONE":
            result.append({"row": row, "label": label, "marker": marker, "kind": kind})
    return result


def _article_chapter_scopes(ws,ribbon,stop,dce_roles=None,ds=None,de=None):
    """Mode technique CHx/STOT si présent, sinon paires visibles titre/Total titre."""
    technical=_markers_available(ws,ribbon+1,stop-1,{"CH3","CH4","CH5","CH6","STOT"})
    if not technical:
        if dce_roles is None or ds is None or de is None:return []
        return _visible_chapter_scopes(ws,ribbon,stop,dce_roles,ds,de)
    stack=[];scopes=[]
    for row in range(ribbon+1,stop):
        marker=row_marker(ws,row)
        if marker in {"CH3","CH4","CH5","CH6"}:
            level=int(marker[-1]);stack=[x for x in stack if x["level"]<level];stack.append({"level":level,"start":row});continue
        if marker!="STOT" or not stack:continue
        current=stack[-1]
        article_rows=elementary_article_rows(ws,current["start"]+1,row-1,dce_roles,ds,de)
        if article_rows:scopes.append({"row":row,"start":current["start"],"level":current["level"],"article_rows":article_rows,"mode":"TECHNICAL_MARKERS"})
        stack.pop()
    return scopes

def _chapter_row_kind(ws,row,dce_roles,ds,de):
    """Classification stricte commune : TTC > TVA > HT > chapitre marque."""
    label=_row_text_label(ws,row,dce_roles,ds,de) or ""
    return _article_classify_total(label,row_marker(ws,row))

def chapter_and_lot_total_checks(ws,ribbon,stop,company,roles,start,end,dce_roles,ds,de,store,stats):
    """Sous-totaux et HT recomposés depuis les articles visibles ou marqués."""
    layout=_article_amount_layout(ws,roles,ribbon)
    target_col=layout.get("total") or (roles.get("amount",[]) or [None])[-1]
    if target_col is None:return
    for scope in _article_chapter_scopes(ws,ribbon,stop,dce_roles,ds,de):
        result=_article_scope_sum(ws,scope["article_rows"],roles,ribbon)
        if result["status"]!="CERTAIN":stats["chapter_totals_not_checked_incomplete_articles"]+=1;continue
        declared=number(ws.cell(scope["row"],target_col).value)
        if declared is None:stats["chapter_totals_not_checked_non_numeric"]+=1;continue
        expected=result["amount"];stats["chapter_totals_checked"]+=1
        tolerance=max(TOTAL_ABSOLUTE_TOLERANCE,abs(expected)*TOTAL_RELATIVE_TOLERANCE)
        if abs(expected-declared)>tolerance:
            label=_row_text_label(ws,scope["row"],dce_roles,ds,de) or f"Ligne {scope['row']}"
            mode="marqueurs techniques" if scope["mode"]=="TECHNICAL_MARKERS" else "paire exacte titre / total"
            detail=f"Somme de {len(scope['article_rows'])} article(s) élémentaire(s) ACT, périmètre établi par {mode} (lignes {scope['start']} à {scope['row']})."
            add_total_error(store,company,scope["row"],label,"CHAPTER_TOTAL_ERROR","chapter",expected,declared,ws,target_col,detail,roles,start,end);stats["chapter_total_errors"]+=1
    financial=_article_financial_rows(ws,ribbon,stop,dce_roles,ds,de)
    ht_rows=[x for x in financial if x["kind"]=="HT"]
    if len(ht_rows)!=1:stats["ht_not_checked_ambiguous_row"]+=1;return
    ht=ht_rows[0]
    first=header_limit(ws,ds,de,ribbon)+1
    article_rows=elementary_article_rows(ws,first,ht["row"]-1,dce_roles,ds,de)
    result=_article_scope_sum(ws,article_rows,roles,ribbon)
    if result["status"]!="CERTAIN":stats["ht_not_checked_incomplete_articles"]+=1;return
    declared=number(ws.cell(ht["row"],target_col).value)
    if declared is None:stats["ht_not_checked_non_numeric"]+=1;return
    expected=result["amount"];stats["ht_totals_checked"]+=1
    tolerance=max(TOTAL_ABSOLUTE_TOLERANCE,abs(expected)*TOTAL_RELATIVE_TOLERANCE)
    if abs(expected-declared)>tolerance:
        scope=f"Somme de {len(article_rows)} article(s) élémentaire(s) ACT reconnus dans le lot, chacun compté une seule fois."
        add_total_error(store,company,ht["row"],ht["label"],"HT_TOTAL_ERROR","ht",expected,declared,ws,target_col,scope,roles,start,end);stats["ht_total_errors"]+=1

def vat_and_ttc_checks(ws,ribbon,stop,company,roles,start,end,dce_roles,ds,de,store,stats):
    """TVA/TTC depuis la même somme adaptative d'articles que le HT."""
    layout=_article_amount_layout(ws,roles,ribbon)
    target_col=layout.get("total") or (roles.get("amount",[]) or [None])[-1]
    if target_col is None:return
    financial=_article_financial_rows(ws,ribbon,stop,dce_roles,ds,de)
    ht_rows=[x for x in financial if x["kind"]=="HT"];vat_rows=[x for x in financial if x["kind"]=="VAT"];ttc_rows=[x for x in financial if x["kind"]=="TTC"]
    if len(ht_rows)!=1:stats["vat_ttc_not_checked_ht_ambiguous"]+=1;return
    ht=ht_rows[0];first=header_limit(ws,ds,de,ribbon)+1
    article_rows=elementary_article_rows(ws,first,ht["row"]-1,dce_roles,ds,de)
    article_sum=_article_scope_sum(ws,article_rows,roles,ribbon)
    if article_sum["status"]!="CERTAIN":stats["vat_ttc_not_checked_incomplete_articles"]+=1;return
    article_ht=article_sum["amount"];checked=[]
    if len(vat_rows)==1:
        vat=vat_rows[0];rate=_article_explicit_vat_rate(vat["label"]);declared=number(ws.cell(vat["row"],target_col).value)
        if rate is None:stats["vat_not_checked_missing_explicit_rate"]+=1
        elif declared is None:stats["vat_not_checked_non_numeric"]+=1
        else:
            expected=article_ht*rate;checked.append((vat,declared,expected,rate));stats["vat_totals_checked"]+=1
            tolerance=max(TOTAL_ABSOLUTE_TOLERANCE,abs(expected)*TOTAL_RELATIVE_TOLERANCE)
            if abs(expected-declared)>tolerance:
                scope=f"Somme certaine de {len(article_rows)} article(s) ACT × taux explicite de {money(rate*100)} %."
                add_total_error(store,company,vat["row"],vat["label"],"VAT_TOTAL_ERROR","vat",expected,declared,ws,target_col,scope,roles,start,end);stats["vat_total_errors"]+=1
    elif vat_rows:stats["vat_not_checked_multiple_rates_without_article_bases"]+=1
    if len(ttc_rows)!=1:return
    if len(vat_rows)!=1 or len(checked)!=1:stats["ttc_not_checked_vat_ambiguous"]+=1;return
    ttc=ttc_rows[0];declared=number(ws.cell(ttc["row"],target_col).value)
    if declared is None:stats["ttc_not_checked_non_numeric"]+=1;return
    expected=article_ht+checked[0][2];stats["ttc_totals_checked"]+=1
    tolerance=max(TOTAL_ABSOLUTE_TOLERANCE,abs(expected)*TOTAL_RELATIVE_TOLERANCE)
    if abs(expected-declared)>tolerance:
        rate=checked[0][3];scope=f"Somme certaine des articles ACT + TVA attendue au taux explicite de {money(rate*100)} %."
        add_total_error(store,company,ttc["row"],ttc["label"],"TTC_TOTAL_ERROR","ttc",expected,declared,ws,target_col,scope,roles,start,end);stats["ttc_total_errors"]+=1

def update_average_pu(ws,dce_roles,company_roles,stop,dce_start,dce_end):
    from openpyxl.comments import Comment
    cols=dce_roles.get("average_pu",[])
    if not cols:return {"column":None,"written":0,"cleared":0,"article_rows":0,"contributors":{}}
    target=cols[0];written=cleared=article_rows=0;contributors={}
    for row in range(1,stop):
        cell=ws.cell(row,target);old=cell.value
        if not article_row(ws,row,dce_roles,dce_start,dce_end):
            if row>header_limit(ws,dce_start,dce_end,1) and old not in (None,""):
                cell.value=None;cell.comment=None;cleared+=1
            continue
        article_rows+=1;representatives=[]
        for roles in company_roles:
            values=[]
            for col in roles.get("unit_price",[]):
                n=number(ws.cell(row,col).value)
                if n is not None and math.isfinite(n) and n>0:values.append(n)
            if values:representatives.append(sum(values)/len(values))
        new=round(sum(representatives)/len(representatives),2) if representatives else None
        contributors[row]=len(representatives)
        if new is None:
            if old not in (None,""):cell.value=None;cleared+=1
            cell.comment=None
        else:
            cell.value=new;cell.number_format='#,##0.00';written+=1
            cell.comment=Comment(f"Moyenne calculée sur {len(representatives)} entreprise(s) sur {len(company_roles)}. PU nuls, vides et textuels exclus.","AnalyseAO")
    return {"column":target,"written":written,"cleared":cleared,"article_rows":article_rows,"contributors":contributors}
# HT_CUMULATIVE_TOTAL_V2
ENABLE_CHAPTER_TOTAL_CHECKS=False
ENABLE_VAT_TTC_CHECKS=False
ENABLE_HT_TOTAL_CHECK=True


def _ht_row_label(ws,row,dce_roles,ds,de):
    return (_row_text_label(ws,row,dce_roles,ds,de) or "").strip()


def _ht_explicit_label(label,marker=""):
    """HT explicitement nommé. La normalisation absorbe ponctuation/casse/accents."""
    marker=str(marker or "").strip().upper()
    if marker=="TOTHT":return True
    text=_article_total_text(label)
    if any(token in text for token in ("tva","ttc","toutes taxes")):return False
    return bool(re.search(r"\b(?:montant|total)\s+(?:ht|hors\s+taxes?)\b",text))


def _ht_tax_row(label,marker=""):
    marker=str(marker or "").strip().upper()
    if marker in {"TVA","TOTTTC"}:return True
    text=_article_total_text(label)
    return bool(re.search(r"\b(?:tva|ttc)\b|toutes\s+taxes",text))


def _ht_summary_row(label,marker=""):
    """Toute ligne qui récapitule d'autres lignes est exclue de la somme."""
    marker=str(marker or "").strip().upper()
    if marker in {"CH3","CH4","CH5","CH6","STOT","TOTHT","TVA","TOTTTC"}:return True
    text=norm(label)
    if not text:return False
    return (
        text=="total" or text.startswith("total ") or
        text=="sous total" or text.startswith("sous total ") or
        text=="cumul" or text.startswith("cumul ") or
        text.startswith("recapitulatif ") or text=="recapitulatif" or
        text.startswith("synthese ") or text=="synthese" or
        _ht_explicit_label(label,marker) or _ht_tax_row(label,marker)
    )


def _ht_matches_previous_heading(ws,row,label,dce_roles,ribbon):
    """Un 'Total X' apparié à un titre X précédent est un sous-total, jamais le HT."""
    text=norm(label)
    if not (text.startswith("total ") or text.startswith("sous total ")):return False
    name=re.sub(r"^(?:sous\s+total|total)\s+","",text).strip()
    if not name:return False
    dcols=dce_roles.get("designation",[])
    if not dcols:return False
    dcol=dcols[0]
    return any(norm(ws.cell(previous,dcol).value)==name for previous in range(ribbon+1,row))


def _ht_terminal_candidate(label,marker=""):
    """Total implicitement HT, confirmé ensuite par sa position terminale et son unicité."""
    marker=str(marker or "").strip().upper()
    if marker in {"TVA","TOTTTC","STOT"}:return False
    text=norm(label)
    if not text or _ht_tax_row(label,marker):return False
    if text in {"total","total general","montant total"}:return True
    if not (text.startswith("total ") or text.startswith("montant ")):return False
    return any(scope in text for scope in ("general","lot","travaux","offre","marche"))


def _find_declared_ht_row(ws,ribbon,stop,dce_roles,ds,de,cumulative_col):
    """HT explicite prioritaire, sinon total général terminal unique."""
    rows=[]
    for row in range(ribbon+1,stop):
        label=_ht_row_label(ws,row,dce_roles,ds,de)
        marker=row_marker(ws,row)
        value=number(ws.cell(row,cumulative_col).value)
        rows.append({"row":row,"label":label,"marker":marker,"value":value})
    explicit=[item for item in rows if item["value"] is not None and _ht_explicit_label(item["label"],item["marker"])]
    if len(explicit)==1:
        item=explicit[0].copy();item["source"]="EXPLICIT_HT";return {"status":"CERTAIN","item":item}
    if len(explicit)>1:return {"status":"NOT_CHECKED","reason":"multiple_explicit_ht","candidates":explicit}

    tax_rows=[item["row"] for item in rows if _ht_tax_row(item["label"],item["marker"])]
    terminal_end=min(tax_rows) if tax_rows else stop
    candidates=[]
    for item in rows:
        if item["row"]>=terminal_end or item["value"] is None:continue
        if not _ht_terminal_candidate(item["label"],item["marker"]):continue
        if _ht_matches_previous_heading(ws,item["row"],item["label"],dce_roles,ribbon):continue
        candidates.append(item)
    # Seuls les candidats terminaux sont retenus : aucun montant direct non récapitulatif après eux.
    terminal=[]
    for item in candidates:
        later_direct=False
        for row in range(item["row"]+1,terminal_end):
            label=_ht_row_label(ws,row,dce_roles,ds,de);value=number(ws.cell(row,cumulative_col).value)
            if value is not None and abs(value)>TOTAL_ABSOLUTE_TOLERANCE and not _ht_summary_row(label,row_marker(ws,row)):
                later_direct=True;break
        if not later_direct:terminal.append(item)
    if len(terminal)==1:
        item=terminal[0].copy();item["source"]="IMPLICIT_TERMINAL_TOTAL";return {"status":"CERTAIN","item":item}
    return {"status":"NOT_CHECKED","reason":"missing_or_ambiguous_terminal_total","candidates":terminal or candidates}


def _sum_cumulative_before_ht(ws,first_row,ht_row,cumulative_col,dce_roles,ds,de):
    """Addition basique : cumul numérique, hors toutes les lignes récapitulatives."""
    total=0.0;contributors=[];excluded=[]
    for row in range(first_row,ht_row):
        value=number(ws.cell(row,cumulative_col).value)
        if value is None or abs(value)<=TOTAL_ABSOLUTE_TOLERANCE:continue
        label=_ht_row_label(ws,row,dce_roles,ds,de);marker=row_marker(ws,row)
        if _ht_summary_row(label,marker):
            excluded.append({"row":row,"label":label,"value":value});continue
        total+=value;contributors.append({"row":row,"label":label,"value":value})
    if not contributors:return {"status":"NOT_CHECKED","reason":"no_cumulative_contributor","contributors":[],"excluded":excluded}
    return {"status":"CERTAIN","amount":total,"contributors":contributors,"excluded":excluded}


def lot_ht_cumulative_check(ws,ribbon,stop,company,roles,start,end,dce_roles,ds,de,store,stats):
    """Seul contrôle vertical global actif : colonne cumul vers HT déclaré."""
    layout=_article_amount_layout(ws,roles,ribbon)
    if layout.get("status")!="CERTAIN" or layout.get("total") is None:
        stats["ht_not_checked_ambiguous_amount_layout"]+=1;return
    cumulative_col=layout["total"]
    found=_find_declared_ht_row(ws,ribbon,stop,dce_roles,ds,de,cumulative_col)
    if found["status"]!="CERTAIN":
        stats["ht_not_checked_"+found["reason"]]+=1;return
    ht=found["item"]
    first=header_limit(ws,ds,de,ribbon)+1
    summed=_sum_cumulative_before_ht(ws,first,ht["row"],cumulative_col,dce_roles,ds,de)
    if summed["status"]!="CERTAIN":
        stats["ht_not_checked_"+summed["reason"]]+=1;return
    declared=ht["value"];expected=summed["amount"]
    stats["ht_totals_checked"]+=1
    if ht["source"]=="IMPLICIT_TERMINAL_TOTAL":stats["ht_implicit_terminal_used"]+=1
    tolerance=max(TOTAL_ABSOLUTE_TOLERANCE,abs(expected)*TOTAL_RELATIVE_TOLERANCE)
    if abs(expected-declared)>tolerance:
        scope=(f"Addition de {len(summed['contributors'])} valeur(s) numériques de la colonne cumul confirmée. "
               f"{len(summed['excluded'])} ligne(s) récapitulative(s) exclue(s). "
               f"Cible HT : {'HT explicite' if ht['source']=='EXPLICIT_HT' else 'total général terminal interprété comme HT'}. "
               f"TVA et TTC non contrôlés.")
        add_total_error(store,company,ht["row"],ht["label"] or "Total HT",
                        "HT_TOTAL_ERROR","ht",expected,declared,ws,cumulative_col,
                        scope,roles,start,end)
        stats["ht_total_errors"]+=1

# LOCAL_HT_FROM_RIBBON_V3
_RIBBON_MONEY_RE=re.compile(r"([+-]?\d[\d .\u00a0]*[,.]\d{2})\s*€",re.I)

def _local_company_name(ribbon_label):
    text=str(ribbon_label or "").strip().replace("\r","\n")
    first=next((part.strip() for part in text.splitlines() if part.strip()),text)
    return re.sub(r"\s+[—-]\s+[+-]?\d[\d .\u00a0]*[,.]\d{2}\s*€.*$","",first).strip()


def _local_ribbon_amount(ribbon_label):
    matches=_RIBBON_MONEY_RE.findall(str(ribbon_label or ""))
    return number(matches[-1]) if matches else None


def _local_row_label(ws,row,roles,start,end):
    """Libellé financier lu exclusivement dans le bloc de l'entreprise."""
    for col in roles.get("designation",[]):
        if start<=col<=end:
            value=ws.cell(row,col).value
            if value not in (None,""):return str(value).strip()
    # Repli local uniquement : première chaîne du bloc, jamais le DCE.
    for col in range(start,end+1):
        value=ws.cell(row,col).value
        if isinstance(value,str) and value.strip():return value.strip()
    return ""


def _local_is_tax(label):
    text=_article_total_text(label)
    return bool(re.search(r"\b(?:tva|ttc)\b|toutes\s+taxes",text))


def _local_is_explicit_ht(label):
    text=_article_total_text(label)
    if not text or _local_is_tax(label):return False
    return bool(re.search(r"\bht\b|\bhors\s+taxes?\b",text)) and bool(re.search(r"\b(?:montant|total)\b",text))


def _local_is_summary(label,marker=""):
    marker=str(marker or "").strip().upper();text=norm(label)
    if marker in {"CH3","CH4","CH5","CH6","STOT","TOTHT","TVA","TOTTTC"}:return True
    if not text:return False
    return (
        text=="total" or text.startswith("total ") or
        text=="sous total" or text.startswith("sous total ") or
        text=="cumul" or text.startswith("cumul ") or
        text=="recapitulatif" or text.startswith("recapitulatif ") or
        text=="synthese" or text.startswith("synthese ") or
        _local_is_explicit_ht(label) or _local_is_tax(label)
    )


def _local_terminal_total(label):
    text=norm(label)
    if not text or _local_is_tax(label):return False
    if text in {"total","total general","montant total"}:return True
    return (text.startswith("total ") or text.startswith("montant ")) and any(x in text for x in ("general","lot","travaux","offre","marche"))


def _local_amount_close(left,right):
    if left is None or right is None:return False
    tolerance=max(TOTAL_ABSOLUTE_TOLERANCE,abs(right)*TOTAL_RELATIVE_TOLERANCE)
    return abs(left-right)<=tolerance


def _find_local_declared_ht(ws, ribbon, stop, ribbon_label, roles, start, end, cumulative_col):
    """Ruban -> HT local explicite ; sinon total terminal local certain.

    Regles de fiabilite :
    1. privilegier une ligne HT explicite correspondant au ruban ;
    2. accepter une ligne HT explicite unique meme si le ruban a ete altere ;
    3. en l'absence de HT explicite, accepter uniquement un total terminal local
       non fiscal et non ambigu ;
    4. ne jamais choisir arbitrairement entre plusieurs candidats.
    """
    ribbon_value = _local_ribbon_amount(ribbon_label)
    rows = []
    for row in range(ribbon + 1, stop):
        label = _local_row_label(ws, row, roles, start, end)
        value = number(ws.cell(row, cumulative_col).value)
        rows.append({
            "row": row,
            "label": label,
            "value": value,
            "marker": row_marker(ws, row),
        })

    explicit = [
        item for item in rows
        if item["value"] is not None and _local_is_explicit_ht(item["label"])
    ]
    matching_explicit = [
        item for item in explicit
        if _local_amount_close(item["value"], ribbon_value)
    ]
    if len(matching_explicit) == 1:
        item = matching_explicit[0].copy()
        item["source"] = "LOCAL_HT_MATCHING_RIBBON"
        item["ribbon_value"] = ribbon_value
        return {"status": "CERTAIN", "item": item}
    if len(matching_explicit) > 1:
        return {
            "status": "NOT_CHECKED",
            "reason": "multiple_local_ht_matching_ribbon",
            "candidates": matching_explicit,
        }
    if len(explicit) == 1:
        item = explicit[0].copy()
        item["source"] = "UNIQUE_LOCAL_HT"
        item["ribbon_value"] = ribbon_value
        return {"status": "CERTAIN", "item": item}
    if len(explicit) > 1:
        return {
            "status": "NOT_CHECKED",
            "reason": "multiple_local_ht_without_ribbon_match",
            "candidates": explicit,
        }

    terminal = [
        item for item in rows
        if item["value"] is not None
        and _local_terminal_total(item["label"])
        and not _local_is_tax(item["label"])
    ]
    matching_terminal = [
        item for item in terminal
        if _local_amount_close(item["value"], ribbon_value)
    ]
    if len(matching_terminal) == 1:
        item = matching_terminal[0].copy()
        item["source"] = "LOCAL_TERMINAL_TOTAL_MATCHING_RIBBON"
        item["ribbon_value"] = ribbon_value
        return {"status": "CERTAIN", "item": item}
    if len(matching_terminal) > 1:
        return {
            "status": "NOT_CHECKED",
            "reason": "multiple_local_terminal_totals_matching_ribbon",
            "candidates": matching_terminal,
        }
    if len(terminal) == 1:
        item = terminal[0].copy()
        item["source"] = "UNIQUE_LOCAL_TERMINAL_TOTAL"
        item["ribbon_value"] = ribbon_value
        return {"status": "CERTAIN", "item": item}
    if len(terminal) > 1:
        return {
            "status": "NOT_CHECKED",
            "reason": "multiple_local_terminal_totals",
            "candidates": terminal,
        }
    return {
        "status": "NOT_CHECKED",
        "reason": "no_local_declared_ht",
        "candidates": [],
    }

def _sum_local_cumulative(ws,first_row,ht_row,roles,start,end,cumulative_col):
    total=0.0;contributors=[];excluded=[]
    for row in range(first_row,ht_row):
        value=number(ws.cell(row,cumulative_col).value)
        if value is None or abs(value)<=TOTAL_ABSOLUTE_TOLERANCE:continue
        label=_local_row_label(ws,row,roles,start,end);marker=row_marker(ws,row)
        if _local_is_summary(label,marker):
            excluded.append({"row":row,"label":label,"value":value});continue
        contributors.append({"row":row,"label":label,"value":value});total+=value
    if not contributors:return {"status":"NOT_CHECKED","reason":"no_local_cumulative_contributor","contributors":[],"excluded":excluded}
    return {"status":"CERTAIN","amount":total,"contributors":contributors,"excluded":excluded}


# HT_CONTROL_AUDIT_VISUAL_V34
HT_SELECTED_BORDER = "548235"
HT_SCOPE_BORDER = "4472C4"
HT_EXCLUDED_FILL = "FCE4D6"
HT_EXCLUDED_BORDER = "C65911"


def _ht_comment_once(cell, message):
    from openpyxl.comments import Comment
    existing = cell.comment.text.splitlines() if cell.comment else []
    if message not in existing:
        text = ("\n".join(existing) + "\n" if existing else "") + message
        cell.comment = Comment(text, "AnalyseAO")


def _ht_border(cell, left=None, right=None, top=None, bottom=None):
    old = cell.border
    cell.border = Border(
        left=left if left is not None else old.left,
        right=right if right is not None else old.right,
        top=top if top is not None else old.top,
        bottom=bottom if bottom is not None else old.bottom,
    )


def _ht_label_cell(ws, row, roles, start, end):
    for col in roles.get("designation", []):
        if start <= col <= end and ws.cell(row, col).value not in (None, ""):
            return ws.cell(row, col)
    for col in range(start, end + 1):
        value = ws.cell(row, col).value
        if isinstance(value, str) and value.strip():
            return ws.cell(row, col)
    return None


def _ht_style_selected(ws, ht, summed, cumulative_col, roles, start, end):
    cell = ws.cell(ht["row"], cumulative_col)
    green = Side(style="thick", color=HT_SELECTED_BORDER)
    _ht_border(cell, green, green, green, green)
    rows = sorted({item["row"] for item in summed["contributors"]})
    message = (
        "TOTAL HT RETENU POUR LE CONTRÔLE — "
        f"Méthode : {ht['source']}. Colonne {cell.column_letter}. "
        f"Périmètre : lignes {rows[0]} à {rows[-1]}, {len(rows)} valeur(s) incluse(s), "
        f"{len(summed['excluded'])} ligne(s) récapitulative(s) exclue(s). "
        "TVA et TTC non contrôlés."
    )
    _ht_comment_once(cell, message)
    label = _ht_label_cell(ws, ht["row"], roles, start, end)
    if label is not None:
        _ht_comment_once(label, message)


def _ht_style_scope(ws, contributors, cumulative_col):
    rows = sorted({item["row"] for item in contributors})
    if not rows:
        return
    blue = Side(style="medium", color=HT_SCOPE_BORDER)
    for row in rows:
        cell = ws.cell(row, cumulative_col)
        _ht_border(
            cell,
            left=blue,
            right=blue,
            top=blue if row == rows[0] else None,
            bottom=blue if row == rows[-1] else None,
        )
    _ht_comment_once(
        ws.cell(rows[0], cumulative_col),
        f"PÉRIMÈTRE HT — {len(rows)} valeur(s) réellement additionnée(s) "
        f"dans la colonne {ws.cell(rows[0], cumulative_col).column_letter}, "
        f"entre les lignes {rows[0]} et {rows[-1]}.",
    )


def _ht_style_excluded(ws, excluded, cumulative_col, selected_coordinate):
    orange = Side(style="medium", color=HT_EXCLUDED_BORDER)
    for item in excluded:
        cell = ws.cell(item["row"], cumulative_col)
        current = str(getattr(cell.fill.fgColor, "rgb", "") or "").upper()[-6:]
        if current != "F4CCCC":
            cell.fill = PatternFill("solid", fgColor=HT_EXCLUDED_FILL)
            _ht_border(cell, orange, orange, orange, orange)
        _ht_comment_once(
            cell,
            "TOTAL EXCLU DU CONTRÔLE HT — Ligne récapitulative non additionnée "
            f"afin d'éviter un double comptage. Total retenu : {selected_coordinate}.",
        )


def _ht_reason(reason):
    labels = {
        "ambiguous_amount_layout": "cartographie des colonnes de montant ambiguë",
        "multiple_local_ht_matching_ribbon": "plusieurs lignes HT correspondent au ruban",
        "multiple_local_ht_without_ribbon_match": "plusieurs lignes HT explicites sont possibles",
        "multiple_local_terminal_totals_matching_ribbon": "plusieurs totaux terminaux correspondent au ruban",
        "multiple_local_terminal_totals": "plusieurs totaux terminaux sont possibles",
        "no_local_declared_ht": "aucune ligne HT locale certaine n'a été qualifiée",
        "no_local_cumulative_contributor": "aucune valeur contributrice certaine n'a été trouvée",
    }
    return labels.get(str(reason), str(reason).replace("_", " "))


def _ht_candidate_rows(ws, ribbon, stop, roles, start, end):
    result = []
    for row in range(ribbon + 1, stop):
        label = _local_row_label(ws, row, roles, start, end)
        if (_local_is_explicit_ht(label) or _local_terminal_total(label)) and not _local_is_tax(label):
            result.append({"row": row, "label": label})
    return result


def _ht_style_not_selected(ws, candidates, roles, start, end, cumulative_col, reason, store, company, stats):
    """Orange seulement les candidats identifiés, jamais toutes les colonnes montant."""
    orange = Side(style="medium", color=HT_EXCLUDED_BORDER)
    seen_rows = set()
    for item in candidates:
        row = item["row"]
        if row in seen_rows:
            continue
        seen_rows.add(row)
        label = item.get("label") or _local_row_label(ws, row, roles, start, end) or "Total HT possible"
        targets = []
        label_cell = _ht_label_cell(ws, row, roles, start, end)
        if label_cell is not None:
            targets.append(label_cell)
        if cumulative_col is not None:
            targets.append(ws.cell(row, cumulative_col))
        message = (
            "CANDIDAT HT NON RETENU — Le moteur n'a sélectionné aucun total avec certitude : "
            f"{_ht_reason(reason)}. Vérification manuelle requise."
        )
        for cell in dict.fromkeys(targets):
            current = str(getattr(cell.fill.fgColor, "rgb", "") or "").upper()[-6:]
            if current != "F4CCCC":
                cell.fill = PatternFill("solid", fgColor=HT_EXCLUDED_FILL)
                _ht_border(cell, orange, orange, orange, orange)
            _ht_comment_once(cell, message)
        add_issue(
            store, company, "TOTAL_NOT_CHECKED", row, label, message,
            ("ht_candidate_not_selected", row, reason),
            extra={"reason": reason, "control_status": "NOT_CHECKED"},
        )
        stats["ht_candidates_not_selected_marked"] += 1


def _ht_visualize_certain(ws, ht, summed, cumulative_col, roles, start, end):
    selected = ws.cell(ht["row"], cumulative_col).coordinate
    _ht_style_scope(ws, summed["contributors"], cumulative_col)
    _ht_style_excluded(ws, summed["excluded"], cumulative_col, selected)
    _ht_style_selected(ws, ht, summed, cumulative_col, roles, start, end)


def _ht_legend_sample(cell, border_color, fill_color=None):
    side = Side(style="medium", color=border_color)
    cell.value = "Exemple"
    cell.border = Border(left=side, right=side, top=side, bottom=side)
    if fill_color:
        cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _ensure_ht_audit_legend(ws):
    """Utilise C1:H1 sans insertion de ligne/colonne et sans modifier les largeurs."""
    if norm(ws["A1"].value) not in {"legende analyse", "legende", "legende d analyse"}:
        return False
    entries = (
        ("C1", "D1", "Total HT retenu pour le contrôle", HT_SELECTED_BORDER, None),
        ("E1", "F1", "Valeurs incluses dans le périmètre HT", HT_SCOPE_BORDER, None),
        ("G1", "H1", "Total exclu ou candidat HT non retenu", HT_EXCLUDED_BORDER, HT_EXCLUDED_FILL),
    )
    for sample_ref, text_ref, label, border_color, fill_color in entries:
        _ht_legend_sample(ws[sample_ref], border_color, fill_color)
        text = ws[text_ref]
        text.value = label
        text.font = Font(bold=True)
        text.alignment = Alignment(vertical="center", wrap_text=True)
    return True


def lot_ht_local_ribbon_check(ws,ribbon,stop,ribbon_label,roles,start,end,store,stats):
    company=_local_company_name(ribbon_label)
    layout=_article_amount_layout(ws,roles,ribbon)
    if layout.get("status")!="CERTAIN" or layout.get("total") is None:
        reason="ambiguous_amount_layout"
        stats["ht_not_checked_"+reason]+=1
        candidates=_ht_candidate_rows(ws,ribbon,stop,roles,start,end)
        # Sans colonne cumul certaine, seul le libellé candidat est marqué.
        _ht_style_not_selected(ws,candidates,roles,start,end,None,reason,store,company,stats)
        return
    cumulative_col=layout["total"]
    found=_find_local_declared_ht(ws,ribbon,stop,ribbon_label,roles,start,end,cumulative_col)
    if found["status"]!="CERTAIN":
        reason=found["reason"]
        stats["ht_not_checked_"+reason]+=1
        candidates=found.get("candidates") or _ht_candidate_rows(ws,ribbon,stop,roles,start,end)
        _ht_style_not_selected(ws,candidates,roles,start,end,cumulative_col,reason,store,company,stats)
        return
    ht=found["item"]
    first=header_limit(ws,start,end,ribbon)+1
    summed=_sum_local_cumulative(ws,first,ht["row"],roles,start,end,cumulative_col)
    if summed["status"]!="CERTAIN":
        reason=summed["reason"]
        stats["ht_not_checked_"+reason]+=1
        _ht_style_not_selected(ws,[ht],roles,start,end,cumulative_col,reason,store,company,stats)
        return
    expected=summed["amount"]
    declared=ht["value"]
    stats["ht_totals_checked"]+=1
    _ht_visualize_certain(ws,ht,summed,cumulative_col,roles,start,end)
    tolerance=max(TOTAL_ABSOLUTE_TOLERANCE,abs(expected)*TOTAL_RELATIVE_TOLERANCE)
    ht_delta=abs(expected-declared)
    _reliability_register(store,company,"HT_TOTAL",min(1.0,ht_delta/tolerance),ht_delta>tolerance)
    if abs(expected-declared)>tolerance:
        scope=(f"Addition de {len(summed['contributors'])} valeur(s) de la colonne cumul confirmée du bloc entreprise. "
               f"{len(summed['excluded'])} ligne(s) récapitulative(s) locale(s) exclue(s). "
               f"HT retenu : {ht['source']}. TVA et TTC non contrôlés.")
        add_total_error(store,company,ht["row"],ht["label"] or "Total HT","HT_TOTAL_ERROR","ht",
                        expected,declared,ws,cumulative_col,scope,roles,start,end)
        stats["ht_total_errors"]+=1


# CONTROL_RELIABILITY_SCORE_V35
RELIABILITY_COLOR_STOPS = (
    (0.0, "F8696B"), (70.0, "F4B183"), (85.0, "FFEB84"),
    (95.0, "A9D18E"), (100.0, "63BE7B"),
)


def _reliability_register(store, company, control_type, usage, failed=False):
    usage = min(1.0, max(0.0, float(usage)))
    registry = store.setdefault("reliability", {})
    data = registry.setdefault(company, {
        "checks": 0, "usage_sum": 0.0, "exact": 0,
        "tolerance_used": 0, "failed": 0, "by_type": {},
    })
    data["checks"] += 1
    data["usage_sum"] += usage
    if failed:
        data["failed"] += 1
    elif usage <= 1e-12:
        data["exact"] += 1
    else:
        data["tolerance_used"] += 1
    detail = data["by_type"].setdefault(control_type, {
        "checks": 0, "usage_sum": 0.0, "failed": 0,
    })
    detail["checks"] += 1
    detail["usage_sum"] += usage
    if failed:
        detail["failed"] += 1


def _reliability_finalize(store):
    result = {}
    for company, data in store.get("reliability", {}).items():
        checks = data["checks"]
        score = None if not checks else 100.0 * (1.0 - data["usage_sum"] / checks)
        result[company] = {
            "score": None if score is None else round(max(0.0, min(100.0, score)), 2),
            "checks": checks,
            "usage_sum": round(data["usage_sum"], 8),
            "exact": data["exact"],
            "tolerance_used": data["tolerance_used"],
            "failed": data["failed"],
            "by_type": {
                name: {"checks": values["checks"],
                       "usage_sum": round(values["usage_sum"], 8),
                       "failed": values["failed"]}
                for name, values in data["by_type"].items()
            },
        }
    return result


def _reliability_color(score):
    def rgb(value):
        return tuple(int(value[i:i + 2], 16) for i in (0, 2, 4))
    score = max(0.0, min(100.0, float(score)))
    for index in range(1, len(RELIABILITY_COLOR_STOPS)):
        low_score, low_color = RELIABILITY_COLOR_STOPS[index - 1]
        high_score, high_color = RELIABILITY_COLOR_STOPS[index]
        if score <= high_score:
            ratio = (score - low_score) / (high_score - low_score)
            values = [a + (b - a) * ratio for a, b in zip(rgb(low_color), rgb(high_color))]
            return "".join(f"{max(0, min(255, round(v))):02X}" for v in values)
    return RELIABILITY_COLOR_STOPS[-1][1]


def _reliability_comment(data):
    return (
        f"FIABILITE DES CONTROLES : {data['score']:.2f} %\n"
        "Le score part de 100 % et diminue selon la part des seuils de tolerance "
        "consommee par les controles arithmetiques executes.\n"
        f"Controles effectues : {data['checks']}\n"
        f"Concordances exactes : {data['exact']}\n"
        f"Controles ayant utilise une tolerance : {data['tolerance_used']}\n"
        f"Erreurs certaines : {data['failed']}\n"
        "Le score complete les alertes detaillees et ne les remplace pas."
    )


# RELIABILITY_SCORE_REAL_V36

def _reliability_recompute_all(ws, result):
    """Rejoue uniquement les contrôles numériques pour produire un score complet."""
    ribbon = result["ribbon"]
    blocks = result["blocks"]
    stop = summary_row(ws) or ws.max_row + 1
    _, ds, de = blocks[0]
    dce = role_map(ws, ds, de, ribbon, True)
    rel_store = {"reliability": {}}

    for label, start, end in blocks[1:]:
        company = str(label).splitlines()[0].strip()
        roles = role_map(ws, start, end, ribbon, False)
        pu_cols = roles.get("unit_price", [])
        qcomp = roles.get("quantity_company", [])
        qref = roles.get("quantity_reference", [])
        mcols = roles.get("amount", [])
        component_mcols = [col for col in mcols if not is_total_column(ws, col, ribbon)]
        common_pu = pu_cols[0] if len(pu_cols) == 1 else None

        for row in range(ribbon + 1, stop):
            if not article_row(ws, row, dce, ds, de):
                continue
            for index, mcol in enumerate(component_mcols):
                pcol = common_pu or (pu_cols[index] if index < len(pu_cols) else None)
                if pcol is None:
                    break
                qcol = qcomp[index] if index < len(qcomp) else (qref[index] if index < len(qref) else None)
                qraw = ws.cell(row, qcol).value if qcol else None
                if qraw in (None,) and index < len(qref):
                    qraw = ws.cell(row, qref[index]).value
                q = number(qraw)
                pu_raw = ws.cell(row, pcol).value
                pu = number(pu_raw)
                amount = number(ws.cell(row, mcol).value)
                active = (
                    (q is not None and abs(q) > 0.000001)
                    or (amount is not None and abs(amount) > 0.000001)
                    or bool(str(pu_raw or "").strip())
                )
                if not active or pu is None or pu == 0 or q is None or amount is None:
                    continue
                expected = q * pu
                delta = abs(expected - amount)
                scale = max(abs(expected), abs(amount))
                relative = (delta / scale) if scale > 0 else 0.0
                usage = min(
                    1.0,
                    max(
                        delta / CALC_ABSOLUTE_SIGNIFICANT,
                        relative / CALC_RELATIVE_SIGNIFICANT,
                    ),
                )
                failed = (
                    (amount == 0 and expected > 0.000001)
                    or delta >= CALC_ABSOLUTE_SIGNIFICANT
                    or relative >= CALC_RELATIVE_SIGNIFICANT
                )
                _reliability_register(rel_store, company, "CALCULATION", usage, failed)

            # Cumul horizontal : même qualification que horizontal_total_checks().
            dce_amounts = dce.get("amount", [])
            act_amounts = roles.get("amount", [])
            dce_totals = [col for col in dce_amounts if is_total_column(ws, col, ribbon)]
            act_totals = [col for col in act_amounts if is_total_column(ws, col, ribbon)]
            if len(dce_totals) == 1 and len(act_totals) == 1:
                dtotal, atotal = dce_totals[0], act_totals[0]
                ddetails = [col for col in dce_amounts if col < dtotal and col not in dce_totals]
                adetails = [col for col in act_amounts if col < atotal and col not in act_totals]
                if len(ddetails) >= 2 and len(ddetails) == len(adetails):
                    values = [number(ws.cell(row, col).value) for col in adetails]
                    declared = number(ws.cell(row, atotal).value)
                    if declared is not None and all(value is not None for value in values):
                        expected = sum(values)
                        tolerance = max(TOTAL_ABSOLUTE_TOLERANCE, abs(expected) * TOTAL_RELATIVE_TOLERANCE)
                        delta = abs(expected - declared)
                        _reliability_register(
                            rel_store, company, "ROW_AGGREGATE",
                            min(1.0, delta / tolerance), delta > tolerance,
                        )

        # Total HT : même sélection que le contrôle local V3.4.
        layout = _article_amount_layout(ws, roles, ribbon)
        if layout.get("status") == "CERTAIN" and layout.get("total") is not None:
            cumulative_col = layout["total"]
            found = _find_local_declared_ht(ws, ribbon, stop, label, roles, start, end, cumulative_col)
            if found.get("status") == "CERTAIN":
                ht = found["item"]
                first = header_limit(ws, start, end, ribbon) + 1
                summed = _sum_local_cumulative(ws, first, ht["row"], roles, start, end, cumulative_col)
                if summed.get("status") == "CERTAIN":
                    expected = summed["amount"]
                    declared = ht["value"]
                    tolerance = max(TOTAL_ABSOLUTE_TOLERANCE, abs(expected) * TOTAL_RELATIVE_TOLERANCE)
                    delta = abs(expected - declared)
                    _reliability_register(
                        rel_store, company, "HT_TOTAL",
                        min(1.0, delta / tolerance), delta > tolerance,
                    )

    return _reliability_finalize(rel_store)


def _insert_reliability_row_above_ribbons(ws, result):
    """Insère une ligne dédiée, sans modifier les propriétés des colonnes."""
    from openpyxl.comments import Comment
    ribbon = result.get("ribbon")
    blocks = result.get("blocks", [])
    reliability = result.get("reliability", {})
    if ribbon is None or len(blocks) < 2:
        return

    # Idempotence.
    if ribbon > 1 and str(ws.cell(ribbon - 1, blocks[0][1]).value or "").startswith("FIABILITE DES CONTROLES"):
        return

    merges = [
        (rg.min_row, rg.max_row, rg.min_col, rg.max_col)
        for rg in list(ws.merged_cells.ranges)
        if rg.min_row >= ribbon
    ]
    for min_row, max_row, min_col, max_col in merges:
        ws.unmerge_cells(
            start_row=min_row, end_row=max_row,
            start_column=min_col, end_column=max_col,
        )
    ws.insert_rows(ribbon, 1)
    for min_row, max_row, min_col, max_col in merges:
        ws.merge_cells(
            start_row=min_row + 1, end_row=max_row + 1,
            start_column=min_col, end_column=max_col,
        )

    # Libellé estimation.
    _, est_start, est_end = blocks[0]
    ws.merge_cells(start_row=ribbon, start_column=est_start, end_row=ribbon, end_column=est_end)
    est = ws.cell(ribbon, est_start, "FIABILITE DES CONTROLES")
    est.fill = PatternFill("solid", fgColor="595959")
    est.font = Font(color="FFFFFF", bold=True)
    est.alignment = Alignment(horizontal="center", vertical="center")

    for label, start, end in blocks[1:]:
        company = str(label).splitlines()[0].strip()
        data = reliability.get(company)
        ws.merge_cells(start_row=ribbon, start_column=start, end_row=ribbon, end_column=end)
        cell = ws.cell(ribbon, start)
        if not data or data.get("score") is None:
            cell.value = "Fiabilite : non calculable"
            color = "D9E1F2"
        else:
            score = data["score"]
            cell.value = f"Fiabilite des controles : {score:.2f} % ({data['checks']} controles)"
            color = _reliability_color(score)
            message = _reliability_comment(data)
            cell.comment = Comment(message, "AnalyseAO")
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(color="1F1F1F", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=True)

    ws.row_dimensions[ribbon].height = 24
    result["ribbon"] = ribbon + 1
    if result.get("summary_start") is not None and result["summary_start"] >= ribbon:
        result["summary_start"] += 1
    summary = result.get("summary")
    if isinstance(summary, dict) and summary.get("summary_row") is not None and summary["summary_row"] >= ribbon:
        summary["summary_row"] += 1
    if ws.freeze_panes:
        coordinate = ws.freeze_panes.coordinate if hasattr(ws.freeze_panes, "coordinate") else str(ws.freeze_panes)
        row, col = coordinate_to_tuple(coordinate)
        if row >= ribbon:
            ws.freeze_panes = f"{get_column_letter(col)}{row + 1}"


def _apply_reliability_to_ribbons(ws, result):
    """V3.6 : score sur une ligne dédiée au-dessus des rubans."""
    result["reliability"] = _reliability_recompute_all(ws, result)
    _insert_reliability_row_above_ribbons(ws, result)



def _ensure_reliability_legend(ws):
    if norm(ws["A1"].value) not in {"legende analyse", "legende", "legende d analyse"}:
        return False
    ws["I1"].value = "100 %"
    ws["I1"].fill = PatternFill("solid", fgColor=_reliability_color(100.0))
    ws["I1"].font = Font(bold=True)
    ws["I1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["J1"].value = "Fiabilite des controles : gradient rouge vers vert"
    ws["J1"].font = Font(bold=True)
    ws["J1"].alignment = Alignment(vertical="center", wrap_text=True)
    return True


def analyze(ws):
    ribbon,blocks=pure_blocks(ws)
    if ribbon is None or len(blocks)<2:raise ValueError("rubans estimation/entreprises introuvables")
    stop=summary_row(ws) or ws.max_row+1
    _,ds,de=blocks[0];dce=role_map(ws,ds,de,ribbon,True)
    company_data=[];roles_list=[];store={"seen":set(),"items":[]};stats=defaultdict(int)
    for label,start,end in blocks[1:]:
        company=str(label).splitlines()[0].strip();roles=role_map(ws,start,end,ribbon,False);roles_list.append(roles);company_data.append((company,start,end,roles))
        column_cardinality_issues(store,company,dce,roles)
        q_pairs=paired(dce.get("quantity_reference",[]),roles.get("quantity_company",[]) or roles.get("quantity_reference",[]))
        quantity_missing_rows=[];round_deltas=[]
        for row in range(ribbon+1,stop):
            if not article_row(ws,row,dce,ds,de):continue
            article=article_label(ws,row,dce,ds,de)
            for role,title,equivalent in (("designation","Désignation",designation_equivalent),("unit","Unité",unit_equivalent)):
                for dcol,acol,_ in paired(dce.get(role,[]),roles.get(role,[])):
                    dv,av=ws.cell(row,dcol).value,ws.cell(row,acol).value
                    if av in (None,""):continue
                    if not equivalent(dv,av):add_issue(store,company,"TEXT",row,article,f"{title} DCE « {display(dv)} » ; ACT « {display(av)} »",(role,dcol,acol))
            qref_cols=roles.get("quantity_reference",[])
            applicable=[(i,dcol,acol) for i,(dcol,acol,_) in enumerate(q_pairs) if number(ws.cell(row,dcol).value) not in (None,0)]
            if applicable and all(ws.cell(row,acol).value in (None,"") for _,_,acol in applicable):
                # Réponse normale de l'entreprise sur le DCE : agrégée une seule fois par entreprise (section 7).
                quantity_missing_rows.append(article)
            else:
                for pair_index,dcol,acol in applicable:
                    dv=ws.cell(row,dcol).value;av=ws.cell(row,acol).value
                    if av in (None,""):
                        if pair_index<len(qref_cols):av=ws.cell(row,qref_cols[pair_index]).value
                    dn,an=number(dv),number(av)
                    if dn is not None and an is not None:
                        stats["quantity_compared"]+=1
                        if abs(dn-an)>0.000001:add_issue(store,company,"DCE",row,article,f"Quantité DCE = {display(dv)} ; quantité ACT = {display(av)}",("q",pair_index))
                    elif norm(dv)!=norm(av):add_issue(store,company,"DCE",row,article,f"Quantité DCE « {display(dv)} » ; quantité ACT « {display(av)} »",("qt",pair_index))
            pu_cols=roles.get("unit_price",[]);qcomp=roles.get("quantity_company",[]);qref=roles.get("quantity_reference",[]);mcols=roles.get("amount",[])
            component_mcols=[c for c in mcols if not is_total_column(ws,c,ribbon)]
            common_pu=pu_cols[0] if len(pu_cols)==1 else None
            for index,mcol in enumerate(component_mcols):
                pcol=common_pu or (pu_cols[index] if index<len(pu_cols) else None)
                if pcol is None:
                    add_issue(store,company,"STRUCTURE",0,"Structure des colonnes","Structure ACT différente du DCE : correspondance quantité / PU / montant ambiguë ; calcul partiel.",("calc_structure",len(qcomp),len(pu_cols),len(component_mcols)))
                    break
                qcol=qcomp[index] if index<len(qcomp) else (qref[index] if index<len(qref) else None)
                qraw=ws.cell(row,qcol).value if qcol else None
                if qraw in (None,"") and index<len(qref):qraw=ws.cell(row,qref[index]).value
                q=number(qraw);pu_raw=ws.cell(row,pcol).value;pu=number(pu_raw);m=number(ws.cell(row,mcol).value)
                active=(q is not None and abs(q)>0.000001) or (m is not None and abs(m)>0.000001) or bool(str(pu_raw or "").strip())
                if not active:continue
                if pu is None:
                    add_issue(store,company,"UNVALUED",row,article,f"Prix unitaire à contrôler : valeur ACT « {display(pu_raw) or 'vide'} ».",("pu",pcol,index));continue
                if pu==0:
                    add_issue(store,company,"UNVALUED",row,article,"Prix unitaire nul à contrôler.",("pu0",pcol,index));continue
                if q is None or m is None:continue
                expected=q*pu;stats["calculations_checked"]+=1
                delta=abs(expected-m);scale=max(abs(expected),abs(m));relative=(delta/scale) if scale>0 else 0.0
                calc_usage=min(1.0,max(delta/CALC_ABSOLUTE_SIGNIFICANT,relative/CALC_RELATIVE_SIGNIFICANT))
                calc_failed=(m==0 and expected>0.000001) or delta>=CALC_ABSOLUTE_SIGNIFICANT or relative>=CALC_RELATIVE_SIGNIFICANT
                _reliability_register(store,company,"CALCULATION",calc_usage,calc_failed)
                if m==0 and expected>0.000001:
                    stats["calc_significant"]+=1
                    add_issue(store,company,"CALC",row,article,f"Calcul incohérent : {q:g} × {pu:g} = {expected:.2f} € ; montant ACT déclaré = {m:.2f} € ; écart = {delta:.2f} €.",(qcol,pcol,mcol))
                elif delta>=CALC_ABSOLUTE_SIGNIFICANT or relative>=CALC_RELATIVE_SIGNIFICANT:
                    stats["calc_significant"]+=1
                    add_issue(store,company,"CALC",row,article,f"Calcul incohérent : {q:g} × {pu:g} = {expected:.2f} € ; montant ACT déclaré = {m:.2f} € ; écart = {delta:.2f} € ({relative*100:.2f} %).",(qcol,pcol,mcol))
                elif delta>CALC_TECHNICAL_MINIMUM:
                    stats["calc_rounding_count"]+=1;stats["calc_rounding_sum"]+=delta;round_deltas.append(delta)
            horizontal_total_checks(ws,row,company,article,dce,roles,ribbon,store,stats)
        if quantity_missing_rows:
            add_issue(store,company,"INFO",0,"Quantités ACT non renseignées",f"Quantités ACT non renseignées sur {len(quantity_missing_rows)} article(s) ; quantités DCE conservées pour les contrôles.",("quantity_missing_summary",))
        if round_deltas:
            add_issue(store,company,"ROUND",0,"Écarts mineurs d'arrondi",f"{len(round_deltas)} écart(s) mineur(s) d'arrondi détecté(s) ; écart cumulé absolu = {sum(round_deltas):.2f} € ; aucune correction appliquée.",("rounding_summary",))
        lot_ht_local_ribbon_check(ws,ribbon,stop,label,roles,start,end,store,stats)
        # Contrôles TVA/TTC désactivés : contrôle vertical HT uniquement
    average=update_average_pu(ws,dce,roles_list,stop,ds,de)
    reliability=_reliability_finalize(store)
    return {"reliability":reliability,"version":VERSION,"companies":len(company_data),"issues":store["items"],"statistics":dict(stats),"average_pu":average,"blocks":blocks,"ribbon":ribbon,"summary_start":stop if stop<=ws.max_row else None}

def rewrite_summary(ws,result):
    start=result.get("summary_start")
    if start is None:raise ValueError("synthèse historique introuvable")
    for rg in list(ws.merged_cells.ranges):
        if rg.min_row>=start:ws.unmerge_cells(str(rg))
    ws.delete_rows(start,ws.max_row-start+1)
    issues=result["issues"];blocks=result["blocks"];by=defaultdict(lambda:defaultdict(list))
    for item in issues:by[item["company"]][item["category"]].append(f"{item['label']} : {item['message']}")
    sections=[("ERREURS DE CALCUL",("CALC","ROUND","CHAPTER_TOTAL_ERROR","HT_TOTAL_ERROR","VAT_TOTAL_ERROR","TTC_TOTAL_ERROR"),"F4CCCC","Pas d’erreur de calcul détectée."),("CUMULS DE LIGNE (HORS TOTAL GÉNÉRAL)",("ROW_AGGREGATE_ERROR",),"FFF2CC","Aucun écart de cumul de ligne détecté."),("ÉCARTS AVEC LE DCE",("DCE","STRUCTURE","INFO"),"FFF2CC","Aucun écart avec le DCE ni limite de comparaison détecté."),("POSTES NON VALORISÉS / VALEURS TEXTUELLES",("UNVALUED",),"F4CCCC","Tous les postes contrôlés sont valorisés numériquement."),("MODIFICATIONS DE TEXTE / UNITÉ",("TEXT",),"FFF2CC","Pas de modification de texte ou d’unité détectée.")]
    prepared=[];heights=[1]*5
    for label,_,_ in blocks[1:]:
        company=str(label).splitlines()[0].strip();entries=[]
        for idx,(heading,cats,color,empty) in enumerate(sections):
            details=[]
            for cat in cats:details.extend(by[company].get(cat,[]))
            # Déduplication d'affichage exacte au sein d'une entreprise/catégorie.
            details=list(dict.fromkeys(details)) or [empty];heights[idx]=max(heights[idx],1+len(details));entries.append((heading,details,color))
        prepared.append((company,entries))
    thin=Side(style="thin",color="808080");border=Border(left=thin,right=thin,top=thin,bottom=thin)
    for block,(company,entries) in zip(blocks[1:],prepared):
        _,cs,ce=block;ws.merge_cells(start_row=start,start_column=cs,end_row=start,end_column=ce);c=ws.cell(start,cs,f"ANALYSE DE L’OFFRE — {company}");c.fill=PatternFill("solid",fgColor="404040");c.font=Font(color="FFFFFF",bold=True);c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True);c.border=border;cursor=start+1
        for idx,(heading,details,color) in enumerate(entries):
            ws.merge_cells(start_row=cursor,start_column=cs,end_row=cursor,end_column=ce);h=ws.cell(cursor,cs,heading);h.fill=PatternFill("solid",fgColor=color);h.font=Font(bold=True);h.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True);h.border=border;cursor+=1
            for offset in range(heights[idx]-1):
                ws.merge_cells(start_row=cursor,start_column=cs,end_row=cursor,end_column=ce);text=details[offset] if offset<len(details) else "";d=ws.cell(cursor,cs,text);d.alignment=Alignment(vertical="top",wrap_text=True);d.border=border;ws.row_dimensions[cursor].height=30 if text else 15;cursor+=1
    _,es,ee=blocks[0];ws.merge_cells(start_row=start,start_column=es,end_row=start,end_column=ee);c=ws.cell(start,es,"SYNTHÈSE DES CONTRÔLES PAR ENTREPRISE");c.fill=PatternFill("solid",fgColor="595959");c.font=Font(color="FFFFFF",bold=True);c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True);c.border=border;ws.row_dimensions[start].height=28
    return {"summary_source":"DIRECT_V3","summary_row":start,"issue_count":len(issues)}

def process(ws):
    _ensure_ht_audit_legend(ws)
    _ensure_reliability_legend(ws)
    result=analyze(ws)
    result["summary"]=rewrite_summary(ws,result)
    _apply_reliability_to_ribbons(ws,result)
    return result

