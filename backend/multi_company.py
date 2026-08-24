# -*- coding: utf-8 -*-
from __future__ import annotations

from copy import copy
from pathlib import Path
from collections import defaultdict
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from workbook_catalog import inspect_workbook
from side_by_side import (
    business_sheet, scope_type, pair_sheets, technical_columns,
    visible_table_end, meaningful_rows, row_descriptor, align_sequences,
    copy_block_row, copy_dimensions, copy_merges,
    resolve_orphan_attachments, parent_attachment_map,
)

ORANGE = "FA4616"
SEPARATOR = "E7E6E6"


def normalize_lot(value):
    match = re.search(r"\d{1,3}", str(value or ""))
    return match.group(0).zfill(2) if match else ""


def safe_title(value):
    value = re.sub(r"[\\/*?:\[\]]", "_", str(value or "")).strip()
    return value[:31] or "Lot"


def sheet_infos(path, lot=None, role="ACT"):
    wb=load_workbook(path,data_only=True)
    catalog={entry.sheet:entry for entry in inspect_workbook(path,role)}
    result=[]
    for ws in wb.worksheets:
        entry=catalog.get(ws.title)
        if not entry or entry.category=="DOCUMENT": continue
        if lot and entry.lot and normalize_lot(entry.lot)!=normalize_lot(lot): continue
        result.append((ws,entry.category))
    result.sort(key=lambda item: catalog[item[0].title].score,reverse=True)
    return result

def select_sheet(infos, scope, preferred_name=None):
    candidates = [ws for ws, item_scope in infos if item_scope == scope]
    if not candidates:
        return None
    if preferred_name:
        exact = [ws for ws in candidates if ws.title == preferred_name]
        if exact:
            return exact[0]
    return candidates[0]


def descriptors(ws):
    if ws is None:
        return [], 0, None, None
    marker, article_id = technical_columns(ws)
    end = visible_table_end(ws, marker)
    rows = meaningful_rows(ws, end, marker)
    return [row_descriptor(ws, row, end, marker, article_id) for row in rows], end, marker, article_id


def alignment_map(dce_desc, act_desc):
    """Retourne ACT par ligne DCE et lignes ACT supplémentaires avant/après les ancres DCE."""
    paired = align_sequences(dce_desc, act_desc)
    by_dce = {}
    extras_before = defaultdict(list)
    trailing = []
    next_dce = None
    # Parcours inverse pour rattacher une ligne ACT supplémentaire à la prochaine ligne DCE.
    for left, right in reversed(paired):
        if left is not None:
            next_dce = left["row"]
            if right is not None:
                by_dce[left["row"]] = right["row"]
        elif right is not None:
            if next_dce is None:
                trailing.append(right["row"])
            else:
                extras_before[next_dce].append(right["row"])
    for key in list(extras_before):
        extras_before[key].reverse()
    trailing.reverse()
    return by_dce, extras_before, trailing


def is_meaningful_extra(item, previous_key):
    """Garde-fou avant d'injecter un extra ACT dans le plan commun.

    Les lignes décoratives (zéros, orphelines rattachées) sont déjà écartées en
    amont par meaningful_rows/_numeric_orphan/resolve_orphan_attachments. Cette
    fonction ne rejette donc que les cas résiduels sans contenu visible, ou une
    ligne strictement identique au poste ACT qui vient juste d'être inséré pour
    la même entreprise (titre ou article dupliqué dans le même intervalle).
    """
    if item is None:
        return False
    if not str(item.get("visible") or "").strip():
        return False
    if previous_key is not None and item.get("key") == previous_key:
        return False
    return True


def build_row_plan(dce_desc, company_maps):
    """Ordre DCE conservé, avec union des ajouts ACT de toutes les entreprises."""
    plan = []
    seen_extra = set()
    lookup = [{d["row"]: d for d in mapping.get("desc", [])} for mapping in company_maps]
    last_key = [None] * len(company_maps)

    def _try_append(company_index, act_row):
        key = (company_index, act_row)
        if key in seen_extra:
            return
        item = lookup[company_index].get(act_row)
        if not is_meaningful_extra(item, last_key[company_index]):
            return
        plan.append((None, company_index, act_row))
        seen_extra.add(key)
        last_key[company_index] = item.get("key")

    for dce in dce_desc:
        dce_row = dce["row"]
        for company_index, mapping in enumerate(company_maps):
            for act_row in mapping["extras"].get(dce_row, []):
                _try_append(company_index, act_row)
        plan.append((dce_row, None, None))
    for company_index, mapping in enumerate(company_maps):
        for act_row in mapping["trailing"]:
            _try_append(company_index, act_row)
    return plan


def copy_merged_ranges(src, dst, row_map, start_col, end_col):
    if src is None:
        return
    copy_merges(src, dst, row_map, start_col, end_col)


def add_header(ws, row, start, end, text):
    if end < start:
        return
    ws.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)
    cell = ws.cell(row, start, text)
    cell.fill = PatternFill("solid", fgColor=ORANGE)
    cell.font = Font(color="FFFFFF", bold=True, size=12)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def build_scope_sheet(final, lot, scope, dce_ws, companies):
    title = safe_title(f"LOT {lot} {scope}")
    ws = final.create_sheet(title)
    ws.sheet_view.showGridLines = False

    dce_desc, dce_end, dce_mark, _ = descriptors(dce_ws)
    dattach = parent_attachment_map(resolve_orphan_attachments(dce_ws, dce_end, dce_mark)) if dce_ws else {}
    company_data = []
    for company in companies:
        act_ws = company["sheet"]
        act_desc, act_end, act_mark, _ = descriptors(act_ws)
        attach = parent_attachment_map(resolve_orphan_attachments(act_ws, act_end, act_mark))
        by_dce, extras, trailing = alignment_map(dce_desc, act_desc) if dce_ws else ({}, {}, [d["row"] for d in act_desc])
        company_data.append({**company, "desc": act_desc, "end": act_end, "by_dce": by_dce, "extras": extras, "trailing": trailing, "attach": attach})

    plan = build_row_plan(dce_desc, company_data) if dce_ws else [
        (None, ci, row) for ci, company in enumerate(company_data) for row in company["trailing"]
    ]

    starts = []
    current_col = 1
    if dce_ws:
        dce_start = current_col
        current_col += dce_end
        starts.append(("ESTIMATION", dce_start, dce_end, dce_ws))
        current_col += 2
    else:
        dce_start = None

    for company in company_data:
        start = current_col
        starts.append((company["enterprise"], start, company["end"], company["sheet"]))
        current_col += company["end"] + 2

    for label, start, width, source in starts:
        add_header(ws, 1, start, start + width - 1, label)
        copy_dimensions(source, ws, start, width)
    ws.row_dimensions[1].height = 24

    out_row = 2
    dce_row_map = {}
    company_row_maps = [dict() for _ in company_data]
    for dce_row, extra_company, extra_act_row in plan:
        if dce_row is not None:
            if dce_ws:
                copy_block_row(dce_ws, dce_row, ws, out_row, dce_start, dce_end, dattach.get(dce_row))
                dce_row_map[dce_row] = out_row
            for ci, company in enumerate(company_data):
                act_row = company["by_dce"].get(dce_row)
                if act_row:
                    start = next(item[1] for item in starts if item[0] == company["enterprise"])
                    copy_block_row(company["sheet"], act_row, ws, out_row, start, company["end"], company["attach"].get(act_row))
                    company_row_maps[ci][act_row] = out_row
        else:
            company = company_data[extra_company]
            start = next(item[1] for item in starts if item[0] == company["enterprise"])
            copy_block_row(company["sheet"], extra_act_row, ws, out_row, start, company["end"], company["attach"].get(extra_act_row))
            company_row_maps[extra_company][extra_act_row] = out_row
        out_row += 1

    if dce_ws:
        copy_merged_ranges(dce_ws, ws, dce_row_map, dce_start, dce_end)
    for ci, company in enumerate(company_data):
        start = next(item[1] for item in starts if item[0] == company["enterprise"])
        copy_merged_ranges(company["sheet"], ws, company_row_maps[ci], start, company["end"])

    # Séparateurs gris entre ESTIMATION et les entreprises.
    block_ends = [(start + width - 1) for _, start, width, _ in starts]
    for block_end in block_ends[:-1]:
        for col in range(block_end + 1, block_end + 3):
            ws.column_dimensions[get_column_letter(col)].width = 2.5
            for row in range(1, ws.max_row + 1):
                ws.cell(row, col).fill = PatternFill("solid", fgColor=SEPARATOR)

    ws.freeze_panes = "A2"
    ws.page_setup.orientation = "landscape"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    return ws


def unique_title(workbook, base, current_sheet=None):
    used={s for s in workbook.sheetnames if current_sheet is None or s!=current_sheet.title}
    if base not in used: return base
    i=2
    while True:
        suffix=f" {i}"; candidate=base[:31-len(suffix)]+suffix
        if candidate not in used: return candidate
        i+=1


def build_multi_company_workbook(jobs, output_path, temp_dir=None):
    """Une feuille par lot, catégorie et occurrence; toutes les entreprises côte à côte."""
    grouped=defaultdict(list)
    for job in jobs: grouped[normalize_lot(job["lot"])].append(job)
    final=Workbook(); final.remove(final.active)
    category_order=["BASE","OPTION","VARIANTE","COMPLEMENT","REMISE","RABAIS","PLUS-VALUE","MOINS-VALUE"]
    for lot in sorted(grouped):
        lot_jobs=sorted(grouped[lot],key=lambda j:j["enterprise"].casefold())
        dce_path=lot_jobs[0]["dce_path"]
        dce_infos=sheet_infos(dce_path,lot,"DCE")
        act_infos=[{"job":j,"infos":sheet_infos(j["act_path"],lot,"ACT")} for j in lot_jobs]
        categories=[]
        present={scope for _,scope in dce_infos}
        for item in act_infos: present.update(scope for _,scope in item["infos"])
        for c in category_order:
            if c in present: categories.append(c)
        categories.extend(sorted(present-set(categories)-{"DOCUMENT","RECAP","ANNEXE"}))
        for category in categories:
            dce_sheets=[ws for ws,scope in dce_infos if scope==category]
            if category=="BASE" and dce_sheets:
                dce_sheets=dce_sheets[:1]
            act_sheets={item["job"]["enterprise"]:[ws for ws,scope in item["infos"] if scope==category] for item in act_infos}
            if category=="BASE":
                act_sheets={enterprise:(sheets[:1] if sheets else []) for enterprise,sheets in act_sheets.items()}
            occurrences=max([len(dce_sheets)]+[len(v) for v in act_sheets.values()] or [0])
            for occurrence in range(occurrences):
                dce_ws=dce_sheets[occurrence] if occurrence<len(dce_sheets) else None
                companies=[]
                for item in act_infos:
                    seq=act_sheets[item["job"]["enterprise"]]
                    if occurrence<len(seq): companies.append({"enterprise":item["job"]["enterprise"],"sheet":seq[occurrence]})
                # STABILITY_FIX_AFTER_RESTORE_V1
                # Une feuille comparative exige un DCE et au moins un ACT compatible.
                if dce_ws is None or not companies: continue
                sheet=build_scope_sheet(final,lot,category,dce_ws,companies)
                suffix=f" {occurrence+1}" if occurrences>1 else ""
                sheet.title=unique_title(final,safe_title(f"LOT {lot} {category}{suffix}"),sheet)
    final.calculation.fullCalcOnLoad=True; final.calculation.forceFullCalc=True; final.calculation.calcMode="auto"
    if not final.sheetnames:
        raise ValueError("Aucune feuille comparative produite : aucun couple DCE/ACT compatible.")
    output_path=Path(output_path); output_path.parent.mkdir(parents=True,exist_ok=True); final.save(output_path); return output_path

# TCE_VIRTUAL_SHEETS_V22
_v22_legacy_build_multi_company_workbook = build_multi_company_workbook


def _v22_exact_sheet(path, sheet_name):
    workbook = load_workbook(path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"Feuille demandée introuvable : {sheet_name}")
    return workbook[sheet_name]


def build_multi_company_workbook(jobs, output_path, temp_dir=None):
    """Jobs TCE par feuilles exactes, ancien pipeline inchangé sinon."""
    if not any(job.get("virtual_sheet_job") for job in jobs):
        return _v22_legacy_build_multi_company_workbook(jobs, output_path, temp_dir)

    grouped = defaultdict(list)
    for job in jobs:
        grouped[(normalize_lot(job["lot"]), job.get("category") or "BASE")].append(job)

    final = Workbook()
    final.remove(final.active)
    opened = []
    try:
        category_order = [
            "BASE", "BPU", "OPTION", "VARIANTE", "COMPLEMENT",
            "REMISE", "RABAIS", "PLUS-VALUE", "MOINS-VALUE",
        ]
        order = {name: index for index, name in enumerate(category_order)}
        for (lot, category), lot_jobs in sorted(
            grouped.items(), key=lambda item: (item[0][0], order.get(item[0][1], 999), item[0][1])
        ):
            lot_jobs = sorted(lot_jobs, key=lambda job: job["enterprise"].casefold())
            dce_book = load_workbook(lot_jobs[0]["dce_path"], data_only=True)
            opened.append(dce_book)
            dce_name = lot_jobs[0].get("dce_sheet")
            if not dce_name or dce_name not in dce_book.sheetnames:
                raise ValueError(f"Lot {lot} {category} : feuille DCE explicite introuvable")
            dce_ws = dce_book[dce_name]

            companies = []
            for job in lot_jobs:
                act_book = load_workbook(job["act_path"], data_only=True)
                opened.append(act_book)
                act_name = job.get("act_sheet")
                if not act_name or act_name not in act_book.sheetnames:
                    raise ValueError(
                        f"Lot {lot} {category} - {job['enterprise']} : feuille ACT explicite introuvable"
                    )
                companies.append({"enterprise": job["enterprise"], "sheet": act_book[act_name]})

            if not companies:
                continue
            sheet = build_scope_sheet(final, lot, category, dce_ws, companies)
            sheet.title = unique_title(final, safe_title(f"LOT {lot} {category}"), sheet)

        if not final.sheetnames:
            raise ValueError("Aucune feuille comparative produite : aucun couple DCE/ACT compatible.")
        final.calculation.fullCalcOnLoad = True
        final.calculation.forceFullCalc = True
        final.calculation.calcMode = "auto"
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        final.save(output_path)
        return output_path
    finally:
        for workbook in opened:
            try:
                workbook.close()
            except Exception:
                pass

