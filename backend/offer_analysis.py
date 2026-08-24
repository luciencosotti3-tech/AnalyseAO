# Légende analyse en haut à gauche V2
# -*- coding: utf-8 -*-
from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Any
import math
import re
import unicodedata

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill

FORMULA_TOLERANCE = 0.02
DELTA_THRESHOLD = 0.30

CHEAPEST_FILL = PatternFill("solid", fgColor="C6E0B4")
MOST_EXPENSIVE_FILL = PatternFill("solid", fgColor="FFC7CE")
QUANTITY_FILL = PatternFill("solid", fgColor="FFF2CC")
TEXT_FILL = PatternFill("solid", fgColor="FFF2CC")
UNIT_FILL = PatternFill("solid", fgColor="FFE699")
ERROR_FILL = PatternFill("solid", fgColor="F4CCCC")
ABOVE_FONT = Font(color="FF0000", bold=True)
BELOW_FONT = Font(color="0070C0", bold=True)
CHANGE_FONT = Font(color="C00000", bold=True)


@dataclass
class Block:
    label: str
    start: int
    end: int
    roles: dict[str, list[int]]
    total_ht: float | None = None


def norm(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKD", str(value))
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = text.casefold().replace("²", "2")
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9%]+", " ", text)).strip()


def to_float(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = str(value).strip().replace("€", "").replace("\u00a0", "").replace(" ", "")
    if not text:
        return None
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def block_ranges(ws):
    merged = [r for r in ws.merged_cells.ranges if r.min_row == 1 and r.max_row == 1]
    ranges = []
    for item in sorted(merged, key=lambda r: r.min_col):
        label = str(ws.cell(1, item.min_col).value or "").strip()
        excluded = {"rouge", "orange", "i", "legende analyse", "classement ht"}
        if label and norm(label) not in excluded and label != "!":
            ranges.append((label, item.min_col, item.max_col))
    return ranges


def detect_roles(ws, start, end):
    roles = {key: [] for key in ("reference", "designation", "unit", "quantity", "unit_price", "amount", "vat")}
    for col in range(start, end + 1):
        fragments = []
        for row in range(2, min(ws.max_row, 12) + 1):
            value = ws.cell(row, col).value
            if value is not None:
                fragments.append(norm(value))
        header = " ".join(fragments)
        last = next((frag for frag in reversed(fragments) if frag), "")
        if "designation" in header:
            roles["designation"].append(col)
        elif last in {"u", "unite"} or " unite " in f" {header} ":
            roles["unit"].append(col)
        elif any(token in header for token in ("quantite", "qte", "qty")):
            roles["quantity"].append(col)
        elif last in {"p u", "pu", "prix unitaire"} or "prix unitaire" in header:
            roles["unit_price"].append(col)
        elif "montant" in header and "tva" not in header:
            roles["amount"].append(col)
        elif "tva" in header:
            roles["vat"].append(col)
        elif col == start:
            roles["reference"].append(col)
    return roles


def detect_blocks(ws):
    blocks = []
    for label, start, end in block_ranges(ws):
        blocks.append(Block(label, start, end, detect_roles(ws, start, end)))
    return blocks


def value(ws, row, block, role, prefer_last=True):
    cols = block.roles.get(role, [])
    ordered = reversed(cols) if prefer_last else cols
    for col in ordered:
        raw = ws.cell(row, col).value
        if raw is not None and str(raw).strip() != "":
            return raw, col
    return None, (cols[-1] if cols else None)


def is_article(ws, row, estimation):
    designation, _ = value(ws, row, estimation, "designation")
    label = norm(designation)
    if not label or any(token in label for token in ("total", "tva", "ttc", "montant ht")):
        return False
    numeric = any(to_float(value(ws, row, estimation, role)[0]) is not None for role in ("quantity", "unit_price", "amount"))
    ref, _ = value(ws, row, estimation, "reference")
    return numeric or bool(re.search(r"\d", str(ref or "")))


def explicit_total_ht(ws, block):
    amount_cols = block.roles.get("amount", [])
    if not amount_cols:
        return None
    candidates = []
    for row in range(2, ws.max_row + 1):
        designation, _ = value(ws, row, block, "designation")
        label = norm(designation)
        if ("montant ht" in label or "total ht" in label) and not any(x in label for x in ("option", "tva", "ttc")):
            amount, _ = value(ws, row, block, "amount")
            number = to_float(amount)
            if number is not None:
                candidates.append(number)
    return candidates[-1] if candidates else None


def add_comment(cell, message):
    current = cell.comment.text + "\n" if cell.comment else ""
    cell.comment = Comment(current + message, "AnalyseAO")



# ---------- Légende conditionnelle analyse V1.1 ----------
def _rgb(value):
    if value is None:
        return ""
    rgb = getattr(value, "rgb", None)
    if not rgb:
        return ""
    return str(rgb).upper()[-6:]


def _analysis_legend_entries(ws):
    """Détecte uniquement les styles métier réellement présents."""
    present = set()
    for row in ws.iter_rows():
        for cell in row:
            fill = _rgb(cell.fill.fgColor)
            font = _rgb(cell.font.color)
            if fill == "C6E0B4":
                present.add("CHEAPEST")
            elif fill == "FFC7CE":
                present.add("EXPENSIVE")
            elif fill in {"FFF2CC", "FFE699"}:
                present.add("CHANGED")
            elif fill == "F4CCCC":
                present.add("ERROR")
            if cell.font.bold and font == "FF0000":
                present.add("ABOVE_30")
            elif cell.font.bold and font == "0070C0":
                present.add("BELOW_30")

    definitions = [
        ("CHEAPEST", "C6E0B4", None, "Montant le moins cher parmi les entreprises"),
        ("EXPENSIVE", "FFC7CE", None, "Montant le plus cher parmi les entreprises"),
        ("CHANGED", "FFF2CC", "C00000", "Désignation, unité ou quantité modifiée"),
        ("ERROR", "F4CCCC", None, "Erreur de calcul ou poste non valorisé"),
        ("ABOVE_30", None, "FF0000", "Montant supérieur de plus de 30 % à l'estimation"),
        ("BELOW_30", None, "0070C0", "Montant inférieur de plus de 30 % à l'estimation"),
    ]
    return [entry for entry in definitions if entry[0] in present]



def _write_analysis_legend(ws, start_col=None):
    """Insère une légende conditionnelle dans le coin supérieur gauche.

    Le tableau existant est décalé de trois colonnes sans modifier les lignes.
    Les fusions et largeurs de colonnes sont déplacées avec les données.
    """
    entries = _analysis_legend_entries(ws)
    if not entries:
        return

    # Idempotence : une seconde analyse ne décale pas une nouvelle fois la feuille.
    if str(ws["A1"].value or "").strip().upper() in {"LÉGENDE ANALYSE", "LEGENDE ANALYSE"}:
        return

    shift = 3
    old_max_col = ws.max_column
    old_max_row = ws.max_row
    get_column_letter = __import__(
        "openpyxl.utils.cell", fromlist=["get_column_letter"]
    ).get_column_letter
    coordinate_to_tuple = __import__(
        "openpyxl.utils.cell", fromlist=["coordinate_to_tuple"]
    ).coordinate_to_tuple

    merged = [
        (item.min_row, item.max_row, item.min_col, item.max_col)
        for item in list(ws.merged_cells.ranges)
    ]
    for item in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(item))

    dimensions = {}
    for col in range(1, old_max_col + 1):
        letter = get_column_letter(col)
        dim = ws.column_dimensions[letter]
        dimensions[col] = {
            "width": dim.width,
            "hidden": dim.hidden,
            "bestFit": dim.bestFit,
            "outlineLevel": dim.outlineLevel,
        }

    end_letter = get_column_letter(old_max_col)
    ws.move_range(
        f"A1:{end_letter}{old_max_row}",
        rows=0,
        cols=shift,
        translate=True,
    )

    for min_row, max_row, min_col, max_col in merged:
        ws.merge_cells(
            start_row=min_row,
            end_row=max_row,
            start_column=min_col + shift,
            end_column=max_col + shift,
        )

    for old_col, values in dimensions.items():
        target = ws.column_dimensions[get_column_letter(old_col + shift)]
        target.width = values["width"]
        target.hidden = values["hidden"]
        target.bestFit = values["bestFit"]
        target.outlineLevel = values["outlineLevel"]

    # Espace réservé à la légende.
    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 3

    if ws.freeze_panes:
        coordinate = (
            ws.freeze_panes.coordinate
            if hasattr(ws.freeze_panes, "coordinate")
            else str(ws.freeze_panes)
        )
        row, col = coordinate_to_tuple(coordinate)
        ws.freeze_panes = f"{get_column_letter(col + shift)}{row}"

    descriptions = {
        "CHEAPEST": ("C6E0B4", None, "Montant le moins cher parmi les entreprises"),
        "EXPENSIVE": ("FFC7CE", None, "Montant le plus cher parmi les entreprises"),
        "CHANGED": ("FFF2CC", None, "Désignation, unité ou quantité modifiée"),
        "ERROR": ("F4CCCC", None, "Erreur de calcul ou poste non valorisé"),
        "ABOVE_30": (None, "FF0000", "Montant supérieur de plus de 30 % à l'estimation"),
        "BELOW_30": (None, "0070C0", "Montant inférieur de plus de 30 % à l'estimation"),
    }
    order = ["CHEAPEST", "EXPENSIVE", "CHANGED", "ERROR", "ABOVE_30", "BELOW_30"]

    ws.merge_cells("A1:B1")
    title = ws["A1"]
    title.value = "LÉGENDE ANALYSE"
    title.fill = PatternFill("solid", fgColor="595959")
    title.font = Font(color="FFFFFF", bold=True)

    row = 2
    for key in order:
        if key not in entries:
            continue
        fill_color, font_color, description = descriptions[key]
        sample = ws.cell(row, 1, "Exemple")
        label = ws.cell(row, 2, description)
        if fill_color:
            sample.fill = PatternFill("solid", fgColor=fill_color)
        if font_color:
            sample.font = Font(color=font_color, bold=True)
        label.font = Font(color="000000")
        row += 1

def analyse_sheet(ws):
    blocks = detect_blocks(ws)
    if len(blocks) < 2:
        return {"sheet": ws.title, "status": "IGNORED", "issues": 0}
    estimation = blocks[0]
    enterprises = blocks[1:]
    for block in enterprises:
        block.total_ht = explicit_total_ht(ws, block)
    enterprises.sort(key=lambda b: (float("inf") if b.total_ht is None else b.total_ht, b.label.casefold()))

    issues = 0
    for row in range(2, ws.max_row + 1):
        if not is_article(ws, row, estimation):
            continue
        est_designation, _ = value(ws, row, estimation, "designation")
        est_unit, _ = value(ws, row, estimation, "unit")
        est_quantity, _ = value(ws, row, estimation, "quantity")
        est_amount, _ = value(ws, row, estimation, "amount")
        est_amount_n = to_float(est_amount)

        row_amounts = []
        for block in enterprises:
            ent_designation, des_col = value(ws, row, block, "designation")
            ent_unit, unit_col = value(ws, row, block, "unit")
            ent_quantity, qty_col = value(ws, row, block, "quantity")
            ent_pu, pu_col = value(ws, row, block, "unit_price")
            ent_amount, amount_col = value(ws, row, block, "amount")
            q, pu, amount = to_float(ent_quantity), to_float(ent_pu), to_float(ent_amount)

            if norm(est_designation) and norm(ent_designation) and norm(est_designation) != norm(ent_designation):
                cell = ws.cell(row, des_col)
                cell.fill, cell.font = copy(TEXT_FILL), copy(CHANGE_FONT)
                add_comment(cell, "Désignation modifiée par rapport à l'estimation")
                issues += 1
            if norm(est_unit) and norm(ent_unit) and norm(est_unit) != norm(ent_unit):
                cell = ws.cell(row, unit_col)
                cell.fill, cell.font = copy(UNIT_FILL), copy(CHANGE_FONT)
                add_comment(cell, "Unité modifiée par rapport à l'estimation")
                issues += 1
            eq = to_float(est_quantity)
            if eq is not None and q is not None and abs(eq - q) > FORMULA_TOLERANCE:
                cell = ws.cell(row, qty_col)
                cell.fill, cell.font = copy(QUANTITY_FILL), copy(CHANGE_FONT)
                add_comment(cell, f"Quantité estimation = {eq:g}; quantité entreprise = {q:g}")
                issues += 1
            if q is not None and q != 0 and (pu is None or pu == 0) and (amount is None or amount == 0):
                cell = ws.cell(row, pu_col or block.start)
                cell.fill = copy(ERROR_FILL)
                add_comment(cell, "Prix unitaire manquant / poste non valorisé")
                issues += 1
            if q is not None and pu is not None and amount is not None:
                expected = q * pu
                if abs(expected - amount) > FORMULA_TOLERANCE:
                    cell = ws.cell(row, amount_col)
                    cell.fill = copy(ERROR_FILL)
                    add_comment(cell, f"Calcul incohérent : {q:g} x {pu:g} = {expected:.2f}, montant annoncé = {amount:.2f}")
                    issues += 1
            if est_amount_n and est_amount_n > 0 and amount and amount > 0 and amount_col:
                if amount > est_amount_n * 1.30:
                    ws.cell(row, amount_col).font = copy(ABOVE_FONT)
                    add_comment(ws.cell(row, amount_col), "Montant supérieur de plus de 30 % à l'estimation")
                    issues += 1
                elif amount < est_amount_n * 0.70:
                    ws.cell(row, amount_col).font = copy(BELOW_FONT)
                    add_comment(ws.cell(row, amount_col), "Montant inférieur de plus de 30 % à l'estimation")
                    issues += 1
            if amount is not None and amount_col:
                row_amounts.append((amount, amount_col))

        positive = [(amount, col) for amount, col in row_amounts if amount > 0]
        if positive:
            min_value = min(amount for amount, _ in positive)
            max_value = max(amount for amount, _ in positive)
            if min_value != max_value:
                for amount, col in positive:
                    if amount == min_value:
                        ws.cell(row, col).fill = copy(CHEAPEST_FILL)
                    if amount == max_value:
                        ws.cell(row, col).fill = copy(MOST_EXPENSIVE_FILL)

    # Classement HT dans un petit bloc à droite, sans feuille Synthèse.
    start_col = ws.max_column + 2
    ws.cell(1, start_col, "CLASSEMENT HT")
    ws.cell(1, start_col).font = Font(bold=True, color="FFFFFF")
    ws.cell(1, start_col).fill = PatternFill("solid", fgColor="1F4E78")
    for index, block in enumerate(enterprises, 1):
        ws.cell(index + 1, start_col, index)
        ws.cell(index + 1, start_col + 1, block.label)
        ws.cell(index + 1, start_col + 2, block.total_ht)
        ws.cell(index + 1, start_col + 2).number_format = '#,##0.00 [$€-fr-FR]'

    # La légende est placée après le classement et uniquement si une catégorie
    # de marquage est réellement présente dans la feuille.
    _write_analysis_legend(ws, start_col + 5)
    return {"sheet": ws.title, "status": "OK", "issues": issues}




# Mise en page analyse triee + prix rubans V1
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter, coordinate_to_tuple

# L'ancienne légende ajoutait de grandes colonnes à gauche. Elle est neutralisée.
def _write_analysis_legend(ws, start_col=None):
    return


def _money_fr(value):
    if value is None:
        return "Prix HT non détecté"
    return f"{value:,.2f} €".replace(",", "X").replace(".", ",").replace("X", " ")


def _remove_old_analysis_extras(ws):
    """Supprime les anciennes légendes et le classement latéral."""
    # Ancienne légende verticale issue du patch précédent.
    if str(ws["A1"].value or "").strip().upper() in {"LÉGENDE ANALYSE", "LEGENDE ANALYSE"}:
        merged = list(ws.merged_cells.ranges)
        for rg in merged:
            if rg.max_col <= 3:
                ws.unmerge_cells(str(rg))
        ws.delete_cols(1, 3)
        if ws.freeze_panes:
            row, col = coordinate_to_tuple(
                ws.freeze_panes.coordinate if hasattr(ws.freeze_panes, "coordinate") else str(ws.freeze_panes)
            )
            ws.freeze_panes = f"{get_column_letter(max(1, col - 3))}{row}"

    # Tableau CLASSEMENT HT devenu inutile puisque les blocs sont réellement triés.
    rank_ranges = []
    for rg in list(ws.merged_cells.ranges):
        if rg.min_row == 1 and rg.max_row == 1:
            label = norm(ws.cell(1, rg.min_col).value)
            if label == "classement ht":
                rank_ranges.append((rg.min_col, rg.max_col, str(rg)))
    for start, end, address in sorted(rank_ranges, reverse=True):
        ws.unmerge_cells(address)
        ws.delete_cols(start, end - start + 1)


def _reorder_offer_blocks(ws, blocks):
    """Réordonne les blocs complets sans perdre styles, commentaires ou fusions."""
    if len(blocks) < 2:
        return blocks

    estimation = blocks[0]
    enterprises = blocks[1:]
    for block in enterprises:
        block.total_ht = explicit_total_ht(ws, block)
    enterprises.sort(key=lambda block: (
        float("inf") if block.total_ht is None else block.total_ht,
        block.label.casefold(),
    ))
    ordered = [estimation] + enterprises

    old_start = min(block.start for block in blocks)
    max_row = ws.max_row
    temp_col = ws.max_column + 5
    stored = []

    # Mémorise et retire les fusions entièrement incluses dans les blocs.
    merges = []
    for rg in list(ws.merged_cells.ranges):
        owner = next((block for block in blocks if rg.min_col >= block.start and rg.max_col <= block.end), None)
        if owner is not None:
            merges.append((owner, rg.min_row, rg.max_row, rg.min_col - owner.start, rg.max_col - owner.start))
            ws.unmerge_cells(str(rg))

    # Déplace d'abord chaque bloc dans une zone temporaire non chevauchante.
    cursor = temp_col
    for block in ordered:
        width = block.end - block.start + 1
        dimensions = []
        for col in range(block.start, block.end + 1):
            dim = ws.column_dimensions[get_column_letter(col)]
            dimensions.append((dim.width, dim.hidden, dim.bestFit, dim.outlineLevel))
        ws.move_range(
            f"{get_column_letter(block.start)}1:{get_column_letter(block.end)}{max_row}",
            rows=0, cols=cursor - block.start, translate=True,
        )
        stored.append((block, cursor, width, dimensions))
        cursor += width

    # Replace les blocs dans l'ordre demandé, juste après l'estimation.
    cursor = old_start
    relocated = []
    for block, source, width, dimensions in stored:
        ws.move_range(
            f"{get_column_letter(source)}1:{get_column_letter(source + width - 1)}{max_row}",
            rows=0, cols=cursor - source, translate=True,
        )
        for offset, values in enumerate(dimensions):
            dim = ws.column_dimensions[get_column_letter(cursor + offset)]
            dim.width, dim.hidden, dim.bestFit, dim.outlineLevel = values
        block.start, block.end = cursor, cursor + width - 1
        block.roles = detect_roles(ws, block.start, block.end)
        relocated.append(block)
        cursor += width

    # Recrée les fusions à la nouvelle position de chaque bloc.
    for owner, min_row, max_row_merge, rel_min, rel_max in merges:
        ws.merge_cells(
            start_row=min_row, end_row=max_row_merge,
            start_column=owner.start + rel_min, end_column=owner.start + rel_max,
        )
    return relocated


def _decorate_ribbons(ws, blocks):
    """Ajoute le total HT au ruban de chaque entreprise."""
    if not blocks:
        return
    estimation = blocks[0]
    est_cell = ws.cell(1, estimation.start)
    est_cell.value = "ESTIMATION"
    est_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    enterprises = blocks[1:]
    priced = [block for block in enterprises if block.total_ht is not None]
    cheapest = priced[0] if priced else None
    most_expensive = priced[-1] if priced else None
    for block in enterprises:
        cell = ws.cell(1, block.start)
        cell.value = f"{block.label}\n{_money_fr(block.total_ht)}"
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.font = Font(color="FFFFFF", bold=True)
        if block is cheapest:
            fill = PatternFill("solid", fgColor="548235")
        elif block is most_expensive:
            fill = PatternFill("solid", fgColor="C00000")
        else:
            fill = PatternFill("solid", fgColor="4472C4")
        for col in range(block.start, block.end + 1):
            ws.cell(1, col).fill = fill
    ws.row_dimensions[1].height = 34


def _compact_legend(ws, entries):
    """Ajoute une légende horizontale compacte sur trois lignes au-dessus du tableau."""
    items = [
        ("CHEAPEST", "C6E0B4", None, "Prix unitaire le moins cher"),
        ("EXPENSIVE", "FFC7CE", None, "Prix unitaire le plus cher"),
        ("CHANGED", "FFF2CC", None, "Donnée modifiée"),
        ("ERROR", "F4CCCC", None, "Erreur ou valeur absente"),
        ("ABOVE_30", None, "FF0000", "Écart > +30 % / estimation"),
        ("BELOW_30", None, "0070C0", "Écart < -30 % / estimation"),
    ]
    visible = [item for item in items if item[0] in entries]
    if not visible:
        return

    old_merges = [(rg.min_row, rg.max_row, rg.min_col, rg.max_col) for rg in list(ws.merged_cells.ranges)]
    for rg in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rg))
    ws.insert_rows(1, 3)
    for min_row, max_row, min_col, max_col in old_merges:
        ws.merge_cells(
            start_row=min_row + 3, end_row=max_row + 3,
            start_column=min_col, end_column=max_col,
        )

    ws.merge_cells("A1:F1")
    title = ws["A1"]
    title.value = "LÉGENDE D'ANALYSE"
    title.fill = PatternFill("solid", fgColor="404040")
    title.font = Font(color="FFFFFF", bold=True)
    title.alignment = Alignment(horizontal="left", vertical="center")

    thin = Side(style="thin", color="BFBFBF")
    for index, (_, fill_color, font_color, text) in enumerate(visible):
        row = 2 + index // 3
        col = 1 + (index % 3) * 2
        sample = ws.cell(row, col, "Aa")
        label = ws.cell(row, col + 1, text)
        if fill_color:
            sample.fill = PatternFill("solid", fgColor=fill_color)
        sample.font = Font(color=font_color or "000000", bold=True)
        sample.alignment = Alignment(horizontal="center", vertical="center")
        label.alignment = Alignment(vertical="center", wrap_text=True)
        sample.border = label.border = Border(bottom=thin)
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22

    if ws.freeze_panes:
        row, col = coordinate_to_tuple(
            ws.freeze_panes.coordinate if hasattr(ws.freeze_panes, "coordinate") else str(ws.freeze_panes)
        )
        ws.freeze_panes = f"{get_column_letter(col)}{row + 3}"


_analyse_sheet_before_sorted_layout = analyse_sheet


def analyse_sheet(ws):
    # L'analyse métier existante applique d'abord couleurs et commentaires.
    result = _analyse_sheet_before_sorted_layout(ws)
    if result.get("status") == "IGNORED":
        return result

    entries = _analysis_legend_entries(ws)
    _remove_old_analysis_extras(ws)
    blocks = detect_blocks(ws)
    if len(blocks) < 2:
        return result
    blocks = _reorder_offer_blocks(ws, blocks)
    _decorate_ribbons(ws, blocks)
    _compact_legend(ws, entries)
    result["physical_order"] = [block.label for block in blocks]
    return result




# Correctif final legende originale sans classement lateral V5
def _find_ribbon_row(ws):
    for row in range(1, min(ws.max_row, 20) + 1):
        for cell in ws[row]:
            if norm(cell.value) == "estimation":
                return row
    return None


def _delete_top_legend_variants(ws):
    """Retire toute ancienne variante placée avant le ruban ESTIMATION."""
    ribbon_row = _find_ribbon_row(ws)
    if ribbon_row and ribbon_row > 1:
        for rg in list(ws.merged_cells.ranges):
            if rg.min_row < ribbon_row:
                ws.unmerge_cells(str(rg))
        ws.delete_rows(1, ribbon_row - 1)


def _delete_side_analysis_tables(ws):
    """Supprime CLASSEMENT HT et toute ancienne légende latérale."""
    targets = []
    for rg in list(ws.merged_cells.ranges):
        if rg.min_row != 1 or rg.max_row != 1:
            continue
        label = norm(ws.cell(1, rg.min_col).value)
        if label in {"classement ht", "legende analyse", "legende d analyse"}:
            # CLASSEMENT HT occupe trois colonnes. La première légende latérale
            # occupe deux colonnes. On conserve une colonne blanche de sécurité.
            width = 3 if label == "classement ht" else 2
            targets.append((rg.min_col, max(rg.max_col, rg.min_col + width - 1), str(rg)))

    for start, end, address in sorted(targets, key=lambda item: item[0], reverse=True):
        if address in {str(rg) for rg in ws.merged_cells.ranges}:
            ws.unmerge_cells(address)
        ws.delete_cols(start, end - start + 1)

    # Sécurité si le bandeau n'était pas fusionné. Recherche sur les premières
    # lignes, car des variantes de légende peuvent avoir décalé le ruban.
    plain_targets = []
    for row in range(1, min(ws.max_row, 15) + 1):
        for col in range(1, ws.max_column + 1):
            label = norm(ws.cell(row, col).value)
            if label == "classement ht":
                plain_targets.append((col, 3))
            elif label in {"legende analyse", "legende d analyse"}:
                plain_targets.append((col, 2))
    for col, width in sorted(set(plain_targets), reverse=True):
        ws.delete_cols(col, width)


def _insert_original_locked_legend(ws, entries):
    """Première version de la légende, au début de la feuille et figée."""
    definitions = [
        ("CHEAPEST", "C6E0B4", None, "Montant le moins cher parmi les entreprises"),
        ("EXPENSIVE", "FFC7CE", None, "Montant le plus cher parmi les entreprises"),
        ("CHANGED", "FFF2CC", "C00000", "Désignation, unité ou quantité modifiée"),
        ("ERROR", "F4CCCC", None, "Erreur de calcul ou poste non valorisé"),
        ("ABOVE_30", None, "FF0000", "Montant supérieur de plus de 30 % à l'estimation"),
        ("BELOW_30", None, "0070C0", "Montant inférieur de plus de 30 % à l'estimation"),
    ]
    visible = [item for item in definitions if item[0] in entries]
    if not visible:
        return

    count = 1 + len(visible)
    merges = [(rg.min_row, rg.max_row, rg.min_col, rg.max_col) for rg in list(ws.merged_cells.ranges)]
    for rg in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rg))
    ws.insert_rows(1, count)
    for min_row, max_row, min_col, max_col in merges:
        ws.merge_cells(
            start_row=min_row + count,
            end_row=max_row + count,
            start_column=min_col,
            end_column=max_col,
        )

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    title = ws.cell(1, 1, "LÉGENDE ANALYSE")
    title.fill = PatternFill("solid", fgColor="595959")
    title.font = Font(color="FFFFFF", bold=True)
    title.alignment = Alignment(horizontal="center", vertical="center")

    for row, (_, fill_color, font_color, description) in enumerate(visible, 2):
        sample = ws.cell(row, 1, "Exemple")
        label = ws.cell(row, 2, description)
        if fill_color:
            sample.fill = PatternFill("solid", fgColor=fill_color)
        sample.font = Font(color=font_color or "000000", bold=bool(font_color))
        sample.alignment = Alignment(horizontal="center", vertical="center")
        label.alignment = Alignment(vertical="center")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 52
    # La légende et le ruban juste dessous restent visibles pendant le défilement.
    ws.freeze_panes = f"A{count + 2}"


_analyse_sheet_before_final_legend_cleanup = analyse_sheet


def analyse_sheet(ws):
    result = _analyse_sheet_before_final_legend_cleanup(ws)
    if result.get("status") == "IGNORED":
        return result

    entries = _analysis_legend_entries(ws)
    _delete_top_legend_variants(ws)
    _delete_side_analysis_tables(ws)
    _insert_original_locked_legend(ws, entries)
    result["side_ranking_removed"] = True
    return result


def analyse_workbook(source, output):
    source, output = Path(source), Path(output)
    wb = load_workbook(source, data_only=False)
    report = []
    for ws in wb.worksheets:
        if norm(ws.title).startswith("lot "):
            report.append(analyse_sheet(ws))
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return output, report


# FORCE_DEFAULT_LEGEND_ACTIVE_V7
# Dernière définition du module : elle prime sur tous les anciens wrappers.
_v7_previous_analyse_sheet = analyse_sheet


def _v7_remove_existing_top_legend(ws):
    """Retire une légende déjà placée au-dessus du ruban ESTIMATION."""
    if norm(ws["A1"].value) not in {"legende analyse", "legende", "legende d analyse"}:
        return
    estimation_row = None
    for row in range(2, min(ws.max_row, 20) + 1):
        if any(norm(cell.value) == "estimation" for cell in ws[row]):
            estimation_row = row
            break
    if estimation_row is None:
        return
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row < estimation_row:
            ws.unmerge_cells(str(merged))
    ws.delete_rows(1, estimation_row - 1)


def _v7_insert_default_legend(ws):
    """Insère toujours la première légende complète, puis la verrouille."""

    _v7_remove_existing_top_legend(ws)

    # Légende PAR DÉFAUT : aucune dépendance à la détection des couleurs.
    definitions = [
        ("C6E0B4", None, "Montant le moins cher parmi les entreprises"),
        ("FFC7CE", None, "Montant le plus cher parmi les entreprises"),
        ("FFF2CC", "C00000", "Désignation, unité ou quantité modifiée"),
        ("F4CCCC", None, "Erreur de calcul ou poste non valorisé"),
        (None, "FF0000", "Montant supérieur de plus de 30 % à l'estimation"),
        (None, "0070C0", "Montant inférieur de plus de 30 % à l'estimation"),
    ]
    definitions[3] = ("F4CCCC", None, "Erreur de calcul, total incohérent ou poste non valorisé")
    inserted = 1 + len(definitions)

    merges = [
        (merged.min_row, merged.max_row, merged.min_col, merged.max_col)
        for merged in list(ws.merged_cells.ranges)
    ]
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))
    ws.insert_rows(1, inserted)
    for min_row, max_row, min_col, max_col in merges:
        ws.merge_cells(
            start_row=min_row + inserted,
            end_row=max_row + inserted,
            start_column=min_col,
            end_column=max_col,
        )

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    title = ws.cell(1, 1, "LÉGENDE ANALYSE")
    title.fill = PatternFill("solid", fgColor="595959")
    title.font = Font(color="FFFFFF", bold=True)
    title.alignment = Alignment(horizontal="center", vertical="center")

    for row, (fill_color, font_color, description) in enumerate(definitions, 2):
        sample = ws.cell(row, 1, "Exemple")
        label = ws.cell(row, 2, description)
        if fill_color:
            sample.fill = PatternFill("solid", fgColor=fill_color)
        sample.font = Font(color=font_color or "000000", bold=bool(font_color))
        sample.alignment = Alignment(horizontal="center", vertical="center")
        label.alignment = Alignment(vertical="center")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 52
    # 7 lignes de légende + le ruban en ligne 8 restent visibles.
    ws.freeze_panes = "A9"


def analyse_sheet(ws):
    result = _v7_previous_analyse_sheet(ws)
    if result.get("status") != "IGNORED":
        _v7_insert_default_legend(ws)
        result["default_legend_v7"] = True
    return result


# COMPANY_ANALYSIS_SUMMARY_V8
_company_summary_previous_analyse_sheet = analyse_sheet


def _v8_ribbon_blocks(ws):
    """Retourne les blocs d'offres depuis le ruban ESTIMATION/entreprises."""
    ribbon_row=None
    for row in range(1,min(ws.max_row,20)+1):
        if any(norm(cell.value)=="estimation" for cell in ws[row]):
            ribbon_row=row; break
    if ribbon_row is None:
        return None,[]
    ranges=[]
    for merged in ws.merged_cells.ranges:
        if merged.min_row==ribbon_row and merged.max_row==ribbon_row:
            label=str(ws.cell(ribbon_row,merged.min_col).value or '').strip()
            if label:
                ranges.append((label,merged.min_col,merged.max_col))
    return ribbon_row,sorted(ranges,key=lambda item:item[1])


def _v8_designation_column(ws,start,end,ribbon_row):
    for row in range(ribbon_row+1,min(ws.max_row,ribbon_row+15)+1):
        for col in range(start,end+1):
            if norm(ws.cell(row,col).value)=="designation":
                return col
    return min(end,start+1)


def _v8_article_label(ws,row,start,end,designation_col):
    designation=str(ws.cell(row,designation_col).value or '').strip()
    reference=''
    for col in range(start,min(end,designation_col-1)+1):
        raw=ws.cell(row,col).value
        if raw not in (None,''):
            reference=str(raw).strip(); break
    text=' '.join(part for part in (reference,designation) if part)
    return text or f"Ligne {row}"


def _v8_parse_ribbon_total(label):
    lines=str(label or '').splitlines()
    if len(lines)<2: return None
    text=lines[-1].replace('€','').replace('\u00a0','').replace(' ','').replace(',','.')
    try: return float(text)
    except ValueError: return None


def _v8_collect_company(ws,block,ribbon_row):
    label,start,end=block
    designation_col=_v8_designation_column(ws,start,end,ribbon_row)
    result={'name':label.splitlines()[0].strip(),'errors':[],'qty':[],'unvalued':[],'text':[],'delta':0.0}
    seen={key:set() for key in ('errors','qty','unvalued','text')}
    calc_pattern=re.compile(r'Calcul incohérent\s*:\s*(.+?)\s*=\s*([0-9., ]+)\s*,\s*montant annoncé\s*=\s*([0-9., ]+)',re.I)
    for row in range(ribbon_row+1,ws.max_row+1):
        article=_v8_article_label(ws,row,start,end,designation_col)
        comments=[]
        for col in range(start,end+1):
            comment=ws.cell(row,col).comment
            if comment: comments.extend(line.strip() for line in comment.text.splitlines() if line.strip())
        for message in comments:
            low=norm(message)
            category=None
            if low.startswith('calcul incoherent'):
                category='errors'
                match=calc_pattern.search(message)
                if match:
                    def number(value):
                        return float(value.replace(' ','').replace(',','.'))
                    try: result['delta']+=number(match.group(2))-number(match.group(3))
                    except ValueError: pass
            elif low.startswith('quantite estimation'):
                category='qty'
            elif 'prix unitaire manquant' in low or 'poste non valorise' in low:
                category='unvalued'
            elif low.startswith('designation modifiee') or low.startswith('unite modifiee'):
                category='text'
            if category:
                detail=f"{article} : {message}"
                if detail not in seen[category]:
                    seen[category].add(detail); result[category].append(detail)
    total=_v8_parse_ribbon_total(label)
    result['total']=total
    result['corrected_total']=(total+result['delta']) if total is not None else None
    return result


def _v8_money(value):
    if value is None:return 'non calculable'
    return f"{value:,.2f} € HT".replace(',','X').replace('.',',').replace('X',' ')


def _v8_summary_lines(data):
    sections=[]
    errors=data['errors']
    if errors:
        correction=(f"Montant HT corrigé indicatif : {_v8_money(data['corrected_total'])} "
                    f"(écart cumulé : {_v8_money(data['delta'])})")
        sections.append(('ERREURS DE CALCUL',errors+[correction],'F4CCCC'))
    else:
        sections.append(('ERREURS DE CALCUL',['Pas d’erreur de calcul détectée.'],'E2F0D9'))
    sections.append(('MODIFICATIONS DE QUANTITÉS',data['qty'] or ['Pas de modification de quantité détectée.'],'FFF2CC' if data['qty'] else 'E2F0D9'))
    sections.append(('POSTES NON VALORISÉS',data['unvalued'] or ['Tous les postes contrôlés sont valorisés.'],'F4CCCC' if data['unvalued'] else 'E2F0D9'))
    sections.append(('MODIFICATIONS DE TEXTE / UNITÉ',data['text'] or ['Pas de modification de texte ou d’unité détectée.'],'FFF2CC' if data['text'] else 'E2F0D9'))
    return sections


def _v8_write_company_summaries(ws):
    ribbon_row,blocks=_v8_ribbon_blocks(ws)
    if ribbon_row is None or len(blocks)<2:return
    enterprises=blocks[1:]
    analyses=[_v8_collect_company(ws,block,ribbon_row) for block in enterprises]
    section_sets=[_v8_summary_lines(data) for data in analyses]
    heights=[]
    for section_index in range(4):
        heights.append(1+max(len(sections[section_index][1]) for sections in section_sets))
    start_row=ws.max_row+2
    title_row=start_row
    thin=Side(style='thin',color='808080')
    border=Border(left=thin,right=thin,top=thin,bottom=thin)
    for block,data,sections in zip(enterprises,analyses,section_sets):
        _,start,end=block
        ws.merge_cells(start_row=title_row,start_column=start,end_row=title_row,end_column=end)
        cell=ws.cell(title_row,start,f"ANALYSE DE L’OFFRE — {data['name']}")
        cell.fill=PatternFill('solid',fgColor='404040'); cell.font=Font(color='FFFFFF',bold=True)
        cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); cell.border=border
        cursor=title_row+1
        for section_index,(heading,details,color) in enumerate(sections):
            ws.merge_cells(start_row=cursor,start_column=start,end_row=cursor,end_column=end)
            head=ws.cell(cursor,start,heading); head.fill=PatternFill('solid',fgColor=color); head.font=Font(bold=True)
            head.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); head.border=border
            cursor+=1
            max_details=heights[section_index]-1
            for index in range(max_details):
                ws.merge_cells(start_row=cursor,start_column=start,end_row=cursor,end_column=end)
                detail=details[index] if index<len(details) else ''
                c=ws.cell(cursor,start,detail); c.alignment=Alignment(vertical='top',wrap_text=True); c.border=border
                ws.row_dimensions[cursor].height=30 if detail else 15
                cursor+=1
    # Estimation : cartouche explicatif en face des analyses entreprises.
    _,est_start,est_end=blocks[0]
    ws.merge_cells(start_row=title_row,start_column=est_start,end_row=title_row,end_column=est_end)
    c=ws.cell(title_row,est_start,'SYNTHÈSE DES CONTRÔLES PAR ENTREPRISE')
    c.fill=PatternFill('solid',fgColor='595959'); c.font=Font(color='FFFFFF',bold=True)
    c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=border
    ws.row_dimensions[title_row].height=28


_v8_summary_previous_analyse_sheet=analyse_sheet


def analyse_sheet(ws):
    result=_v8_summary_previous_analyse_sheet(ws)
    if result.get('status')!='IGNORED':
        _v8_write_company_summaries(ws)
        result['company_summaries_v8']=True
    return result


# FIX_TOTALS_AND_AGGREGATES_V9
_v9_old_ribbon_blocks = _v8_ribbon_blocks
_v9_old_collect_company = _v8_collect_company
_v9_previous_analyse_sheet = analyse_sheet


def _v9_money(value):
    return f"{value:,.2f} €".replace(',', 'X').replace('.', ',').replace('X', ' ')


def _v9_find_offer_total(ws,start,end,ribbon_row):
    """Cherche le total métier du lot, pas un simple sous-total de bâtiment."""
    candidates=[]
    stop=ws.max_row
    for row in range(ribbon_row+1,ws.max_row+1):
        # Ne pas lire les synthèses ajoutées sous le tableau.
        if any(str(ws.cell(row,col).value or '').startswith('ANALYSE DE L’OFFRE') for col in range(start,end+1)):
            stop=row-1; break
    for row in range(ribbon_row+1,stop+1):
        texts=[str(ws.cell(row,col).value or '').strip() for col in range(start,end+1)]
        joined=' '.join(texts).casefold()
        priority=0
        if 'total' in joined and ('lot' in joined or 'gros œuvre' in joined or 'gros oeuvre' in joined): priority=4
        elif 'montant ht du lot' in joined or 'montant h.t. du lot' in joined: priority=4
        elif joined.startswith('total '): priority=3
        elif 'montant total' in joined: priority=2
        if not priority: continue
        nums=[]
        for col in range(start,end+1):
            value=ws.cell(row,col).value
            if isinstance(value,(int,float)) and not isinstance(value,bool): nums.append(float(value))
        if nums: candidates.append((priority,row,nums[-1]))
    if candidates:
        candidates.sort(key=lambda item:(item[0],item[1]),reverse=True)
        return candidates[0][2]
    return None


def _v8_ribbon_blocks(ws):
    """Version V9 : répare les prix de ruban avant la synthèse V8."""
    ribbon_row,blocks=_v9_old_ribbon_blocks(ws)
    if ribbon_row is None:return ribbon_row,blocks
    repaired=[]
    for index,(label,start,end) in enumerate(blocks):
        if index==0:
            repaired.append((label,start,end));continue
        name=str(label or '').splitlines()[0].strip()
        total=_v9_find_offer_total(ws,start,end,ribbon_row)
        if total is None:
            # Conserve le prix fiable déjà calculé par le module.
            lines=str(label or '').splitlines()
            new_label=label if len(lines)>1 else f"{name}\nPrix HT non détecté"
        else:
            new_label=f"{name}\n{_v9_money(total)}"
        ws.cell(ribbon_row,start).value=new_label
        repaired.append((new_label,start,end))
    return ribbon_row,repaired


def _v9_is_aggregate_total_row(ws,row,start,end,ribbon_row):
    """Le dernier montant d'un bloc multi-bâtiments est une somme, pas Q×PU."""
    # Plusieurs en-têtes Montant HT signifient plusieurs bâtiments + total final.
    amount_headers=0
    for header_row in range(ribbon_row+1,min(ws.max_row,ribbon_row+15)+1):
        amount_headers=max(amount_headers,sum(1 for col in range(start,end+1) if norm(ws.cell(header_row,col).value)=='montant ht'))
    if amount_headers<2:return False
    # La cellule commentée doit être la dernière valeur numérique du bloc sur la ligne.
    numeric_cols=[col for col in range(start,end+1) if isinstance(ws.cell(row,col).value,(int,float)) and not isinstance(ws.cell(row,col).value,bool)]
    return bool(numeric_cols and numeric_cols[-1]==end)


def _v8_collect_company(ws,block,ribbon_row):
    """Version V9 : ignore les faux calculs sur les colonnes TOTAL multi-bâtiments."""
    _,start,end=block
    hidden=[]
    for row in range(ribbon_row+1,ws.max_row+1):
        if not _v9_is_aggregate_total_row(ws,row,start,end,ribbon_row):continue
        cell=ws.cell(row,end)
        if cell.comment and any(norm(line).startswith('calcul incoherent') for line in cell.comment.text.splitlines()):
            hidden.append((cell,cell.comment))
            other=[line for line in cell.comment.text.splitlines() if not norm(line).startswith('calcul incoherent')]
            if other:
                cell.comment=Comment('\n'.join(other),hidden[-1][1].author or 'AnalyseAO')
            else: cell.comment=None
    try:
        return _v9_old_collect_company(ws,block,ribbon_row)
    finally:
        for cell,comment in hidden:cell.comment=comment


def _v9_clean_aggregate_false_errors(ws):
    """Retire des cellules Excel les faux commentaires/rouges sur les totaux agrégés."""
    ribbon_row,blocks=_v8_ribbon_blocks(ws)
    if ribbon_row is None:return
    for _,start,end in blocks[1:]:
        for row in range(ribbon_row+1,ws.max_row+1):
            if not _v9_is_aggregate_total_row(ws,row,start,end,ribbon_row):continue
            cell=ws.cell(row,end)
            if not cell.comment:continue
            lines=cell.comment.text.splitlines()
            kept=[line for line in lines if not norm(line).startswith('calcul incoherent')]
            if len(kept)==len(lines):continue
            cell.comment=Comment('\n'.join(kept),cell.comment.author or 'AnalyseAO') if kept else None
            # Le rouge F4CCCC était appliqué uniquement à cause de ce faux calcul.
            if str(getattr(cell.fill.fgColor,'rgb','')).upper().endswith('F4CCCC'):
                cell.fill=PatternFill(fill_type=None)


def analyse_sheet(ws):
    result=_v9_previous_analyse_sheet(ws)
    if result.get('status')!='IGNORED':
        _v9_clean_aggregate_false_errors(ws)
        result['totals_and_aggregates_v9']=True
    return result


# RIBBON_SINGLE_LINE_AND_TOTAL_COLUMN_V10
_v10_previous_collect_company = _v8_collect_company
_v10_previous_analyse_sheet = analyse_sheet


def _v10_amount_header_columns(ws,start,end,ribbon_row):
    """Colonnes Montant HT du bloc, dont la dernière est le total agrégé."""
    best=[]
    for row in range(ribbon_row+1,min(ws.max_row,ribbon_row+15)+1):
        cols=[col for col in range(start,end+1) if norm(ws.cell(row,col).value)=='montant ht']
        if len(cols)>len(best):best=cols
    return best


def _v10_is_aggregate_cell(ws,_row,col,start,end,ribbon_row):
    amount_cols=_v10_amount_header_columns(ws,start,end,ribbon_row)
    # Au moins 2 montants de bâtiments + 1 total final.
    return len(amount_cols)>=3 and col==amount_cols[-1]


def _v8_collect_company(ws,block,ribbon_row):
    """V10 : masque les faux calculs de la vraie colonne TOTAL avant la synthèse."""
    _,start,end=block
    hidden=[]
    for row in range(ribbon_row+1,ws.max_row+1):
        amount_cols=_v10_amount_header_columns(ws,start,end,ribbon_row)
        if len(amount_cols)<3:break
        total_col=amount_cols[-1]
        cell=ws.cell(row,total_col)
        if not cell.comment:continue
        lines=cell.comment.text.splitlines()
        calc=[line for line in lines if norm(line).startswith('calcul incoherent')]
        if not calc:continue
        hidden.append((cell,cell.comment))
        kept=[line for line in lines if not norm(line).startswith('calcul incoherent')]
        cell.comment=Comment('\n'.join(kept),cell.comment.author or 'AnalyseAO') if kept else None
    try:
        return _v10_previous_collect_company(ws,block,ribbon_row)
    finally:
        for cell,comment in hidden:cell.comment=comment


def _v10_clean_total_column(ws):
    """Supprime commentaires et rouge uniquement dans la colonne TOTAL agrégée."""
    ribbon_row,blocks=_v8_ribbon_blocks(ws)
    if ribbon_row is None:return
    for _,start,end in blocks[1:]:
        amount_cols=_v10_amount_header_columns(ws,start,end,ribbon_row)
        if len(amount_cols)<3:continue
        total_col=amount_cols[-1]
        for row in range(ribbon_row+1,ws.max_row+1):
            cell=ws.cell(row,total_col)
            if not cell.comment:continue
            lines=cell.comment.text.splitlines()
            kept=[line for line in lines if not norm(line).startswith('calcul incoherent')]
            if len(kept)==len(lines):continue
            cell.comment=Comment('\n'.join(kept),cell.comment.author or 'AnalyseAO') if kept else None
            if str(getattr(cell.fill.fgColor,'rgb','')).upper().endswith('F4CCCC'):
                cell.fill=PatternFill(fill_type=None)


def _v10_single_line_ribbons(ws):
    """Nom et prix sur une seule ligne, séparés par un tiret long."""
    ribbon_row,blocks=_v8_ribbon_blocks(ws)
    if ribbon_row is None:return
    for index,(label,start,_end) in enumerate(blocks):
        if index==0:continue
        parts=[part.strip() for part in str(label or '').splitlines() if part.strip()]
        if len(parts)>=2:
            value=f"{parts[0]} — {' — '.join(parts[1:])}"
        else:
            value=parts[0] if parts else ''
        cell=ws.cell(ribbon_row,start)
        cell.value=value
        cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=False,shrink_to_fit=True)
    ws.row_dimensions[ribbon_row].height=22


def analyse_sheet(ws):
    result=_v10_previous_analyse_sheet(ws)
    if result.get('status')!='IGNORED':
        _v10_clean_total_column(ws)
        _v10_single_line_ribbons(ws)
        result['ribbon_single_line_and_total_v10']=True
    return result


# TRUE_BUILDING_CALCULATIONS_V11
_v11_previous_analyse_sheet = analyse_sheet


def _v11_ribbon_row_blocks(ws):
    ribbon_row=None
    for row in range(1,min(ws.max_row,20)+1):
        if any(norm(cell.value)=='estimation' for cell in ws[row]):
            ribbon_row=row;break
    if ribbon_row is None:return None,[]
    blocks=[]
    for merged in ws.merged_cells.ranges:
        if merged.min_row==ribbon_row and merged.max_row==ribbon_row:
            label=str(ws.cell(ribbon_row,merged.min_col).value or '').strip()
            if label:blocks.append((label,merged.min_col,merged.max_col))
    return ribbon_row,sorted(blocks,key=lambda item:item[1])


def _v11_header_mapping(ws,start,end,ribbon_row):
    """Repère chaque triplet Qté entreprise / PU / Montant HT."""
    best=None
    for row in range(ribbon_row+1,min(ws.max_row,ribbon_row+15)+1):
        amounts=[col for col in range(start,end+1) if norm(ws.cell(row,col).value)=='montant ht']
        if not amounts:continue
        current=[]
        for amount_col in amounts:
            pu_col=next((col for col in range(amount_col-1,start-1,-1) if norm(ws.cell(row,col).value) in {'p u','pu'}),None)
            qty_col=next((col for col in range((pu_col or amount_col)-1,start-1,-1) if norm(ws.cell(row,col).value) in {'qte ent','quantite entreprise','quantite'}),None)
            current.append({'amount':amount_col,'pu':pu_col,'qty':qty_col})
        if best is None or len(current)>len(best[1]):best=(row,current)
    return best or (None,[])


def _v11_remove_old_summaries(ws):
    """Supprime uniquement la synthèse V8 afin de la reconstruire après contrôle."""
    ribbon_row,blocks=_v11_ribbon_row_blocks(ws)
    if ribbon_row is None:return
    starts=[]
    for row in range(ribbon_row+1,ws.max_row+1):
        for _,start,_ in blocks[1:]:
            if str(ws.cell(row,start).value or '').startswith('ANALYSE DE L’OFFRE'):
                starts.append(row)
    if starts:
        first=min(starts)
        for merged in list(ws.merged_cells.ranges):
            if merged.min_row>=first:ws.unmerge_cells(str(merged))
        ws.delete_rows(first,ws.max_row-first+1)


def _v11_strip_calc_comments(ws):
    for row in ws.iter_rows():
        for cell in row:
            if not cell.comment:continue
            lines=cell.comment.text.splitlines()
            kept=[line for line in lines if not norm(line).startswith('calcul incoherent') and not norm(line).startswith('total incoherent')]
            if len(kept)==len(lines):continue
            cell.comment=Comment('\n'.join(kept),cell.comment.author or 'AnalyseAO') if kept else None
            if str(getattr(cell.fill.fgColor,'rgb','')).upper().endswith('F4CCCC'):
                cell.fill=PatternFill(fill_type=None)


def _v11_add_comment(cell,message):
    if cell.comment:
        lines=cell.comment.text.splitlines()
        if message not in lines:cell.comment=Comment(cell.comment.text+'\n'+message,cell.comment.author or 'AnalyseAO')
    else:cell.comment=Comment(message,'AnalyseAO')
    cell.fill=PatternFill('solid',fgColor='F4CCCC')


def _v11_num(value):
    return float(value) if isinstance(value,(int,float)) and not isinstance(value,bool) else None


def _v11_validate_block(ws,start,end,ribbon_row,last_table_row):
    """Contrôle chaque bâtiment puis la somme finale, sans confondre les deux."""
    _,mapping=_v11_header_mapping(ws,start,end,ribbon_row)
    if not mapping:return 0
    aggregate=mapping[-1] if len(mapping)>=2 and mapping[-1]['pu'] is None else None
    components=mapping[:-1] if aggregate else mapping
    errors=0
    for row in range(ribbon_row+1,last_table_row+1):
        # Calcul de chaque bâtiment : Qté entreprise x PU = montant du bâtiment.
        for item in components:
            if not item['qty'] or not item['pu']:continue
            qty=_v11_num(ws.cell(row,item['qty']).value); pu=_v11_num(ws.cell(row,item['pu']).value); amount=_v11_num(ws.cell(row,item['amount']).value)
            if qty is None or pu is None or amount is None:continue
            expected=qty*pu; tolerance=max(0.02,abs(expected)*0.0001)
            if abs(expected-amount)>tolerance:
                message=(f"Calcul incohérent : {qty:g} x {pu:g} = {expected:.2f}, "
                         f"montant annoncé = {amount:.2f}")
                _v11_add_comment(ws.cell(row,item['amount']),message);errors+=1
        # Total agrégé : somme des montants des bâtiments = total final.
        if aggregate:
            values=[_v11_num(ws.cell(row,item['amount']).value) for item in components]
            total=_v11_num(ws.cell(row,aggregate['amount']).value)
            numeric=[value for value in values if value is not None]
            if total is not None and len(numeric)>=2:
                expected=sum(numeric); tolerance=max(0.02,abs(expected)*0.0001)
                if abs(expected-total)>tolerance:
                    message=(f"Total incohérent : somme des bâtiments = {expected:.2f}, "
                             f"total annoncé = {total:.2f}")
                    _v11_add_comment(ws.cell(row,aggregate['amount']),message);errors+=1
    return errors


def _v11_offer_total(ws,start,end,ribbon_row,last_table_row):
    """Retient le dernier montant des lignes TOTAL du lot."""
    candidates=[]
    for row in range(ribbon_row+1,last_table_row+1):
        texts=[str(ws.cell(row,col).value or '').strip() for col in range(start,end+1)]
        joined=' '.join(texts).casefold()
        priority=0
        if 'total' in joined and ('lot' in joined or 'gros œuvre' in joined or 'gros oeuvre' in joined or 'etancheite' in joined):priority=5
        elif 'montant ht du lot' in joined:priority=5
        elif joined.startswith('total '):priority=4
        elif 'montant total' in joined:priority=3
        if not priority:continue
        nums=[float(ws.cell(row,col).value) for col in range(start,end+1) if isinstance(ws.cell(row,col).value,(int,float)) and not isinstance(ws.cell(row,col).value,bool)]
        if nums:candidates.append((priority,row,nums[-1]))
    if not candidates:return None
    return sorted(candidates,key=lambda item:(item[0],item[1]),reverse=True)[0][2]


def _v11_money(value):
    return f"{value:,.2f} €".replace(',','X').replace('.',',').replace('X',' ')


def _v11_repair_ribbons(ws,ribbon_row,blocks,last_table_row):
    for index,(label,start,end) in enumerate(blocks):
        if index==0:continue
        name=str(label).split(' — ')[0].splitlines()[0].strip()
        total=_v11_offer_total(ws,start,end,ribbon_row,last_table_row)
        if total is None:
            old=str(label).replace('\n',' — ')
            value=old
        else:value=f"{name} — {_v11_money(total)}"
        cell=ws.cell(ribbon_row,start);cell.value=value
        cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=False,shrink_to_fit=True)


def analyse_sheet(ws):
    result=_v11_previous_analyse_sheet(ws)
    if result.get('status')=='IGNORED':return result
    _v11_remove_old_summaries(ws)
    _v11_strip_calc_comments(ws)
    ribbon_row,blocks=_v11_ribbon_row_blocks(ws)
    if ribbon_row is None:return result
    last_table_row=ws.max_row
    error_count=0
    for _,start,end in blocks[1:]:error_count+=_v11_validate_block(ws,start,end,ribbon_row,last_table_row)
    _v11_repair_ribbons(ws,ribbon_row,blocks,last_table_row)
    # Reconstruit les synthèses APRÈS l'ajout des vrais commentaires.
    _v8_write_company_summaries(ws)
    result['true_building_calculations_v11']=error_count
    return result


# CONSOLIDATED_FINAL_PIPELINE_V12
# Dernier wrapper du module : laisse les traitements historiques s'exécuter,
# puis supprime leurs erreurs arithmétiques/synthèses et reconstruit le résultat.
_v12_previous_analyse_sheet = analyse_sheet


def _v12_norm_ribbon_name(label):
    text=str(label or '').replace('\r','\n')
    first=next((part.strip() for part in text.splitlines() if part.strip()),'')
    # Supprime tout montant déjà suffixé afin d'éviter NOM — PRIX — PRIX.
    first=re.sub(r'\s+[—-]\s+[^—\n]*?€(?:\s*HT)?\s*$', '', first).strip()
    return first


def _v12_ribbon_layout(ws):
    ribbon=None
    for row in range(1,min(ws.max_row,20)+1):
        if any(norm(c.value)=='estimation' for c in ws[row]): ribbon=row; break
    if ribbon is None:return None,[]
    blocks=[]
    for rg in ws.merged_cells.ranges:
        if rg.min_row==ribbon and rg.max_row==ribbon:
            label=str(ws.cell(ribbon,rg.min_col).value or '').strip()
            if label:blocks.append((label,rg.min_col,rg.max_col))
    return ribbon,sorted(blocks,key=lambda x:x[1])


def _v12_remove_summaries(ws,blocks,ribbon):
    starts=[]
    for row in range(ribbon+1,ws.max_row+1):
        for _,start,_ in blocks[1:]:
            if str(ws.cell(row,start).value or '').startswith('ANALYSE DE L’OFFRE'):
                starts.append(row)
    if not starts:return
    first=min(starts)
    for rg in list(ws.merged_cells.ranges):
        if rg.min_row>=first:ws.unmerge_cells(str(rg))
    ws.delete_rows(first,ws.max_row-first+1)


def _v12_clear_arithmetic(ws):
    for row in ws.iter_rows():
        for cell in row:
            if not cell.comment:continue
            lines=cell.comment.text.splitlines()
            kept=[x for x in lines if not norm(x).startswith('calcul incoherent') and not norm(x).startswith('total incoherent')]
            if len(kept)==len(lines):continue
            cell.comment=Comment('\n'.join(kept),cell.comment.author or 'AnalyseAO') if kept else None
            if str(getattr(cell.fill.fgColor,'rgb','')).upper().endswith('F4CCCC'):
                cell.fill=PatternFill(fill_type=None)


def _v12_header_groups(ws,start,end,ribbon):
    """Triplets strictement adjacents Qté ent. / P.U. / Montant HT + total final."""
    best=[]
    for row in range(ribbon+1,min(ws.max_row,ribbon+15)+1):
        groups=[]
        amount_cols=[c for c in range(start,end+1) if norm(ws.cell(row,c).value)=='montant ht']
        for amount in amount_cols:
            pu=amount-1 if amount-1>=start and norm(ws.cell(row,amount-1).value) in {'p u','pu','prix unitaire'} else None
            qty=amount-2 if pu and amount-2>=start and norm(ws.cell(row,amount-2).value) in {'qte ent','quantite entreprise','quantite'} else None
            groups.append({'amount':amount,'pu':pu,'qty':qty})
        if len(groups)>len(best):best=groups
    components=[g for g in best if g['qty'] and g['pu']]
    aggregate=next((g for g in reversed(best) if not g['qty'] and not g['pu']),None)
    return components,aggregate


def _v12_number(value):
    return float(value) if isinstance(value,(int,float)) and not isinstance(value,bool) else None


def _v12_add_error(cell,message):
    current=cell.comment.text.splitlines() if cell.comment else []
    if message not in current:
        cell.comment=Comment(('\n'.join(current)+'\n' if current else '')+message,cell.comment.author if cell.comment else 'AnalyseAO')
    cell.fill=PatternFill('solid',fgColor='F4CCCC')


def _v12_validate_block(ws,start,end,ribbon,last_row):
    components,aggregate=_v12_header_groups(ws,start,end,ribbon)
    errors=0
    for row in range(ribbon+1,last_row+1):
        component_amounts=[]
        for group in components:
            qty=_v12_number(ws.cell(row,group['qty']).value)
            pu=_v12_number(ws.cell(row,group['pu']).value)
            amount=_v12_number(ws.cell(row,group['amount']).value)
            if amount is not None:component_amounts.append(amount)
            if qty is None or pu is None or amount is None:continue
            expected=qty*pu; tolerance=max(0.02,abs(expected)*0.0001)
            if abs(expected-amount)>tolerance:
                msg=f"Calcul incohérent : {qty:g} x {pu:g} = {expected:.2f}, montant annoncé = {amount:.2f}"
                _v12_add_error(ws.cell(row,group['amount']),msg);errors+=1
        if aggregate:
            total=_v12_number(ws.cell(row,aggregate['amount']).value)
            if total is not None and len(component_amounts)>=2:
                expected=sum(component_amounts); tolerance=max(0.02,abs(expected)*0.0001)
                if abs(expected-total)>tolerance:
                    msg=f"Total incohérent : somme des bâtiments = {expected:.2f}, total annoncé = {total:.2f}"
                    _v12_add_error(ws.cell(row,aggregate['amount']),msg);errors+=1
    return errors


def _v12_offer_total(ws,start,end,ribbon,last_row):
    candidates=[]
    for row in range(ribbon+1,last_row+1):
        values=[str(ws.cell(row,c).value or '').strip() for c in range(start,end+1)]
        joined=' '.join(values).casefold()
        priority=0
        if 'montant ht du lot' in joined or 'montant h t du lot' in joined:priority=6
        elif 'total' in joined and any(x in joined for x in ('lot','gros œuvre','gros oeuvre','etancheite')):priority=5
        elif joined.startswith('total '):priority=4
        elif 'montant total' in joined:priority=3
        if not priority:continue
        nums=[float(ws.cell(row,c).value) for c in range(start,end+1) if isinstance(ws.cell(row,c).value,(int,float)) and not isinstance(ws.cell(row,c).value,bool)]
        if nums:candidates.append((priority,row,nums[-1]))
    return max(candidates,key=lambda x:(x[0],x[1]))[2] if candidates else None


def _v12_money(value):
    return f"{value:,.2f} €".replace(',','X').replace('.',',').replace('X',' ')


def _v12_rebuild_ribbons(ws,ribbon,blocks,last_row):
    for index,(label,start,end) in enumerate(blocks):
        if index==0:continue
        name=_v12_norm_ribbon_name(label)
        total=_v12_offer_total(ws,start,end,ribbon,last_row)
        # Reprend le dernier montant existant si aucune ligne Total n'est trouvée.
        if total is None:
            found=re.findall(r'([0-9][0-9 .\u00a0]*,[0-9]{2})\s*€',str(label))
            if found:
                try:total=float(found[-1].replace('\u00a0','').replace(' ','').replace(',','.'))
                except ValueError:pass
        suffix=_v12_money(total) if total is not None else 'Prix HT non détecté'
        cell=ws.cell(ribbon,start);cell.value=f"{name} — {suffix}"
        cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=False,shrink_to_fit=True)
    ws.row_dimensions[ribbon].height=22


def analyse_sheet(ws):
    result=_v12_previous_analyse_sheet(ws)
    if result.get('status')=='IGNORED':return result
    ribbon,blocks=_v12_ribbon_layout(ws)
    if ribbon is None:return result
    _v12_remove_summaries(ws,blocks,ribbon)
    _v12_clear_arithmetic(ws)
    # Relit les blocs après suppression des synthèses.
    ribbon,blocks=_v12_ribbon_layout(ws);last_row=ws.max_row
    errors=sum(_v12_validate_block(ws,s,e,ribbon,last_row) for _,s,e in blocks[1:])
    _v12_rebuild_ribbons(ws,ribbon,blocks,last_row)
    # La synthèse est la toute dernière écriture.
    _v8_write_company_summaries(ws)
    result['consolidated_v12_errors']=errors
    return result


# POST_SUMMARY_RIBBON_AND_TOTALS_V13
_v13_previous_analyse_sheet=analyse_sheet


def _v13_plain_blocks(ws):
    """Lecture pure des rubans, sans appeler la version V9 qui les modifie."""
    ribbon=None
    for row in range(1,min(ws.max_row,20)+1):
        if any(norm(c.value)=='estimation' for c in ws[row]):ribbon=row;break
    if ribbon is None:return None,[]
    blocks=[]
    for rg in ws.merged_cells.ranges:
        if rg.min_row==ribbon and rg.max_row==ribbon:
            label=str(ws.cell(ribbon,rg.min_col).value or '').strip()
            if label:blocks.append((label,rg.min_col,rg.max_col))
    return ribbon,sorted(blocks,key=lambda x:x[1])


def _v13_final_ribbons(ws):
    ribbon,blocks=_v13_plain_blocks(ws)
    if ribbon is None:return
    # Stop avant les synthèses.
    last=ws.max_row
    for row in range(ribbon+1,ws.max_row+1):
        if any(str(ws.cell(row,s).value or '').startswith('ANALYSE DE L’OFFRE') for _,s,_ in blocks[1:]):
            last=row-1;break
    for index,(label,start,end) in enumerate(blocks):
        if index==0:continue
        name=_v12_norm_ribbon_name(label)
        total=_v12_offer_total(ws,start,end,ribbon,last)
        if total is None:
            found=re.findall(r'([0-9][0-9 .\u00a0]*,[0-9]{2})\s*€',str(label))
            if found:
                try:total=float(found[0].replace('\u00a0','').replace(' ','').replace(',','.'))
                except ValueError:pass
        suffix=_v12_money(total) if total is not None else 'Prix HT non détecté'
        cell=ws.cell(ribbon,start);cell.value=f"{name} — {suffix}"
        cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=False,shrink_to_fit=True)
    ws.row_dimensions[ribbon].height=22


def _v13_clean_false_totals(ws):
    """Supprime les totaux V12 calculés sur une cartographie incomplète."""
    ribbon,blocks=_v13_plain_blocks(ws)
    if ribbon is None:return
    for _,start,end in blocks[1:]:
        components,aggregate=_v12_header_groups(ws,start,end,ribbon)
        # Un total n'est contrôlable que si chaque montant de bâtiment attendu
        # est représenté par un vrai triplet Qté ent./PU/Montant.
        amount_headers=_v10_amount_header_columns(ws,start,end,ribbon)
        complete=aggregate is not None and len(components)==len(amount_headers)-1
        if complete:continue
        for row in range(ribbon+1,ws.max_row+1):
            for col in range(start,end+1):
                cell=ws.cell(row,col)
                if not cell.comment:continue
                lines=cell.comment.text.splitlines()
                kept=[x for x in lines if not norm(x).startswith('total incoherent')]
                if len(kept)==len(lines):continue
                cell.comment=Comment('\n'.join(kept),cell.comment.author or 'AnalyseAO') if kept else None
                if str(getattr(cell.fill.fgColor,'rgb','')).upper().endswith('F4CCCC'):
                    cell.fill=PatternFill(fill_type=None)


def analyse_sheet(ws):
    result=_v13_previous_analyse_sheet(ws)
    if result.get('status')!='IGNORED':
        _v13_clean_false_totals(ws)
        # Appelé après la synthèse V8/V12 : aucune fonction ne peut réécrire ensuite.
        _v13_final_ribbons(ws)
        result['post_summary_v13']=True
    return result


# FINAL_FALSE_TOTAL_CLEANUP_V14
_v14_previous_analyse_sheet=analyse_sheet


def _v14_remove_all_total_alerts(ws):
    """Les tableaux sources ont des en-têtes irréguliers : ne valide pas leurs totaux agrégés."""
    for row in ws.iter_rows():
        for cell in row:
            if not cell.comment:continue
            lines=cell.comment.text.splitlines()
            kept=[line for line in lines if not norm(line).startswith('total incoherent')]
            if len(kept)==len(lines):continue
            cell.comment=Comment('\n'.join(kept),cell.comment.author or 'AnalyseAO') if kept else None
            if str(getattr(cell.fill.fgColor,'rgb','')).upper().endswith('F4CCCC'):
                cell.fill=PatternFill(fill_type=None)


def analyse_sheet(ws):
    result=_v14_previous_analyse_sheet(ws)
    if result.get('status')=='IGNORED':return result
    ribbon,blocks=_v13_plain_blocks(ws)
    if ribbon is None:return result
    # Retire la synthèse construite avant le nettoyage.
    _v12_remove_summaries(ws,blocks,ribbon)
    _v14_remove_all_total_alerts(ws)
    # Conserve les vrais 'Calcul incohérent' sur les montants élémentaires.
    _v8_write_company_summaries(ws)
    # Dernière écriture : rubans sur une ligne, sans prix dupliqué.
    _v13_final_ribbons(ws)
    result['false_totals_removed_v14']=True
    return result


# LAYOUT_ENHANCEMENTS_V15
_v15_previous_analyse_sheet = analyse_sheet


def _v15_header_row(ws, ribbon_row):
    """Repère la ligne d'en-têtes qui contient Désignation et les colonnes P.U."""
    best = None
    for row in range(ribbon_row + 1, min(ws.max_row, ribbon_row + 20) + 1):
        values = [norm(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        score = values.count("designation") * 10 + values.count("p u") + values.count("montant ht")
        if best is None or score > best[0]:
            best = (score, row)
    return best[1] if best and best[0] >= 10 else None


def _v15_table_end(ws, ribbon_row):
    """Dernière ligne métier avant les synthèses d'entreprises."""
    _, blocks = _v13_plain_blocks(ws)
    for row in range(ribbon_row + 1, ws.max_row + 1):
        if any(str(ws.cell(row, start).value or "").startswith("ANALYSE DE L’OFFRE") for _, start, _ in blocks[1:]):
            return row - 1
    return ws.max_row


def _v15_shift_right(ws, insert_col):
    """Insère une colonne en conservant fusions, largeurs et formules relatives."""

    old_max_col = ws.max_column
    old_max_row = ws.max_row
    merges = [
        (rg.min_row, rg.max_row, rg.min_col, rg.max_col)
        for rg in list(ws.merged_cells.ranges)
    ]
    for rg in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rg))

    dimensions = {}
    for col in range(insert_col, old_max_col + 1):
        dim = ws.column_dimensions[get_column_letter(col)]
        dimensions[col] = (dim.width, dim.hidden, dim.bestFit, dim.outlineLevel, dim.collapsed)

    ws.move_range(
        f"{get_column_letter(insert_col)}1:{get_column_letter(old_max_col)}{old_max_row}",
        rows=0,
        cols=1,
        translate=True,
    )

    for min_row, max_row, min_col, max_col in merges:
        if min_col >= insert_col:
            min_col += 1
            max_col += 1
        elif max_col >= insert_col:
            max_col += 1
        ws.merge_cells(
            start_row=min_row,
            end_row=max_row,
            start_column=min_col,
            end_column=max_col,
        )

    for old_col, values in dimensions.items():
        target = ws.column_dimensions[get_column_letter(old_col + 1)]
        target.width, target.hidden, target.bestFit, target.outlineLevel, target.collapsed = values


def _v15_average_pu_column(ws):
    """Ajoute une moyenne Excel des P.U. proposés par les entreprises."""

    ribbon_row, blocks = _v13_plain_blocks(ws)
    if ribbon_row is None or len(blocks) < 2:
        return None
    header_row = _v15_header_row(ws, ribbon_row)
    if header_row is None:
        return None

    estimation_start, estimation_end = blocks[0][1], blocks[0][2]
    designation_col = next(
        (col for col in range(estimation_start, estimation_end + 1)
         if norm(ws.cell(header_row, col).value) == "designation"),
        None,
    )
    if designation_col is None:
        return None

    insert_col = designation_col + 1
    _v15_shift_right(ws, insert_col)

    # Recalcule les blocs et les P.U. après décalage.
    ribbon_row, blocks = _v13_plain_blocks(ws)
    header_row = _v15_header_row(ws, ribbon_row)
    table_end = _v15_table_end(ws, ribbon_row)
    enterprise_pu_cols = []
    for _, start, end in blocks[1:]:
        enterprise_pu_cols.extend(
            col for col in range(start, end + 1)
            if norm(ws.cell(header_row, col).value) == "p u"
        )

    header = ws.cell(header_row, insert_col, "Moyenne P.U. entreprises")
    header.fill = PatternFill("solid", fgColor="D9EAF7")
    header.font = Font(bold=True, color="1F1F1F")
    header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(insert_col)].width = 18

    for row in range(header_row + 1, table_end + 1):
        cell = ws.cell(row, insert_col)
        # Une ligne TOTAL déjà fusionnée sur tout le bloc absorbe la colonne
        # insérée : la cellule devient une MergedCell en lecture seule.
        if cell.__class__.__name__ == "MergedCell":
            continue
        refs = [f"{get_column_letter(col)}{row}" for col in enterprise_pu_cols]
        cell.value = f'=IFERROR(AVERAGE({",".join(refs)}),"")' if refs else ""
        cell.number_format = '#,##0.00 [$€-fr-FR]'
        cell.alignment = Alignment(horizontal="right", vertical="center")
    return insert_col


def _v15_neutralize_input_fills(ws):
    """Neutralise les couleurs sources, sans effacer les marquages d'analyse."""

    ribbon_row, _ = _v13_plain_blocks(ws)
    if ribbon_row is None:
        return
    header_row = _v15_header_row(ws, ribbon_row)
    if header_row is None:
        return
    table_end = _v15_table_end(ws, ribbon_row)
    analysis_fills = {"C6E0B4", "FFC7CE", "FFF2CC", "FFE699", "F4CCCC"}

    for row in range(header_row + 1, table_end + 1):
        for cell in ws[row]:
            rgb = str(getattr(cell.fill.fgColor, "rgb", "")).upper()[-6:]
            if rgb not in analysis_fills:
                cell.fill = PatternFill(fill_type=None)


def _v15_apply_column_outline(ws, average_col):
    """Plan niveau 1 : articles estimatifs, moyennes, quantités, P.U. et montants."""

    ribbon_row, _ = _v13_plain_blocks(ws)
    if ribbon_row is None:
        return
    header_row = _v15_header_row(ws, ribbon_row)
    if header_row is None:
        return

    visible = {average_col} if average_col else set()
    visible.update(
        col for col in range(1, ws.max_column + 1)
        if norm(ws.cell(header_row, col).value) in {
            "designation", "quantite", "qte ent", "p u", "montant ht"
        }
    )

    # Les colonnes non essentielles sont regroupées au niveau 2. Un clic sur
    # le bouton de plan "1" affiche seulement les colonnes essentielles.
    ws.sheet_properties.outlinePr.summaryRight = True
    for col in range(1, ws.max_column + 1):
        dim = ws.column_dimensions[get_column_letter(col)]
        dim.outlineLevel = 0 if col in visible else 1
        dim.hidden = False
        dim.collapsed = False


def _v15_tricolor_ribbons(ws):
    """Alterne bleu, vert et jaune sur les rubans des entreprises externes."""

    ribbon_row, blocks = _v13_plain_blocks(ws)
    if ribbon_row is None:
        return
    palette = [
        ("4472C4", "FFFFFF"),  # bleu
        ("70AD47", "FFFFFF"),  # vert
        ("FFD966", "1F1F1F"),  # jaune
    ]
    for index, (_, start, end) in enumerate(blocks[1:]):
        fill_color, font_color = palette[index % len(palette)]
        fill = PatternFill("solid", fgColor=fill_color)
        for col in range(start, end + 1):
            cell = ws.cell(ribbon_row, col)
            cell.fill = fill
            cell.font = Font(color=font_color, bold=True)
            cell.alignment = Alignment(
                horizontal="center", vertical="center",
                wrap_text=False, shrink_to_fit=True,
            )
    ws.row_dimensions[ribbon_row].height = 22


def analyse_sheet(ws):
    result = _v15_previous_analyse_sheet(ws)
    if result.get("status") == "IGNORED":
        return result

    average_col = _v15_average_pu_column(ws)
    _v15_neutralize_input_fills(ws)
    _v15_apply_column_outline(ws, average_col)
    _v15_tricolor_ribbons(ws)
    result["layout_enhancements_v15"] = True
    return result


# REFINED_COLUMN_GROUPS_V16
_v16_previous_analyse_sheet = analyse_sheet


def _v16_refine_column_groups(ws):
    """Groupe uniquement les colonnes redondantes des retours entreprises."""

    ribbon_row, blocks = _v13_plain_blocks(ws)
    if ribbon_row is None or len(blocks) < 2:
        return
    header_row = _v15_header_row(ws, ribbon_row)
    if header_row is None:
        return

    # Réinitialise intégralement le plan trop large créé par V15.
    for col in range(1, ws.max_column + 1):
        dim = ws.column_dimensions[get_column_letter(col)]
        dim.outlineLevel = 0
        dim.hidden = False
        dim.collapsed = False

    # Toute la zone ESTIMATION reste visible, y compris la colonne A qui porte
    # la légende et les codes d'articles.
    estimation_end = blocks[0][2]
    for col in range(1, estimation_end + 1):
        ws.column_dimensions[get_column_letter(col)].outlineLevel = 0

    hide_headers = {
        "rep", "repere", "reference", "code",
        "designation", "u", "unite", "quantite",
    }
    keep_headers = {
        "qte ent", "quantite entreprise", "p u", "pu",
        "prix unitaire", "montant ht",
    }

    for _, start, end in blocks[1:]:
        # La première colonne d'un bloc est généralement le code/repère, même
        # lorsque son en-tête est vide dans certains fichiers entreprises.
        first_col = start
        for col in range(start, end + 1):
            header = norm(ws.cell(header_row, col).value)
            should_group = (
                col == first_col
                or header in hide_headers
            ) and header not in keep_headers

            dim = ws.column_dimensions[get_column_letter(col)]
            dim.outlineLevel = 1 if should_group else 0
            dim.hidden = False
            dim.collapsed = False

    # Verrou absolu : la colonne A n'est jamais masquée ou groupée.
    ws.column_dimensions["A"].outlineLevel = 0
    ws.column_dimensions["A"].hidden = False
    ws.column_dimensions["A"].collapsed = False
    ws.sheet_properties.outlinePr.summaryRight = True


def analyse_sheet(ws):
    result = _v16_previous_analyse_sheet(ws)
    if result.get("status") != "IGNORED":
        _v16_refine_column_groups(ws)
        result["refined_column_groups_v16"] = True
    return result


# EMPTY_NUMERIC_GROUPS_V17
_v17_previous_analyse_sheet = analyse_sheet


def _v17_is_empty_numeric_column(ws, col, header_row, table_end):
    """Vrai si une colonne chiffrée entreprise ne contient aucune donnée utile."""
    meaningful = False
    for row in range(header_row + 1, table_end + 1):
        value = ws.cell(row, col).value
        if value is None or value == "":
            continue
        if isinstance(value, str):
            text = value.strip().casefold()
            # Ces marqueurs indiquent explicitement une absence de donnée chiffrée.
            if text in {"so", "pm", "mo", "non renseigne", "non renseigné", "nr", "-"}:
                continue
            # Une formule est une donnée, même si son résultat visuel est vide.
            if text.startswith("="):
                meaningful = True
                break
            # Tout autre texte dans une colonne chiffrée est conservé accessible.
            meaningful = True
            break
        # Les zéros explicites sont des données renseignées et restent visibles.
        meaningful = True
        break
    return not meaningful


def _v17_refine_groups_with_empty_numeric(ws):
    """Étend V16 aux colonnes chiffrées entièrement non renseignées."""

    ribbon_row, blocks = _v13_plain_blocks(ws)
    if ribbon_row is None or len(blocks) < 2:
        return
    header_row = _v15_header_row(ws, ribbon_row)
    if header_row is None:
        return
    table_end = _v15_table_end(ws, ribbon_row)

    # Repart du plan V16, qui garantit la colonne A et l'estimation visibles.
    _v16_refine_column_groups(ws)

    numeric_headers = {
        "qte ent", "quantite entreprise", "p u", "pu",
        "prix unitaire", "montant ht",
    }
    estimation_end = blocks[0][2]

    for _, start, end in blocks[1:]:
        for col in range(start, end + 1):
            header = norm(ws.cell(header_row, col).value)
            if header not in numeric_headers:
                continue
            if _v17_is_empty_numeric_column(ws, col, header_row, table_end):
                dim = ws.column_dimensions[get_column_letter(col)]
                dim.outlineLevel = 1
                dim.hidden = False
                dim.collapsed = False

    # Verrous absolus demandés.
    for col in range(1, estimation_end + 1):
        dim = ws.column_dimensions[get_column_letter(col)]
        dim.outlineLevel = 0
        dim.hidden = False
        dim.collapsed = False
    ws.column_dimensions["A"].outlineLevel = 0
    ws.column_dimensions["A"].hidden = False
    ws.column_dimensions["A"].collapsed = False


def analyse_sheet(ws):
    result = _v17_previous_analyse_sheet(ws)
    if result.get("status") != "IGNORED":
        _v17_refine_groups_with_empty_numeric(ws)
        result["empty_numeric_groups_v17"] = True
    return result


# UNTITLED_ESTIMATION_GROUPS_V18
_v18_previous_analyse_sheet = analyse_sheet


def _v18_group_untitled_estimation_columns(ws):
    """Groupe les colonnes d'estimation sans titre, sauf la colonne A."""

    ribbon_row, blocks = _v13_plain_blocks(ws)
    if ribbon_row is None or not blocks:
        return
    header_row = _v15_header_row(ws, ribbon_row)
    if header_row is None:
        return

    estimation_start, estimation_end = blocks[0][1], blocks[0][2]

    # V17 a déjà construit les groupes entreprises. On ne réinitialise rien.
    # On ajoute seulement les colonnes sans intitulé de la zone estimation.
    for col in range(estimation_start, estimation_end + 1):
        if col == 1:
            continue
        header = ws.cell(header_row, col).value
        if header is None or str(header).strip() == "":
            dim = ws.column_dimensions[get_column_letter(col)]
            dim.outlineLevel = 1
            dim.hidden = False
            dim.collapsed = False

    # Verrou absolu demandé : légende et codes de la première colonne visibles.
    ws.column_dimensions["A"].outlineLevel = 0
    ws.column_dimensions["A"].hidden = False
    ws.column_dimensions["A"].collapsed = False
    ws.sheet_properties.outlinePr.summaryRight = True


def analyse_sheet(ws):
    result = _v18_previous_analyse_sheet(ws)
    if result.get("status") != "IGNORED":
        _v18_group_untitled_estimation_columns(ws)
        result["untitled_estimation_groups_v18"] = True
    return result


# CONSISTENT_BLOCKS_AND_GROUPS_V19
_v19_previous_analyse_sheet=analyse_sheet


# ROLE_BASED_COLUMN_GROUPS_FIX_V1
def _v19_labeled_subgroups(ws,ribbon,start,end):
    """Repère les libellés fusionnés certains (ex: "Bâtiment D", "Cumul des classeurs").

    Une limite n'est retenue que si elle provient d'une fusion explicite sur une seule
    ligne d'en-tête : aucune heuristique de position n'intervient, le repérage est donc
    certain et ne force aucune structure sur les feuilles qui n'ont pas ce libellé.
    """
    found=[]
    limit=min(ws.max_row,ribbon+6)
    for row in range(ribbon+1,limit+1):
        for rg in ws.merged_cells.ranges:
            if rg.min_row!=row or rg.max_row!=row:continue
            if rg.min_col<start or rg.max_col>end:continue
            text=str(ws.cell(row,rg.min_col).value or '').strip()
            if not text:continue
            label=norm(text)
            is_total=('cumul' in label) or label.startswith('total')
            is_building=(not is_total) and ('batiment' in label or label.startswith('bat '))
            if is_total or is_building:
                found.append({'text':text,'start':rg.min_col,'end':rg.max_col,'is_total':is_total})
    found.sort(key=lambda item:item['start'])
    return found


def _v19_mark_group_boundary(ws,header_row,start,end,comment_text=None):
    """Bordure épaisse et commentaire de repli au début d'un groupe replié contigu."""
    thick=Side(style='medium',color='808080')
    left=ws.cell(header_row,start)
    right=ws.cell(header_row,end)
    left.border=Border(left=thick,top=left.border.top,bottom=left.border.bottom,right=left.border.right)
    right.border=Border(right=thick,top=right.border.top,bottom=right.border.bottom,left=right.border.left)
    if comment_text and not left.comment:
        left.comment=Comment(comment_text,'AnalyseAO')


def _v19_group_consistently(ws):
    try:
        from direct_anomaly_engine import role_map as _dae_role_map, header_limit as _dae_header_limit
    except ImportError:
        return

    ribbon,blocks=_v13_plain_blocks(ws)
    if ribbon is None or len(blocks)<2:return
    header=_v15_header_row(ws,ribbon)
    if header is None:return

    # Repart d'un plan vierge, puis reconstruit sans dépendre des versions précédentes.
    for col in range(1,ws.max_column+1):
        dim=ws.column_dimensions[get_column_letter(col)]
        dim.outlineLevel=0;dim.hidden=False;dim.collapsed=False

    est_start,est_end=blocks[0][1],blocks[0][2]
    # Estimation (références DCE) : A et tous les en-têtes métier visibles, quelle que
    # soit la ligne exacte où ils apparaissent (en-têtes parfois répartis sur deux
    # lignes, ex: P.U./Quantité/Montant en ligne N, Désignation en ligne N+1). Seules
    # les colonnes réellement sans titre (hors A) sont repliables.
    header_limit_est=_dae_header_limit(ws,est_start,est_end,ribbon)
    for col in range(est_start,est_end+1):
        has_title=any(str(ws.cell(row,col).value or '').strip() for row in range(ribbon+1,header_limit_est+1))
        group=(col!=1 and not has_title)
        ws.column_dimensions[get_column_letter(col)].outlineLevel=1 if group else 0

    # Colonnes toujours masquables (répétitions de référence côté entreprise) et
    # colonnes toujours visibles (quantité entreprise, prix unitaire, montant, y
    # compris le montant total/cumul final), classées par rôle réel plutôt que par
    # correspondance exacte de texte sur une seule ligne (ce qui échouait sur les
    # en-têtes répartis sur plusieurs lignes).
    hide_role_names={'designation','unit','quantity_reference','vat'}
    keep_role_names={'quantity_company','unit_price','amount'}

    for _,start,end in blocks[1:]:
        roles=_dae_role_map(ws,start,end,ribbon,False)
        role_of_col={}
        for role_name,cols in roles.items():
            for col in cols:role_of_col[col]=role_name
        header_limit_c=_dae_header_limit(ws,start,end,ribbon)
        titled=[col for col in range(start,end+1) if any(str(ws.cell(row,col).value or '').strip() for row in range(ribbon+1,header_limit_c+1))]
        business_end=max(titled) if titled else start
        for col in range(start,end+1):
            role_name=role_of_col.get(col)
            if col==start:
                group=True
            elif role_name in keep_role_names:
                group=False
            elif role_name in hide_role_names:
                group=True
            elif col>business_end:
                group=True
            else:
                # Rôle non reconnu : visible par prudence, jamais masqué par défaut.
                group=False
            dim=ws.column_dimensions[get_column_letter(col)]
            dim.outlineLevel=1 if group else 0
            dim.hidden=False;dim.collapsed=False

    ws.column_dimensions['A'].outlineLevel=0
    ws.column_dimensions['A'].hidden=False
    ws.column_dimensions['A'].collapsed=False
    ws.sheet_properties.outlinePr.summaryRight=True

    # Renfort visuel des limites de groupe + repli par défaut, uniquement sur les
    # libellés certains "Bâtiment X" (jamais sur une structure devinée) : ceci ne
    # s'active donc que sur les feuilles qui présentent réellement cette structure
    # (ex: 22204), sans jamais toucher les feuilles qui ne l'ont pas (ex: Moirans,
    # lot isolé), qui gardent le comportement précédent (groupes ouverts).
    for _,start,end in blocks:
        for group in _v19_labeled_subgroups(ws,ribbon,start,end):
            if group['is_total']:
                for col in range(group['start'],group['end']+1):
                    ws.column_dimensions[get_column_letter(col)].outlineLevel=0
                continue
            cols_in_group=list(range(group['start'],group['end']+1))
            if any(ws.column_dimensions[get_column_letter(c)].outlineLevel==0 for c in cols_in_group):
                continue
            comment=f"Groupe repliable : quantité, prix unitaire et montant du {group['text']}."
            _v19_mark_group_boundary(ws,header,group['start'],group['end'],comment)
            for col in cols_in_group:
                ws.column_dimensions[get_column_letter(col)].hidden=True
            after=group['end']+1
            if after<=ws.max_column:
                ws.column_dimensions[get_column_letter(after)].collapsed=True


def analyse_sheet(ws):
    result=_v19_previous_analyse_sheet(ws)
    if result.get('status')!='IGNORED':
        _v19_group_consistently(ws)
        result['consistent_blocks_v19']=True
    return result

# FINAL_SORT_RELIABLE_TOTAL_V20
# Redéfinition volontaire : la première analyse appelle explicit_total_ht via
# le namespace global. Cette version fiable s'applique donc avant le tri et le
# déplacement historique des blocs, sans toucher aux wrappers V7 à V19.
def explicit_total_ht(ws, block):
    """Retourne le total HT final du lot pour le tri initial des entreprises.

    Priorise les lignes explicitement liées au total du lot. Les sous-totaux,
    options, TVA et TTC sont exclus. Si aucun libellé fort n'est trouvé, reprend
    le montant déjà présent dans le ruban, puis l'ancienne logique prudente.
    """
    candidates = []
    start, end = block.start, block.end

    for row in range(2, ws.max_row + 1):
        texts = [str(ws.cell(row, col).value or "").strip() for col in range(start, end + 1)]
        joined = norm(" ".join(texts))
        if not joined:
            continue
        if any(token in joined for token in ("tva", "ttc", "option", "variante")):
            continue

        priority = 0
        if "montant ht du lot" in joined or "montant h t du lot" in joined:
            priority = 100
        elif "total" in joined and "lot" in joined:
            priority = 95
        elif joined.startswith("total ") or joined == "total":
            priority = 80
        elif "montant total" in joined:
            priority = 70
        elif "total ht" in joined or "montant ht" in joined:
            priority = 60

        if not priority:
            continue

        numbers = []
        for col in range(start, end + 1):
            number = to_float(ws.cell(row, col).value)
            if number is not None:
                numbers.append(number)
        if numbers:
            # Dernière valeur numérique : total transversal du bloc multi-bâtiments.
            candidates.append((priority, row, numbers[-1]))

    if candidates:
        candidates.sort(key=lambda item: (item[0], item[1]), reverse=True)
        return candidates[0][2]

    # Le ruban peut déjà porter un prix exploitable dans certains classeurs.
    ribbon = str(getattr(block, "label", "") or "")
    matches = re.findall(r"([0-9][0-9 .\u00a0]*,[0-9]{2})\s*€", ribbon)
    if matches:
        try:
            return float(matches[-1].replace("\u00a0", "").replace(" ", "").replace(",", "."))
        except ValueError:
            pass

    # Dernier recours : comportement historique, limité aux colonnes Montant HT.
    amount_cols = block.roles.get("amount", [])
    fallback = []
    for row in range(2, ws.max_row + 1):
        designation, _ = value(ws, row, block, "designation")
        label = norm(designation)
        if not label or any(token in label for token in ("option", "tva", "ttc")):
            continue
        if "montant ht" not in label and "total ht" not in label:
            continue
        for col in reversed(amount_cols):
            number = to_float(ws.cell(row, col).value)
            if number is not None:
                fallback.append(number)
                break
    return fallback[-1] if fallback else None

# TOTAL_RANKING_AND_LEGEND_V21
# Applique aux lignes de total/sous-total les mêmes couleurs de classement (moins
# cher / plus cher / écart +-30 % vs estimation) qu'aux articles. S'exécute avant le
# moteur direct : toute cellule que le moteur direct marquera ensuite en erreur de
# total (fond rouge) prime naturellement sur ce classement, car il s'applique après.
_v21_previous_analyse_sheet = analyse_sheet


def _v21_total_marker_rows(ws):
    try:
        from direct_anomaly_engine import row_marker as _dae_row_marker
    except ImportError:
        return []
    return [row for row in range(2, ws.max_row + 1) if _dae_row_marker(ws, row) in {"STOT", "TOTHT"}]


def _v21_apply_total_ranking_colors(ws):
    # Réutilise exactement le même repérage de blocs/rôles (detect_blocks/value) que le
    # classement par article ci-dessus, plutôt que la détection de ligne d'en-tête
    # utilisée par les groupes de colonnes (V19) : ce repérage est indépendant de la
    # ligne exacte où apparaît chaque en-tête et fonctionne aussi bien sur une ligne
    # de total/sous-total que sur une ligne d'article.
    blocks = detect_blocks(ws)
    if len(blocks) < 2:
        return
    estimation, enterprises = blocks[0], blocks[1:]
    total_rows = _v21_total_marker_rows(ws)
    if not total_rows:
        return
    for row in total_rows:
        est_amount, _ = value(ws, row, estimation, "amount")
        est_amount_n = to_float(est_amount)
        row_amounts = []
        for block in enterprises:
            amount, amount_col = value(ws, row, block, "amount")
            amount_n = to_float(amount)
            if amount_n is None or not amount_col:
                continue
            cell = ws.cell(row, amount_col)
            existing_fill = str(getattr(cell.fill.fgColor, "rgb", "") or "").upper()[-6:]
            if existing_fill == "F4CCCC":
                # Une erreur de total pourra être appliquée par le moteur direct ensuite
                # (qui s'exécute après ce classement) : ne jamais peindre par-dessus une
                # marque d'erreur déjà présente.
                continue
            if est_amount_n and est_amount_n > 0 and amount_n > 0:
                if amount_n > est_amount_n * (1 + DELTA_THRESHOLD):
                    cell.font = copy(ABOVE_FONT)
                    add_comment(cell, "Montant supérieur de plus de 30 % à l'estimation")
                elif amount_n < est_amount_n * (1 - DELTA_THRESHOLD):
                    cell.font = copy(BELOW_FONT)
                    add_comment(cell, "Montant inférieur de plus de 30 % à l'estimation")
            if amount_n > 0:
                row_amounts.append((amount_n, cell))
        if len(row_amounts) < 2:
            continue
        min_value = min(amount for amount, _ in row_amounts)
        max_value = max(amount for amount, _ in row_amounts)
        if min_value == max_value:
            continue
        for amount_n, cell in row_amounts:
            existing_fill = str(getattr(cell.fill.fgColor, "rgb", "") or "").upper()[-6:]
            if existing_fill == "F4CCCC":
                continue
            if amount_n == min_value:
                cell.fill = copy(CHEAPEST_FILL)
            if amount_n == max_value:
                cell.fill = copy(MOST_EXPENSIVE_FILL)


def analyse_sheet(ws):
    result = _v21_previous_analyse_sheet(ws)
    if result.get("status") != "IGNORED":
        _v21_apply_total_ranking_colors(ws)
        result["total_ranking_colors_v21"] = True
    return result


# BEGIN_DIRECT_ANOMALY_ENGINE_V2
try:
    from direct_anomaly_engine import process as _direct_v2_process
    _direct_v2_previous_analyse_sheet=analyse_sheet
    def analyse_sheet(ws):
        result=_direct_v2_previous_analyse_sheet(ws)
        if not isinstance(result,dict) or result.get("status")=="IGNORED":return result
        try:
            direct=_direct_v2_process(ws);enriched=dict(result);enriched["direct_anomaly_engine"]=direct;enriched["direct_anomaly_engine_enabled"]=True;return enriched
        except Exception as exc:
            legacy=dict(result);legacy["direct_anomaly_engine_enabled"]=False;legacy["direct_anomaly_engine_fallback"]=f"{type(exc).__name__}: {exc}";return legacy
except ImportError:
    pass
# END_DIRECT_ANOMALY_ENGINE_V2

# TOTAL_ERROR_BANNER_V22
# S'exécute après le moteur direct : lit les TOTAL_ERROR qu'il a produits et insère
# un bandeau d'alerte au-dessus du tableau (entre la légende et le ruban) uniquement
# si au moins une erreur de total certaine a été détectée.
_v22_previous_analyse_sheet = analyse_sheet


_GENERAL_TOTAL_CATEGORIES = {"CHAPTER_TOTAL_ERROR", "HT_TOTAL_ERROR", "VAT_TOTAL_ERROR", "TTC_TOTAL_ERROR"}


def _fr_money(value):
    return f"{value:,.2f} €".replace(",", "X").replace(".", ",").replace("X", " ")


def _fr_percent(value):
    return f"{value:,.2f} %".replace(",", "X").replace(".", ",").replace("X", " ")


def _v22_total_errors(result):
    direct = result.get("direct_anomaly_engine") if isinstance(result, dict) else None
    if not isinstance(direct, dict):
        return []
    return [item for item in direct.get("issues", []) if item.get("category") in _GENERAL_TOTAL_CATEGORIES]


def _v22_insert_banner(ws, errors):
    from direct_anomaly_engine import pure_blocks as _dae_pure_blocks

    ribbon_row = _find_ribbon_row(ws)
    if ribbon_row is None:
        return
    # Idempotence : ne pas ré-insérer un bandeau déjà présent (ré-analyse du même classeur).
    for row in range(1, ribbon_row):
        for cell in ws[row]:
            if str(cell.value or "").strip().upper().startswith("⚠ ERREURS DE TOTAUX"):
                return

    _, blocks = _dae_pure_blocks(ws)
    if len(blocks) < 2:
        return
    by_company = {}
    for item in errors:
        by_company.setdefault(str(item.get("company", "")).strip(), []).append(item)

    faulty = [
        (cs, ce, by_company[str(label).splitlines()[0].strip()])
        for label, cs, ce in blocks[1:]
        if str(label).splitlines()[0].strip() in by_company
    ]
    if not faulty:
        return

    height = max(2 + len(items) for _, _, items in faulty)
    # openpyxl.insert_rows() ne déplace pas les fusions existantes : on ne remonte que
    # celles situées à partir de la ligne d'insertion (les fusions de la légende,
    # avant le ruban, restent inchangées).
    merges_to_shift = [
        (rg.min_row, rg.max_row, rg.min_col, rg.max_col)
        for rg in list(ws.merged_cells.ranges) if rg.min_row >= ribbon_row
    ]
    for min_row, max_row, min_col, max_col in merges_to_shift:
        ws.unmerge_cells(start_row=min_row, end_row=max_row, start_column=min_col, end_column=max_col)
    ws.insert_rows(ribbon_row, height)
    for min_row, max_row, min_col, max_col in merges_to_shift:
        ws.merge_cells(
            start_row=min_row + height, end_row=max_row + height,
            start_column=min_col, end_column=max_col,
        )

    for cs, ce, items in faulty:
        title_row = ribbon_row
        ws.merge_cells(start_row=title_row, start_column=cs, end_row=title_row, end_column=ce)
        title = ws.cell(title_row, cs, "⚠ ERREURS DE TOTAUX")
        title.fill = PatternFill("solid", fgColor="C00000")
        title.font = Font(color="FFFFFF", bold=True)
        title.alignment = Alignment(horizontal="left", vertical="center")

        intro_row = ribbon_row + 1
        ws.merge_cells(start_row=intro_row, start_column=cs, end_row=intro_row, end_column=ce)
        intro = ws.cell(
            intro_row, cs,
            f"{len(items)} incohérence(s) de total général détectée(s). Valeurs déclarées conservées "
            "(voir cellules et libellés en rouge, avec commentaire).",
        )
        intro.font = Font(bold=True)
        intro.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        for offset, item in enumerate(items, start=2):
            row = ribbon_row + offset
            ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
            relative_delta = item.get("relative_delta")
            relative_text = _fr_percent(relative_delta * 100) if relative_delta is not None else "n/a"
            line = (
                f"{item.get('label', '')} | attendu {_fr_money(item.get('expected', 0))} | "
                f"déclaré {_fr_money(item.get('declared', 0))} | écart {_fr_money(item.get('delta', 0))} "
                f"({relative_text})"
            )
            cell = ws.cell(row, cs, line)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    if ws.freeze_panes:
        row, col = coordinate_to_tuple(
            ws.freeze_panes.coordinate if hasattr(ws.freeze_panes, "coordinate") else str(ws.freeze_panes)
        )
        ws.freeze_panes = f"{get_column_letter(col)}{row + height}"


def analyse_sheet(ws):
    result = _v22_previous_analyse_sheet(ws)
    if not isinstance(result, dict) or result.get("status") == "IGNORED":
        return result
    errors = _v22_total_errors(result)
    if errors:
        _v22_insert_banner(ws, errors)
        result["total_error_banner_v22"] = len(errors)
    return result
# END_TOTAL_ERROR_BANNER_V22
